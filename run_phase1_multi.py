#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Phase1_Evidence Analysis System - ãƒãƒ«ãƒäº‹ä»¶å¯¾å¿œç‰ˆ

ã€æ¦‚è¦ã€‘
å¤§å…ƒã®å…±æœ‰ãƒ‰ãƒ©ã‚¤ãƒ–IDã®ã¿ã‚’è¨­å®šã—ã€è¤‡æ•°äº‹ä»¶ã‚’ä¸¦è¡Œç®¡ç†ã§ãã‚‹ã‚·ã‚¹ãƒ†ãƒ ã§ã™ã€‚

ã€ä½¿ç”¨æ–¹æ³•ã€‘
    python3 run_phase1_multi.py

ã€æ©Ÿèƒ½ã€‘
    - å…±æœ‰ãƒ‰ãƒ©ã‚¤ãƒ–ã‹ã‚‰äº‹ä»¶ã‚’è‡ªå‹•æ¤œå‡º
    - è¤‡æ•°äº‹ä»¶ã®ä¸¦è¡Œç®¡ç†
    - äº‹ä»¶é¸æŠãƒ»åˆ‡ã‚Šæ›¿ãˆ
    - è¨¼æ‹ åˆ†æã®å®Ÿè¡Œ
"""

import os
import sys
import json
import logging
from datetime import datetime
from typing import List, Optional, Dict

# ãƒ—ãƒ­ã‚¸ã‚§ã‚¯ãƒˆãƒ«ãƒ¼ãƒˆã‚’Pythonãƒ‘ã‚¹ã«è¿½åŠ 
project_root = os.path.dirname(os.path.abspath(__file__))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# è‡ªä½œãƒ¢ã‚¸ãƒ¥ãƒ¼ãƒ«ã®ã‚¤ãƒ³ãƒãƒ¼ãƒˆ
try:
    import global_config as gconfig
    from src.case_manager import CaseManager
    from src.evidence_organizer import EvidenceOrganizer
    from src.gdrive_database_manager import GDriveDatabaseManager, create_database_manager
    # æ—¢å­˜ã®ãƒ¢ã‚¸ãƒ¥ãƒ¼ãƒ«ï¼ˆäº‹ä»¶å›ºæœ‰ã®å‡¦ç†ï¼‰
    from src.metadata_extractor import MetadataExtractor
    from src.file_processor import FileProcessor
    from src.ai_analyzer_complete import AIAnalyzerComplete
    from src.evidence_editor_ai import EvidenceEditorAI
    from src.timeline_builder import TimelineBuilder
except ImportError as e:
    print(f"ã‚¨ãƒ©ãƒ¼: ãƒ¢ã‚¸ãƒ¥ãƒ¼ãƒ«ã®ã‚¤ãƒ³ãƒãƒ¼ãƒˆã«å¤±æ•—ã—ã¾ã—ãŸ: {e}")
    print("\nå¿…è¦ãªãƒ•ã‚¡ã‚¤ãƒ«:")
    print("  - global_config.py")
    print("  - src/case_manager.py")
    print("  - src/evidence_organizer.py")
    print("  - src/metadata_extractor.py")
    print("  - src/file_processor.py")
    print("  - src/ai_analyzer_complete.py")
    print("  - src/evidence_editor_ai.py")
    sys.exit(1)

# ãƒ­ã‚®ãƒ³ã‚°è¨­å®š
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler('phase1_multi.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class Phase1MultiRunner:
    """Phase 1ãƒãƒ«ãƒäº‹ä»¶å¯¾å¿œå®Ÿè¡Œã‚¯ãƒ©ã‚¹"""
    
    def __init__(self):
        """åˆæœŸåŒ–"""
        self.case_manager = CaseManager()
        self.current_case = None
        self.db_manager = None  # äº‹ä»¶é¸æŠå¾Œã«åˆæœŸåŒ–
        self.metadata_extractor = MetadataExtractor()
        self.file_processor = FileProcessor()
        self.ai_analyzer = AIAnalyzerComplete()
        self.evidence_editor = EvidenceEditorAI()
    
    def select_case(self) -> bool:
        """äº‹ä»¶ã‚’é¸æŠã¾ãŸã¯æ–°è¦ä½œæˆ
        
        Returns:
            é¸æŠæˆåŠŸ: True, ã‚­ãƒ£ãƒ³ã‚»ãƒ«: False
        """
        print("\n" + "="*70)
        print("  Phase1_Evidence Analysis System - äº‹ä»¶é¸æŠ")
        print("="*70)
        
        # äº‹ä»¶ã‚’æ¤œå‡º
        cases = self.case_manager.detect_cases()
        
        if not cases:
            print("\näº‹ä»¶ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã§ã—ãŸã€‚")
            print("\nã€é¸æŠã—ã¦ãã ã•ã„ã€‘")
            print("  1. æ–°è¦äº‹ä»¶ã‚’ä½œæˆ")
            print("  2. çµ‚äº†")
            
            choice = input("\né¸æŠã—ã¦ãã ã•ã„ (1-2): ").strip()
            
            if choice == '1':
                # æ–°è¦äº‹ä»¶ä½œæˆ
                return self._create_new_case()
            else:
                print("\nâŒ çµ‚äº†ã—ã¾ã™")
                return False
        
        # äº‹ä»¶ä¸€è¦§ã‚’è¡¨ç¤º
        print("\nã€æ¤œå‡ºã•ã‚ŒãŸäº‹ä»¶ã€‘")
        self.case_manager.display_cases(cases)
        
        print("\nã€é¸æŠã—ã¦ãã ã•ã„ã€‘")
        print(f"  1-{len(cases)}. æ—¢å­˜äº‹ä»¶ã‚’é¸æŠ")
        print(f"  {len(cases)+1}. æ–°è¦äº‹ä»¶ã‚’ä½œæˆ")
        print(f"  0. çµ‚äº†")
        
        # äº‹ä»¶ã‚’é¸æŠ
        selected_case = self.case_manager.select_case_interactive(cases, allow_new=True)
        
        if selected_case == "new":
            # æ–°è¦äº‹ä»¶ä½œæˆ
            return self._create_new_case()
        elif not selected_case:
            print("\nâŒ äº‹ä»¶ãŒé¸æŠã•ã‚Œã¾ã›ã‚“ã§ã—ãŸ")
            return False
        
        self.current_case = selected_case
        
        # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ãƒãƒãƒ¼ã‚¸ãƒ£ãƒ¼ã‚’åˆæœŸåŒ–
        self.db_manager = create_database_manager(self.case_manager, selected_case)
        if not self.db_manager:
            logger.warning(" ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ãƒãƒãƒ¼ã‚¸ãƒ£ãƒ¼ã®åˆæœŸåŒ–ã«å¤±æ•—ã—ã¾ã—ãŸ")
        
        # äº‹ä»¶è¨­å®šãƒ•ã‚¡ã‚¤ãƒ«ã‚’ç”Ÿæˆ
        self.case_manager.generate_case_config(selected_case, "current_case.json")
        
        return True
    
    def _create_new_case(self) -> bool:
        """æ–°è¦äº‹ä»¶ã‚’ä½œæˆ
        
        Returns:
            ä½œæˆæˆåŠŸ: True, ã‚­ãƒ£ãƒ³ã‚»ãƒ«: False
        """
        print("\n" + "="*70)
        print("  æ–°è¦äº‹ä»¶ã®ä½œæˆ")
        print("="*70)
        
        try:
            # äº‹ä»¶æƒ…å ±ã‚’å…¥åŠ›
            print("\näº‹ä»¶æƒ…å ±ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„")
            
            case_id = input("\näº‹ä»¶IDï¼ˆä¾‹: 2025_001ï¼‰: ").strip()
            if not case_id:
                print("ã‚¨ãƒ©ãƒ¼: äº‹ä»¶IDã¯å¿…é ˆã§ã™")
                return False
            
            case_name = input("äº‹ä»¶åï¼ˆä¾‹: æå®³è³ å„Ÿè«‹æ±‚äº‹ä»¶ï¼‰: ").strip()
            if not case_name:
                print("ã‚¨ãƒ©ãƒ¼: äº‹ä»¶åã¯å¿…é ˆã§ã™")
                return False
            
            case_number = input("äº‹ä»¶ç•ªå·ï¼ˆä¾‹: ä»¤å’Œ7å¹´(ãƒ¯)ç¬¬1å·ï¼‰[çœç•¥å¯]: ").strip()
            court = input("è£åˆ¤æ‰€ï¼ˆä¾‹: æ±äº¬åœ°æ–¹è£åˆ¤æ‰€ï¼‰[çœç•¥å¯]: ").strip()
            plaintiff = input("åŸå‘Šï¼ˆä¾‹: å±±ç”°å¤ªéƒï¼‰[çœç•¥å¯]: ").strip()
            defendant = input("è¢«å‘Šï¼ˆä¾‹: æ ªå¼ä¼šç¤¾ã€‡ã€‡ï¼‰[çœç•¥å¯]: ").strip()
            
            # ç¢ºèª
            print("\nå…¥åŠ›å†…å®¹ã®ç¢ºèª:")
            print(f"  äº‹ä»¶ID: {case_id}")
            print(f"  äº‹ä»¶å: {case_name}")
            if case_number:
                print(f"  äº‹ä»¶ç•ªå·: {case_number}")
            if court:
                print(f"  è£åˆ¤æ‰€: {court}")
            if plaintiff:
                print(f"  åŸå‘Š: {plaintiff}")
            if defendant:
                print(f"  è¢«å‘Š: {defendant}")
            
            confirm = input("\nã“ã®å†…å®¹ã§ä½œæˆã—ã¾ã™ã‹ï¼Ÿ (y/n): ").strip().lower()
            if confirm != 'y':
                print("ã‚¨ãƒ©ãƒ¼: ã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
                return False
            
            # ãƒ•ã‚©ãƒ«ãƒ€ä½œæˆ
            print("\nãƒ•ã‚©ãƒ«ãƒ€ã‚’ä½œæˆä¸­...")
            
            service = self.case_manager.get_google_drive_service()
            if not service:
                print("ã‚¨ãƒ©ãƒ¼: Google Driveèªè¨¼ã«å¤±æ•—ã—ã¾ã—ãŸ")
                return False
            
            # äº‹ä»¶ãƒ•ã‚©ãƒ«ãƒ€ã‚’ä½œæˆ
            case_folder_name = f"{case_id}_{case_name}"
            
            folder_metadata = {
                'name': case_folder_name,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [self.case_manager.shared_drive_root_id]
            }
            
            case_folder = service.files().create(
                body=folder_metadata,
                supportsAllDrives=True,
                fields='id, name, webViewLink'
            ).execute()
            
            case_folder_id = case_folder['id']
            print(f"  âœ… äº‹ä»¶ãƒ•ã‚©ãƒ«ãƒ€ä½œæˆ: {case_folder_name}")
            
            # éšå±¤çš„ãƒ•ã‚©ãƒ«ãƒ€æ§‹é€ ã‚’ä½œæˆ
            if gconfig.USE_HIERARCHICAL_FOLDERS:
                print(f"  ğŸ“ éšå±¤çš„ãƒ•ã‚©ãƒ«ãƒ€æ§‹é€ ã‚’ä½œæˆä¸­...")
                
                # ç”²å·è¨¼ãƒ«ãƒ¼ãƒˆãƒ•ã‚©ãƒ«ãƒ€ã‚’ä½œæˆ
                ko_folder_metadata = {
                    'name': 'ç”²å·è¨¼',
                    'mimeType': 'application/vnd.google-apps.folder',
                    'parents': [case_folder_id]
                }
                
                ko_folder = service.files().create(
                    body=ko_folder_metadata,
                    supportsAllDrives=True,
                    fields='id, name'
                ).execute()
                
                ko_folder_id = ko_folder['id']
                print(f"    âœ… ç”²å·è¨¼ãƒ•ã‚©ãƒ«ãƒ€ä½œæˆ")
                
                # ç”²å·è¨¼é…ä¸‹ã«ã‚µãƒ–ãƒ•ã‚©ãƒ«ãƒ€ã‚’ä½œæˆ
                ko_subfolders = {}
                for status_key, folder_name in [('confirmed', 'ç¢ºå®šæ¸ˆã¿'), ('pending', 'æ•´ç†æ¸ˆã¿_æœªç¢ºå®š'), ('unclassified', 'æœªåˆ†é¡')]:
                    subfolder = service.files().create(
                        body={
                            'name': folder_name,
                            'mimeType': 'application/vnd.google-apps.folder',
                            'parents': [ko_folder_id]
                        },
                        supportsAllDrives=True,
                        fields='id, name'
                    ).execute()
                    ko_subfolders[status_key] = subfolder['id']
                    print(f"      âœ… ç”²å·è¨¼/{folder_name}")
                
                # ä¹™å·è¨¼ãƒ«ãƒ¼ãƒˆãƒ•ã‚©ãƒ«ãƒ€ã‚’ä½œæˆ
                otsu_folder_metadata = {
                    'name': 'ä¹™å·è¨¼',
                    'mimeType': 'application/vnd.google-apps.folder',
                    'parents': [case_folder_id]
                }
                
                otsu_folder = service.files().create(
                    body=otsu_folder_metadata,
                    supportsAllDrives=True,
                    fields='id, name'
                ).execute()
                
                otsu_folder_id = otsu_folder['id']
                print(f"    âœ… ä¹™å·è¨¼ãƒ•ã‚©ãƒ«ãƒ€ä½œæˆ")
                
                # ä¹™å·è¨¼é…ä¸‹ã«ã‚µãƒ–ãƒ•ã‚©ãƒ«ãƒ€ã‚’ä½œæˆ
                otsu_subfolders = {}
                for status_key, folder_name in [('confirmed', 'ç¢ºå®šæ¸ˆã¿'), ('pending', 'æ•´ç†æ¸ˆã¿_æœªç¢ºå®š'), ('unclassified', 'æœªåˆ†é¡')]:
                    subfolder = service.files().create(
                        body={
                            'name': folder_name,
                            'mimeType': 'application/vnd.google-apps.folder',
                            'parents': [otsu_folder_id]
                        },
                        supportsAllDrives=True,
                        fields='id, name'
                    ).execute()
                    otsu_subfolders[status_key] = subfolder['id']
                    print(f"      âœ… ä¹™å·è¨¼/{folder_name}")
            else:
                # æ—§å½¢å¼ã®ãƒ•ãƒ©ãƒƒãƒˆãªæ§‹é€ ï¼ˆå¾Œæ–¹äº’æ›æ€§ã®ãŸã‚æ®‹ã™ï¼‰
                ko_folder = service.files().create(
                    body={
                        'name': 'ç”²å·è¨¼',
                        'mimeType': 'application/vnd.google-apps.folder',
                        'parents': [case_folder_id]
                    },
                    supportsAllDrives=True,
                    fields='id, name'
                ).execute()
                
                otsu_folder = service.files().create(
                    body={
                        'name': 'ä¹™å·è¨¼',
                        'mimeType': 'application/vnd.google-apps.folder',
                        'parents': [case_folder_id]
                    },
                    supportsAllDrives=True,
                    fields='id, name'
                ).execute()
                
                # æœªåˆ†é¡ãƒ»æ•´ç†æ¸ˆã¿_æœªç¢ºå®šãƒ•ã‚©ãƒ«ãƒ€ï¼ˆäº‹ä»¶ãƒ•ã‚©ãƒ«ãƒ€ç›´ä¸‹ï¼‰
                service.files().create(
                    body={
                        'name': 'æœªåˆ†é¡',
                        'mimeType': 'application/vnd.google-apps.folder',
                        'parents': [case_folder_id]
                    },
                    supportsAllDrives=True,
                    fields='id, name'
                ).execute()
                
                service.files().create(
                    body={
                        'name': 'æ•´ç†æ¸ˆã¿_æœªç¢ºå®š',
                        'mimeType': 'application/vnd.google-apps.folder',
                        'parents': [case_folder_id]
                    },
                    supportsAllDrives=True,
                    fields='id, name'
                ).execute()
                
                print(f"  âœ… æ—§å½¢å¼ãƒ•ã‚©ãƒ«ãƒ€æ§‹é€ ä½œæˆå®Œäº†")
            
            # äº‹ä»¶æƒ…å ±ã‚’ä¸€æ™‚è¨­å®šï¼ˆdatabaseä½œæˆã®ãŸã‚ï¼‰
            temp_case_info = {
                'case_id': case_id,
                'case_name': case_name,
                'case_folder_id': case_folder_id
            }
            
            # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ãƒãƒãƒ¼ã‚¸ãƒ£ãƒ¼ã‚’åˆæœŸåŒ–ã—ã¦database.jsonã‚’ä½œæˆ
            temp_db_manager = create_database_manager(self.case_manager, temp_case_info)
            if temp_db_manager:
                # ç©ºã®database.jsonãŒGoogle Driveã«ä½œæˆã•ã‚Œã‚‹
                database = temp_db_manager.load_database()
                # case_infoã‚’è¿½åŠ 
                database['case_info'] = {
                    "case_name": case_name,
                    "case_number": case_number or "",
                    "court": court or "",
                    "plaintiff": plaintiff or "",
                    "defendant": defendant or "",
                    "case_summary": ""
                }
                temp_db_manager.save_database(database)
                print(f"  âœ… database.jsonä½œæˆï¼ˆGoogle Driveï¼‰")
            else:
                print(f"  âŒ database.jsonä½œæˆã«å¤±æ•—")
                return False
            
            # äº‹ä»¶æƒ…å ±ã‚’è¨­å®š
            self.current_case = {
                'case_id': case_id,
                'case_name': case_name,
                'case_folder_id': case_folder_id,
                'ko_evidence_folder_id': ko_folder['id'],
                'otsu_evidence_folder_id': otsu_folder['id'],
                'case_folder_url': case_folder.get('webViewLink', '')
            }
            
            # éšå±¤çš„æ§‹é€ ã®å ´åˆã¯ã‚µãƒ–ãƒ•ã‚©ãƒ«ãƒ€æƒ…å ±ã‚’è¿½åŠ 
            if gconfig.USE_HIERARCHICAL_FOLDERS:
                self.current_case['folder_structure'] = 'hierarchical'
                self.current_case['ko_folders'] = ko_subfolders
                self.current_case['otsu_folders'] = otsu_subfolders
            else:
                self.current_case['folder_structure'] = 'legacy'
            
            # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ãƒãƒãƒ¼ã‚¸ãƒ£ãƒ¼ã‚’åˆæœŸåŒ–
            self.db_manager = create_database_manager(self.case_manager, self.current_case)
            if not self.db_manager:
                logger.warning(" ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ãƒãƒãƒ¼ã‚¸ãƒ£ãƒ¼ã®åˆæœŸåŒ–ã«å¤±æ•—ã—ã¾ã—ãŸ")
            
            # äº‹ä»¶è¨­å®šãƒ•ã‚¡ã‚¤ãƒ«ã‚’ç”Ÿæˆ
            self.case_manager.generate_case_config(self.current_case, "current_case.json")
            
            print("\nâœ… æ–°è¦äº‹ä»¶ã‚’ä½œæˆã—ã¾ã—ãŸ")
            print(f" ãƒ•ã‚©ãƒ«ãƒ€URL: {case_folder.get('webViewLink', 'N/A')}")
            
            return True
            
        except Exception as e:
            print(f"\nâŒ ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _upload_database_to_gdrive(self, local_path: str, case_folder_id: str) -> Optional[str]:
        """database.jsonã‚’Google Driveã«ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰
        
        Args:
            local_path: ãƒ­ãƒ¼ã‚«ãƒ«ã®database.jsonãƒ‘ã‚¹
            case_folder_id: äº‹ä»¶ãƒ•ã‚©ãƒ«ãƒ€ID
        
        Returns:
            ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰ã•ã‚ŒãŸãƒ•ã‚¡ã‚¤ãƒ«IDï¼ˆå¤±æ•—æ™‚ã¯Noneï¼‰
        """
        try:
            service = self.case_manager.get_google_drive_service()
            if not service:
                return None
            
            from googleapiclient.http import MediaFileUpload
            
            file_metadata = {
                'name': 'database.json',
                'parents': [case_folder_id],
                'mimeType': 'application/json'
            }
            
            media = MediaFileUpload(local_path, mimetype='application/json', resumable=True)
            
            file = service.files().create(
                body=file_metadata,
                media_body=media,
                supportsAllDrives=True,
                fields='id, name'
            ).execute()
            
            logger.info(f" database.jsonã‚’Google Driveã«ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰: {file['id']}")
            return file['id']
            
        except Exception as e:
            logger.error(f" database.jsonã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰å¤±æ•—: {e}")
            return None
    
    def _download_database_from_gdrive(self, case_folder_id: str) -> Optional[Dict]:
        """Google Driveã‹ã‚‰database.jsonã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰
        
        Args:
            case_folder_id: äº‹ä»¶ãƒ•ã‚©ãƒ«ãƒ€ID
        
        Returns:
            database.jsonã®å†…å®¹ï¼ˆDictï¼‰ã€è¦‹ã¤ã‹ã‚‰ãªã„å ´åˆã¯None
        """
        try:
            service = self.case_manager.get_google_drive_service()
            if not service:
                return None
            
            # database.jsonã‚’æ¤œç´¢
            query = f"name='database.json' and '{case_folder_id}' in parents and trashed=false"
            results = service.files().list(
                q=query,
                corpora='drive',
                driveId=self.case_manager.shared_drive_root_id,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                fields='files(id, name)',
                pageSize=1
            ).execute()
            
            files = results.get('files', [])
            if not files:
                logger.warning(" Google Driveã«database.jsonãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“")
                return None
            
            file_id = files[0]['id']
            
            # ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰
            import io
            from googleapiclient.http import MediaIoBaseDownload
            
            request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            
            done = False
            while not done:
                status, done = downloader.next_chunk()
            
            fh.seek(0)
            database = json.loads(fh.read().decode('utf-8'))
            
            logger.info(f" database.jsonã‚’Google Driveã‹ã‚‰ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰")
            return database
            
        except Exception as e:
            logger.error(f" database.jsonãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰å¤±æ•—: {e}")
            return None
    
    def _update_database_on_gdrive(self, database: Dict, case_folder_id: str) -> bool:
        """Google Driveä¸Šã®database.jsonã‚’æ›´æ–°
        
        Args:
            database: æ›´æ–°ã™ã‚‹database.jsonå†…å®¹
            case_folder_id: äº‹ä»¶ãƒ•ã‚©ãƒ«ãƒ€ID
        
        Returns:
            æˆåŠŸ: Trueã€å¤±æ•—: False
        """
        try:
            service = self.case_manager.get_google_drive_service()
            if not service:
                return False
            
            # æ—¢å­˜ã®database.jsonã‚’æ¤œç´¢
            query = f"name='database.json' and '{case_folder_id}' in parents and trashed=false"
            results = service.files().list(
                q=query,
                corpora='drive',
                driveId=self.case_manager.shared_drive_root_id,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                fields='files(id, name)',
                pageSize=1
            ).execute()
            
            files = results.get('files', [])
            
            # ä¸€æ™‚ãƒ•ã‚¡ã‚¤ãƒ«ã«ä¿å­˜
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as tmp:
                json.dump(database, tmp, ensure_ascii=False, indent=2)
                tmp_path = tmp.name
            
            from googleapiclient.http import MediaFileUpload
            media = MediaFileUpload(tmp_path, mimetype='application/json', resumable=True)
            
            if files:
                # æ—¢å­˜ãƒ•ã‚¡ã‚¤ãƒ«ã‚’æ›´æ–°
                file_id = files[0]['id']
                service.files().update(
                    fileId=file_id,
                    media_body=media,
                    supportsAllDrives=True
                ).execute()
                logger.info(f" Google Driveä¸Šã®database.jsonã‚’æ›´æ–°")
            else:
                # æ–°è¦ä½œæˆ
                file_metadata = {
                    'name': 'database.json',
                    'parents': [case_folder_id],
                    'mimeType': 'application/json'
                }
                service.files().create(
                    body=file_metadata,
                    media_body=media,
                    supportsAllDrives=True
                ).execute()
                logger.info(f" database.jsonã‚’Google Driveã«æ–°è¦ä½œæˆ")
            
            # ä¸€æ™‚ãƒ•ã‚¡ã‚¤ãƒ«ã‚’å‰Šé™¤
            os.remove(tmp_path)
            return True
            
        except Exception as e:
            logger.error(f" database.jsonæ›´æ–°å¤±æ•—: {e}")
            return False
    
    def load_database(self) -> dict:
        """database.jsonã®èª­ã¿è¾¼ã¿ï¼ˆGoogle Driveã®ã¿ï¼‰"""
        if not self.current_case:
            raise ValueError("äº‹ä»¶ãŒé¸æŠã•ã‚Œã¦ã„ã¾ã›ã‚“")
        
        if not self.db_manager:
            raise ValueError("ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ãƒãƒãƒ¼ã‚¸ãƒ£ãƒ¼ãŒåˆæœŸåŒ–ã•ã‚Œã¦ã„ã¾ã›ã‚“")
        
        # Google Driveã‹ã‚‰èª­ã¿è¾¼ã¿
        return self.db_manager.load_database()
    
    def save_database(self, database: dict):
        """database.jsonã®ä¿å­˜ï¼ˆGoogle Driveã®ã¿ï¼‰"""
        if not self.current_case:
            raise ValueError("äº‹ä»¶ãŒé¸æŠã•ã‚Œã¦ã„ã¾ã›ã‚“")
        
        if not self.db_manager:
            raise ValueError("ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ãƒãƒãƒ¼ã‚¸ãƒ£ãƒ¼ãŒåˆæœŸåŒ–ã•ã‚Œã¦ã„ã¾ã›ã‚“")
        
        # ãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿æ›´æ–°
        database["metadata"]["last_updated"] = datetime.now().isoformat()
        database["metadata"]["total_evidence_count"] = len(database["evidence"])
        database["metadata"]["completed_count"] = len([
            e for e in database["evidence"] if e.get("status") == "completed"
        ])
        
        # Google Driveã«ä¿å­˜
        if self.db_manager.save_database(database):
            logger.info(f" Google Driveã«database.jsonã‚’ä¿å­˜ã—ã¾ã—ãŸ")
        else:
            logger.error(f" Google Driveä¿å­˜å¤±æ•—")
            raise Exception("database.jsonã®ä¿å­˜ã«å¤±æ•—ã—ã¾ã—ãŸ")
    
    def display_main_menu(self):
        """ãƒ¡ã‚¤ãƒ³ãƒ¡ãƒ‹ãƒ¥ãƒ¼è¡¨ç¤º"""
        if not self.current_case:
            print("\nâŒ ã‚¨ãƒ©ãƒ¼: äº‹ä»¶ãŒé¸æŠã•ã‚Œã¦ã„ã¾ã›ã‚“")
            return
        
        print("\n" + "="*70)
        print(f"  Phase1_Evidence Analysis System - è¨¼æ‹ ç®¡ç†")
        print(f"  äº‹ä»¶: {self.current_case['case_name']}")
        print("="*70)
        print("\nã€è¨¼æ‹ ã®æ•´ç†ãƒ»åˆ†æã€‘")
        print("  1. è¨¼æ‹ æ•´ç† (æœªåˆ†é¡ãƒ•ã‚©ãƒ«ãƒ€ â†’ æ•´ç†æ¸ˆã¿_æœªç¢ºå®š)")
        print("  2. è¨¼æ‹ åˆ†æ (ç•ªå·æŒ‡å®š: tmp_001 / ç¯„å›²æŒ‡å®š: tmp_001-011)")
        print("  3. AIå¯¾è©±å½¢å¼ã§åˆ†æå†…å®¹ã‚’æ”¹å–„")
        print("\nã€è¨¼æ‹ ã®ç¢ºå®šãƒ»ç®¡ç†ã€‘")
        print("  4. æ—¥ä»˜é †ã«ä¸¦ã³æ›¿ãˆã¦ç¢ºå®š (æ•´ç†æ¸ˆã¿_æœªç¢ºå®š â†’ ç”²å·è¨¼/ä¹™å·è¨¼)")
        print("\nã€è¨¼æ‹ ã®é–²è¦§ãƒ»ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ç”Ÿæˆã€‘")
        print("  5. è¨¼æ‹ åˆ†æä¸€è¦§ã‚’è¡¨ç¤º")
        print("  6. è¨¼æ‹ ä¸€è¦§ã‚’ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆï¼ˆCSV/Excelï¼‰")
        print("  7. CSVç·¨é›†å†…å®¹ã‚’database.jsonã«åæ˜ ")
        print("  8. æ™‚ç³»åˆ—ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã®ç”Ÿæˆï¼ˆè¨¼æ‹ ã‚’æ™‚ç³»åˆ—ã§æ•´ç†ï¼‰")
        print("  9. ä¾é ¼è€…ç™ºè¨€ãƒ»ãƒ¡ãƒ¢ã®ç®¡ç†")
        print("\nã€ã‚·ã‚¹ãƒ†ãƒ ç®¡ç†ã€‘")
        print("  10. database.jsonã®çŠ¶æ…‹ç¢ºèª")
        print("  11. äº‹ä»¶ã‚’åˆ‡ã‚Šæ›¿ãˆ")
        print("  0. çµ‚äº†")
        print("-"*70)
    
    def select_evidence_type(self) -> Optional[str]:
        """è¨¼æ‹ ç¨®åˆ¥ã‚’é¸æŠ
        
        Returns:
            'ko': ç”²å·è¨¼, 'otsu': ä¹™å·è¨¼, None: ã‚­ãƒ£ãƒ³ã‚»ãƒ«
        """
        print("\nè¨¼æ‹ ç¨®åˆ¥ã‚’é¸æŠã—ã¦ãã ã•ã„:")
        print("  1. ç”²å·è¨¼ï¼ˆã“ã¡ã‚‰ã®è¨¼æ‹ ï¼‰")
        print("  2. ä¹™å·è¨¼ï¼ˆç›¸æ‰‹æ–¹ã®è¨¼æ‹ ï¼‰")
        print("  3. ã‚­ãƒ£ãƒ³ã‚»ãƒ«")
        
        choice = input("\n> ").strip()
        
        if choice == '1':
            return 'ko'
        elif choice == '2':
            return 'otsu'
        else:
            return None
    
    def get_evidence_number_input(self, evidence_type: str = 'ko') -> Optional[List[str]]:
        """è¨¼æ‹ ç•ªå·ã®å…¥åŠ›å–å¾—ï¼ˆæ”¹å–„ç‰ˆï¼‰
        
        Args:
            evidence_type: 'ko' ã¾ãŸã¯ 'otsu'
        
        Examples:
            ãƒ¦ãƒ¼ã‚¶ãƒ¼å…¥åŠ›: 001      -> ['tmp_ko_001'] ï¼ˆç”²å·è¨¼ã®å ´åˆï¼‰
            ãƒ¦ãƒ¼ã‚¶ãƒ¼å…¥åŠ›: 070-073  -> ['tmp_ko_070', 'tmp_ko_071', 'tmp_ko_072', 'tmp_ko_073']
            ãƒ¦ãƒ¼ã‚¶ãƒ¼å…¥åŠ›: 001-011  -> ['tmp_ko_001', 'tmp_ko_002', ..., 'tmp_ko_011']
        """
        # è¨¼æ‹ ç¨®åˆ¥ã«å¿œã˜ãŸãƒ—ãƒ¬ãƒ•ã‚£ãƒƒã‚¯ã‚¹ã‚’å–å¾—
        prefix = gconfig.TEMP_PREFIX_MAP.get(evidence_type, 'tmp_ko_')
        evidence_type_name = gconfig.EVIDENCE_TYPE_DISPLAY_NAMES.get(evidence_type, 'ç”²å·è¨¼')
        
        print(f"\nè¨¼æ‹ ç•ªå·ã®å…¥åŠ›ï¼ˆ{evidence_type_name}ï¼‰")
        print(f"  å˜ä¸€æŒ‡å®š: 001, 020")
        print(f"  ç¯„å›²æŒ‡å®š: 001-011, 020-030")
        print(f"  â€» è‡ªå‹•çš„ã« '{prefix}' ãŒä»˜åŠ ã•ã‚Œã¾ã™")
        print(f"  ã‚­ãƒ£ãƒ³ã‚»ãƒ«: ç©ºEnter")
        user_input = input("\n> ").strip()
        
        if not user_input:
            return None
        
        # ç¯„å›²æŒ‡å®šã®å‡¦ç†ï¼ˆä¾‹: 001-011, 020-030ï¼‰
        if '-' in user_input and user_input.count('-') == 1:
            try:
                # ç¯„å›²ã®é–‹å§‹ã¨çµ‚äº†ã‚’åˆ†é›¢
                start_str, end_str = user_input.split('-')
                
                # æ•°å€¤å¤‰æ›ï¼ˆæ•°å­—ã®ã¿ã‚’æƒ³å®šï¼‰
                start_num = int(start_str)
                end_num = int(end_str)
                
                if start_num > end_num:
                    logger.error("é–‹å§‹ç•ªå·ã¯çµ‚äº†ç•ªå·ä»¥ä¸‹ã§ãªã‘ã‚Œã°ãªã‚Šã¾ã›ã‚“")
                    print(f"\nã‚¨ãƒ©ãƒ¼: é–‹å§‹ç•ªå·({start_num})ãŒçµ‚äº†ç•ªå·({end_num})ã‚ˆã‚Šå¤§ãã„ã§ã™")
                    return None
                
                # ç¯„å›²ãŒå¤§ãã™ããªã„ã‹ãƒã‚§ãƒƒã‚¯
                range_size = end_num - start_num + 1
                if range_size > 100:
                    print(f"\nã‚¨ãƒ©ãƒ¼: ç¯„å›²ãŒå¤§ãã™ãã¾ã™({range_size}ä»¶)")
                    print("  ä¸€åº¦ã«å‡¦ç†ã§ãã‚‹ã®ã¯100ä»¶ã¾ã§ã§ã™")
                    return None
                
                # ã‚¼ãƒ­åŸ‹ã‚ã®æ¡æ•°ã‚’åˆ¤å®šï¼ˆé–‹å§‹ç•ªå·ã®æ¡æ•°ã‚’ç¶­æŒï¼‰
                width = len(start_str)
                
                # ç¯„å›²å†…ã®è¨¼æ‹ ç•ªå·ãƒªã‚¹ãƒˆã‚’ç”Ÿæˆï¼ˆãƒ—ãƒ¬ãƒ•ã‚£ãƒƒã‚¯ã‚¹ã‚’è‡ªå‹•ä»˜åŠ ï¼‰
                # ä¾‹: tmp_ko_001, tmp_ko_002, ..., tmp_ko_011
                evidence_numbers = [f"{prefix}{i:0{width}d}" for i in range(start_num, end_num + 1)]
                
                print(f"\nâœ… ç”Ÿæˆã•ã‚ŒãŸè¨¼æ‹ ç•ªå·: {evidence_numbers[0]} ã€œ {evidence_numbers[-1]} ({len(evidence_numbers)}ä»¶)")
                return evidence_numbers
                
            except ValueError as e:
                logger.error(f"ç¯„å›²æŒ‡å®šã®å½¢å¼ãŒæ­£ã—ãã‚ã‚Šã¾ã›ã‚“: {e}")
                print("\nã‚¨ãƒ©ãƒ¼: ç¯„å›²æŒ‡å®šã®å½¢å¼ãŒæ­£ã—ãã‚ã‚Šã¾ã›ã‚“")
                print("  æ­£ã—ã„ä¾‹: 001-011, 020-030ï¼ˆæ•°å­—ã®ã¿ï¼‰")
                print(f"  è©³ç´°: {e}")
                return None
        else:
            # å˜ä¸€è¨¼æ‹ ç•ªå·ï¼ˆæ•°å­—ã®ã¿ã®å…¥åŠ›ã‚’æƒ³å®šï¼‰
            try:
                # æ•°å­—éƒ¨åˆ†ã®ã¿ã‹ãƒã‚§ãƒƒã‚¯
                num = int(user_input)
                width = len(user_input)
                evidence_number = f"{prefix}{num:0{width}d}"
                print(f"\nâœ… ç”Ÿæˆã•ã‚ŒãŸè¨¼æ‹ ç•ªå·: {evidence_number}")
                return [evidence_number]
            except ValueError:
                # æ•°å­—ã§ãªã„å ´åˆã¯ãã®ã¾ã¾è¿”ã™ï¼ˆå¾Œæ–¹äº’æ›æ€§ï¼‰
                logger.warning(f"æ—§å½¢å¼ã®è¨¼æ‹ ç•ªå·å…¥åŠ›: {user_input}")
                print(f"\nâš ï¸ æ—§å½¢å¼ã®è¨¼æ‹ ç•ªå·ã¨ã—ã¦å‡¦ç†: {user_input}")
                return [user_input]
    
    def search_evidence_files_from_gdrive(self) -> List[Dict]:
        """Google Driveã‹ã‚‰è¨¼æ‹ ãƒ•ã‚¡ã‚¤ãƒ«ã‚’æ¤œç´¢
        
        Returns:
            æ¤œå‡ºã•ã‚ŒãŸè¨¼æ‹ ãƒ•ã‚¡ã‚¤ãƒ«ã®ãƒªã‚¹ãƒˆ
        """
        if not self.current_case or not self.current_case.get('ko_evidence_folder_id'):
            logger.error(" ç”²å·è¨¼ãƒ•ã‚©ãƒ«ãƒ€IDãŒè¨­å®šã•ã‚Œã¦ã„ã¾ã›ã‚“")
            return []
        
        print("\nGoogle Driveã‹ã‚‰è¨¼æ‹ ãƒ•ã‚¡ã‚¤ãƒ«ã‚’æ¤œç´¢ä¸­...")
        
        service = self.case_manager.get_google_drive_service()
        if not service:
            logger.error(" Google Driveèªè¨¼ã«å¤±æ•—ã—ã¾ã—ãŸ")
            return []
        
        try:
            ko_folder_id = self.current_case['ko_evidence_folder_id']
            query = f"'{ko_folder_id}' in parents and trashed=false and mimeType!='application/vnd.google-apps.folder'"
            
            results = service.files().list(
                q=query,
                corpora='drive',
                driveId=self.case_manager.shared_drive_root_id,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                fields='files(id, name, mimeType, size, createdTime, modifiedTime, webViewLink, webContentLink)',
                pageSize=1000
            ).execute()
            
            files = results.get('files', [])
            print(f"å®Œäº†: {len(files)}ä»¶ã®è¨¼æ‹ ãƒ•ã‚¡ã‚¤ãƒ«ã‚’æ¤œå‡ºã—ã¾ã—ãŸ")
            
            return files
            
        except Exception as e:
            logger.error(f" Google Driveæ¤œç´¢ã‚¨ãƒ©ãƒ¼: {e}")
            return []
    
    def _get_gdrive_info_from_database(self, evidence_number: str, evidence_type: str = 'ko') -> Optional[Dict]:
        """database.jsonã‹ã‚‰è¨¼æ‹ ã®Google Driveæƒ…å ±ã‚’å–å¾—
        
        Args:
            evidence_number: è¨¼æ‹ ç•ªå·ï¼ˆä¾‹: tmp_001, tmp_020ï¼‰
            evidence_type: è¨¼æ‹ ç¨®åˆ¥ ('ko' ã¾ãŸã¯ 'otsu')
        
        Returns:
            Google Driveãƒ•ã‚¡ã‚¤ãƒ«æƒ…å ±ï¼ˆè¦‹ã¤ã‹ã‚‰ãªã„å ´åˆã¯Noneï¼‰
        """
        try:
            database = self.load_database()
            
            # è¨¼æ‹ ç•ªå·ã‚’æ­£è¦åŒ–ï¼ˆç”²001 â†’ ko001ã§çµ±ä¸€ï¼‰
            normalized_number = evidence_number
            if evidence_number.startswith('ç”²'):
                normalized_number = f"ko{evidence_number[1:]}"
            elif evidence_number.startswith('ä¹™'):
                normalized_number = f"otsu{evidence_number[1:]}"
            
            # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã‹ã‚‰è¨¼æ‹ ã‚’æ¤œç´¢
            # 1. evidence_id ã§æ¤œç´¢ï¼ˆç¢ºå®šæ¸ˆã¿è¨¼æ‹ : ko001, ko002...ï¼‰
            # 2. temp_id ã§æ¤œç´¢ï¼ˆæ•´ç†æ¸ˆã¿_æœªç¢ºå®š: tmp_001, tmp_002...ï¼‰
            for evidence in database.get('evidence', []):
                # ç¢ºå®šæ¸ˆã¿è¨¼æ‹ ã®æ¤œç´¢
                if evidence.get('evidence_id') == normalized_number:
                    # Google Driveãƒ•ã‚¡ã‚¤ãƒ«IDã‚’å–å¾—ï¼ˆè¤‡æ•°ã®å ´æ‰€ã‚’ãƒã‚§ãƒƒã‚¯ï¼‰
                    gdrive_file_id = evidence.get('gdrive_file_id')
                    
                    # gdrive_file_idãŒãªã„å ´åˆã€complete_metadata.gdrive.file_idã‚’ãƒã‚§ãƒƒã‚¯
                    if not gdrive_file_id:
                        metadata = evidence.get('complete_metadata', {})
                        gdrive_info = metadata.get('gdrive', {})
                        gdrive_file_id = gdrive_info.get('file_id')
                    
                    if not gdrive_file_id:
                        logger.warning(f" è¨¼æ‹  {evidence_number} ã®Google Driveãƒ•ã‚¡ã‚¤ãƒ«IDãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“")
                        return None
                    
                    # Google Drive APIã§ãƒ•ã‚¡ã‚¤ãƒ«æƒ…å ±ã‚’å–å¾—
                    service = self.case_manager.get_google_drive_service()
                    if not service:
                        logger.error(" Google Driveèªè¨¼ã«å¤±æ•—ã—ã¾ã—ãŸ")
                        return None
                    
                    file_info = service.files().get(
                        fileId=gdrive_file_id,
                        supportsAllDrives=True,
                        fields='id, name, mimeType, size, createdTime, modifiedTime, webViewLink, webContentLink'
                    ).execute()
                    
                    return file_info
                
                # æœªç¢ºå®šè¨¼æ‹ ã®æ¤œç´¢ï¼ˆtemp_id: tmp_001, tmp_002...ï¼‰
                if evidence.get('temp_id') == evidence_number:
                    # Google Driveãƒ•ã‚¡ã‚¤ãƒ«IDã‚’å–å¾—ï¼ˆè¤‡æ•°ã®å ´æ‰€ã‚’ãƒã‚§ãƒƒã‚¯ï¼‰
                    gdrive_file_id = evidence.get('gdrive_file_id')
                    
                    # gdrive_file_idãŒãªã„å ´åˆã€complete_metadata.gdrive.file_idã‚’ãƒã‚§ãƒƒã‚¯
                    if not gdrive_file_id:
                        metadata = evidence.get('complete_metadata', {})
                        gdrive_info = metadata.get('gdrive', {})
                        gdrive_file_id = gdrive_info.get('file_id')
                    
                    if not gdrive_file_id:
                        logger.warning(f" è¨¼æ‹  {evidence_number} ã®Google Driveãƒ•ã‚¡ã‚¤ãƒ«IDãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“")
                        return None
                    
                    # Google Drive APIã§ãƒ•ã‚¡ã‚¤ãƒ«æƒ…å ±ã‚’å–å¾—
                    service = self.case_manager.get_google_drive_service()
                    if not service:
                        logger.error(" Google Driveèªè¨¼ã«å¤±æ•—ã—ã¾ã—ãŸ")
                        return None
                    
                    file_info = service.files().get(
                        fileId=gdrive_file_id,
                        supportsAllDrives=True,
                        fields='id, name, mimeType, size, createdTime, modifiedTime, webViewLink, webContentLink'
                    ).execute()
                    
                    return file_info
            
            logger.warning(f" è¨¼æ‹  {evidence_number} ãŒdatabase.jsonã«è¦‹ã¤ã‹ã‚Šã¾ã›ã‚“")
            return None
            
        except Exception as e:
            logger.error(f" database.jsonèª­ã¿è¾¼ã¿ã‚¨ãƒ©ãƒ¼: {e}")
            return None
    
    def process_evidence(self, evidence_number: str, gdrive_file_info: Dict = None, evidence_type: str = 'ko') -> bool:
        """è¨¼æ‹ ã®å‡¦ç†ï¼ˆå®Œå…¨ç‰ˆï¼‰
        
        Args:
            evidence_number: è¨¼æ‹ ç•ªå·ï¼ˆä¾‹: tmp_001ï¼‰
            gdrive_file_info: Google Driveãƒ•ã‚¡ã‚¤ãƒ«æƒ…å ±ï¼ˆã‚ªãƒ—ã‚·ãƒ§ãƒ³ï¼‰
            evidence_type: è¨¼æ‹ ç¨®åˆ¥ ('ko' ã¾ãŸã¯ 'otsu')
            
        Returns:
            å‡¦ç†æˆåŠŸ: True, å¤±æ•—: False
        """
        logger.info(f"\n{'='*70}")
        logger.info(f"  è¨¼æ‹  {evidence_number} ã®å‡¦ç†é–‹å§‹")
        logger.info(f"{'='*70}")
        
        try:
            # 1. ãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹ã®å–å¾—ï¼ˆGoogle Driveã‹ã‚‰ï¼‰
            if gdrive_file_info:
                logger.info(f"ãƒ•ã‚¡ã‚¤ãƒ«: {gdrive_file_info['name']}")
                logger.info(f"ğŸ”— URL: {gdrive_file_info.get('webViewLink', 'N/A')}")
                
                # Google Driveã‹ã‚‰ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰
                file_path = self._download_file_from_gdrive(gdrive_file_info)
                if not file_path:
                    logger.error(" ãƒ•ã‚¡ã‚¤ãƒ«ã®ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ã«å¤±æ•—ã—ã¾ã—ãŸ")
                    return False
            else:
                # ãƒ­ãƒ¼ã‚«ãƒ«ãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹ã‚’ä½¿ç”¨
                logger.warning(" Google Driveæƒ…å ±ãŒã‚ã‚Šã¾ã›ã‚“ã€‚ãƒ­ãƒ¼ã‚«ãƒ«ãƒ•ã‚¡ã‚¤ãƒ«ã‚’æŒ‡å®šã—ã¦ãã ã•ã„ã€‚")
                print("\n" + "="*70)
                print(f"  è¨¼æ‹  {evidence_number} ã®ãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„")
                print("="*70)
                print("\nä¾‹:")
                print("  /home/user/webapp/evidence_files/tmp_020.pdf")
                print("  /tmp/sample.pdf")
                print("\nã‚­ãƒ£ãƒ³ã‚»ãƒ«: ç©ºEnter")
                
                file_path_input = input("\nãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹ > ").strip()
                
                if not file_path_input:
                    logger.warning(" ãƒ¦ãƒ¼ã‚¶ãƒ¼ãŒã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
                    print("\nâŒ ã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
                    return False
                
                file_path = file_path_input
                
                if not os.path.exists(file_path):
                    logger.error(f" ãƒ•ã‚¡ã‚¤ãƒ«ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“: {file_path}")
                    print(f"\nâŒ ã‚¨ãƒ©ãƒ¼: ãƒ•ã‚¡ã‚¤ãƒ«ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“")
                    print(f"  ãƒ‘ã‚¹: {file_path}")
                    print("\næŒ‡å®šã•ã‚ŒãŸãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹ãŒæ­£ã—ã„ã‹ç¢ºèªã—ã¦ãã ã•ã„")
                    return False
                
                logger.info(f"âœ… ãƒ­ãƒ¼ã‚«ãƒ«ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ä½¿ç”¨: {file_path}")
            
            # 2. ãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿æŠ½å‡º
            logger.info(f"ãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿ã‚’æŠ½å‡ºä¸­...")
            metadata = self.metadata_extractor.extract_complete_metadata(
                file_path,
                gdrive_file_info=gdrive_file_info
            )
            logger.info(f"  - ãƒ•ã‚¡ã‚¤ãƒ«ãƒãƒƒã‚·ãƒ¥(SHA-256): {metadata['hashes']['sha256'][:16]}...")
            logger.info(f"  - ãƒ•ã‚¡ã‚¤ãƒ«ã‚µã‚¤ã‚º: {metadata['basic']['file_size_human']}")
            
            # 3. ãƒ•ã‚¡ã‚¤ãƒ«å‡¦ç†
            logger.info(f"ãƒ•ã‚¡ã‚¤ãƒ«ã‚’å‡¦ç†ä¸­...")
            file_type = self._detect_file_type(file_path)
            processed_data = self.file_processor.process_file(file_path, file_type)
            logger.info(f"  - ãƒ•ã‚¡ã‚¤ãƒ«å½¢å¼: {processed_data['file_type']}")
            
            # 4. AIåˆ†æï¼ˆGPT-4o Visionï¼‰
            logger.info(f"AIåˆ†æã‚’å®Ÿè¡Œä¸­ï¼ˆGPT-4o Visionï¼‰...")
            analysis_result = self.ai_analyzer.analyze_evidence_complete(
                evidence_id=evidence_number,
                file_path=file_path,
                file_type=file_type,
                gdrive_file_info=gdrive_file_info,
                case_info=self.current_case
            )
            
            # 5. å“è³ªè©•ä¾¡
            quality = analysis_result.get('quality_assessment', {})
            logger.info(f"å“è³ªè©•ä¾¡:")
            logger.info(f"  - å®Œå…¨æ€§ã‚¹ã‚³ã‚¢: {quality.get('completeness_score', 0):.1%}")
            logger.info(f"  - ä¿¡é ¼åº¦ã‚¹ã‚³ã‚¢: {quality.get('confidence_score', 0):.1%}")
            logger.info(f"  - è¨€èªåŒ–ãƒ¬ãƒ™ãƒ«: {quality.get('verbalization_level', 0)}")
            
            # 6. database.jsonã«è¿½åŠ 
            logger.info(f"database.jsonã«ä¿å­˜ä¸­...")
            database = self.load_database()
            
            evidence_entry = {
                "evidence_id": evidence_number,
                "evidence_number": f"ç”²{evidence_number.lstrip('ko')}",
                "original_filename": gdrive_file_info['name'] if gdrive_file_info else os.path.basename(file_path),
                "complete_metadata": metadata,
                "phase1_complete_analysis": analysis_result,
                "status": "completed",
                "processed_at": datetime.now().isoformat()
            }
            
            # æ—¢å­˜ã®ã‚¨ãƒ³ãƒˆãƒªã‚’æ›´æ–°ã€ã¾ãŸã¯æ–°è¦è¿½åŠ 
            # temp_id, evidence_id, evidence_number ã®ã„ãšã‚Œã‹ã§ãƒãƒƒãƒãƒ³ã‚°
            existing_index = next(
                (i for i, e in enumerate(database["evidence"]) 
                 if (e.get("evidence_id") == evidence_number or
                     e.get("temp_id") == evidence_number or
                     e.get("evidence_number") == evidence_number)),
                None
            )
            
            if existing_index is not None:
                # æ—¢å­˜ã‚¨ãƒ³ãƒˆãƒªã®temp_idã‚’ä¿æŒ
                old_entry = database["evidence"][existing_index]
                if 'temp_id' in old_entry:
                    evidence_entry['temp_id'] = old_entry['temp_id']
                if 'temp_number' in old_entry:
                    evidence_entry['temp_number'] = old_entry['temp_number']
                
                database["evidence"][existing_index] = evidence_entry
                logger.info(f"  âœ… æ—¢å­˜ã‚¨ãƒ³ãƒˆãƒªã‚’æ›´æ–°ã—ã¾ã—ãŸï¼ˆtemp_id: {old_entry.get('temp_id')}ï¼‰")
            else:
                database["evidence"].append(evidence_entry)
                logger.info(f"  âœ… æ–°è¦ã‚¨ãƒ³ãƒˆãƒªã‚’è¿½åŠ ã—ã¾ã—ãŸ")
            
            self.save_database(database)
            
            logger.info(f"\nâœ… è¨¼æ‹  {evidence_number} ã®å‡¦ç†ãŒå®Œäº†ã—ã¾ã—ãŸï¼")
            return True
            
        except Exception as e:
            logger.error(f" ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {e}", exc_info=True)
            return False
    
    def _download_file_from_gdrive(self, file_info: Dict) -> Optional[str]:
        """Google Driveã‹ã‚‰ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰"""
        try:
            import io
            from googleapiclient.http import MediaIoBaseDownload
            
            service = self.case_manager.get_google_drive_service()
            file_id = file_info['id']
            file_name = file_info['name']
            
            # ä¸€æ™‚ãƒ‡ã‚£ãƒ¬ã‚¯ãƒˆãƒªã«ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰
            temp_dir = gconfig.LOCAL_TEMP_DIR
            os.makedirs(temp_dir, exist_ok=True)
            
            output_path = os.path.join(temp_dir, file_name)
            
            request = service.files().get_media(fileId=file_id)
            fh = io.FileIO(output_path, 'wb')
            downloader = MediaIoBaseDownload(fh, request)
            
            done = False
            while not done:
                status, done = downloader.next_chunk()
                if status:
                    logger.info(f"  ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰é€²æ—: {int(status.progress() * 100)}%")
            
            fh.close()
            logger.info(f"  âœ… ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰å®Œäº†: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f" ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ã‚¨ãƒ©ãƒ¼: {e}")
            return None
    
    def _detect_file_type(self, file_path: str) -> str:
        """ãƒ•ã‚¡ã‚¤ãƒ«å½¢å¼ã‚’æ¤œå‡º"""
        ext = os.path.splitext(file_path)[1].lower()
        
        for file_type, info in gconfig.SUPPORTED_FORMATS.items():
            if ext in info['extensions']:
                return file_type
        
        return 'document'  # ãƒ‡ãƒ•ã‚©ãƒ«ãƒˆ
    
    def finalize_pending_evidence(self):
        """æ•´ç†æ¸ˆã¿_æœªç¢ºå®šã®è¨¼æ‹ ã‚’ä¸¦ã³æ›¿ãˆã¦ç¢ºå®š"""
        print("\n" + "="*70)
        print("  è¨¼æ‹ ã®ä¸¦ã³æ›¿ãˆãƒ»ç¢ºå®š")
        print("="*70)
        
        # database.jsonã‹ã‚‰æœªç¢ºå®šè¨¼æ‹ ã‚’å–å¾—
        database = self.load_database()
        pending_evidence = [e for e in database.get('evidence', []) if e.get('status') == 'pending']
        
        if not pending_evidence:
            print("\næœªç¢ºå®šã®è¨¼æ‹ ã¯ã‚ã‚Šã¾ã›ã‚“")
            return
        
        print(f"\næœªç¢ºå®šè¨¼æ‹ : {len(pending_evidence)}ä»¶")
        print("\nç¾åœ¨ã®é †åº:")
        for idx, evidence in enumerate(pending_evidence, 1):
            print(f"  [{idx}] {evidence['temp_id']} - {evidence['original_filename']}")
            print(f"      ç¨®åˆ¥: {evidence['evidence_type']}, èª¬æ˜: {evidence['description']}")
        
        print("\næ“ä½œã‚’é¸æŠã—ã¦ãã ã•ã„:")
        print("  1: ã“ã®é †åºã§ç¢ºå®š (ç”²001, ç”²002...)")
        print("  2: é †åºã‚’å¤‰æ›´")
        print("  0: ã‚­ãƒ£ãƒ³ã‚»ãƒ«")
        
        while True:
            choice = input("\n> ").strip()
            if choice in ['0', '1', '2']:
                break
            print("ã‚¨ãƒ©ãƒ¼: 0, 1, 2 ã®ã„ãšã‚Œã‹ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„")
        
        if choice == '0':
            print("ã‚¨ãƒ©ãƒ¼: ã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
            return
        
        elif choice == '2':
            # é †åºå¤‰æ›´
            print("\né †åºã‚’å¤‰æ›´ã—ã¾ã™")
            print("   ä¾‹: 1,3,2,4 â†’ 1ç•ªç›®,3ç•ªç›®,2ç•ªç›®,4ç•ªç›®ã®é †")
            new_order_input = input(f"æ–°ã—ã„é †åºã‚’å…¥åŠ› (1-{len(pending_evidence)}ã‚’ã‚«ãƒ³ãƒåŒºåˆ‡ã‚Š): ").strip()
            
            try:
                new_order = [int(x.strip()) for x in new_order_input.split(',')]
                if len(new_order) != len(pending_evidence) or set(new_order) != set(range(1, len(pending_evidence) + 1)):
                    print("ã‚¨ãƒ©ãƒ¼: ç„¡åŠ¹ãªé †åºã§ã™")
                    return
                
                # ä¸¦ã³æ›¿ãˆ
                pending_evidence = [pending_evidence[i-1] for i in new_order]
                
                print("\nâœ… ä¸¦ã³æ›¿ãˆå¾Œ:")
                for idx, evidence in enumerate(pending_evidence, 1):
                    print(f"  [{idx}] {evidence['temp_id']} - {evidence['original_filename']}")
                
            except ValueError:
                print("ã‚¨ãƒ©ãƒ¼: å…¥åŠ›ã‚¨ãƒ©ãƒ¼")
                return
        
        # ç¢ºå®šç¢ºèª
        confirm = input(f"\nã“ã®é †åºã§ç¢ºå®šã—ã¾ã™ã‹ï¼Ÿ (y/n): ").strip().lower()
        if confirm != 'y':
            print("ã‚¨ãƒ©ãƒ¼: ã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
            return
        
        # ç¢ºå®šå‡¦ç†
        print("\nğŸ“¥ è¨¼æ‹ ã‚’ç¢ºå®šä¸­...")
        service = self.case_manager.get_google_drive_service()
        if not service:
            print("ã‚¨ãƒ©ãƒ¼: Google Driveèªè¨¼ã«å¤±æ•—ã—ã¾ã—ãŸ")
            return
        
        success_count = 0
        ko_folder_id = self.current_case['ko_evidence_folder_id']
        
        # æ•´ç†æ¸ˆã¿_æœªç¢ºå®šãƒ•ã‚©ãƒ«ãƒ€IDã‚’å–å¾—
        from evidence_organizer import EvidenceOrganizer
        organizer = EvidenceOrganizer(self.case_manager, self.current_case)
        pending_folder_id = organizer.pending_folder_id
        
        for idx, evidence in enumerate(pending_evidence, 1):
            ko_number = idx
            ko_id = f"ko{ko_number:03d}"
            ko_number_kanji = f"ç”²{ko_number:03d}"
            
            print(f"\n[{idx}/{len(pending_evidence)}] {evidence['temp_id']} â†’ {ko_id}")
            
            try:
                # ãƒ•ã‚¡ã‚¤ãƒ«ã‚’å–å¾—
                file_id = evidence['gdrive_file_id']
                
                # æ–°ã—ã„ãƒ•ã‚¡ã‚¤ãƒ«åã‚’ç”Ÿæˆ
                old_filename = evidence['renamed_filename']
                # tmp_XXX_ ã®éƒ¨åˆ†ã‚’ koXXX_ ã«ç½®æ›
                new_filename = old_filename.replace(evidence['temp_id'], ko_id)
                
                # ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ç§»å‹•ã—ã¦ãƒªãƒãƒ¼ãƒ 
                file = service.files().update(
                    fileId=file_id,
                    addParents=ko_folder_id,
                    removeParents=pending_folder_id,
                    body={'name': new_filename},
                    supportsAllDrives=True,
                    fields='id, name'
                ).execute()
                
                print(f"  âœ… {new_filename}")
                
                # database.jsonã®è¨¼æ‹ æƒ…å ±ã‚’æ›´æ–°
                evidence['evidence_id'] = ko_id
                evidence['evidence_number'] = ko_number_kanji
                evidence['renamed_filename'] = new_filename
                evidence['status'] = 'completed'
                evidence['confirmed_at'] = datetime.now().isoformat()
                
                success_count += 1
                
            except Exception as e:
                print(f"  âŒ ã‚¨ãƒ©ãƒ¼: {e}")
        
        # database.jsonã‚’ä¿å­˜
        self.save_database(database)
        
        print("\n" + "="*70)
        print(f"å®Œäº†: ç¢ºå®šå®Œäº†: {success_count}/{len(pending_evidence)}ä»¶")
        print("="*70)
    
    def analyze_and_sort_pending_evidence(self, evidence_type: str = 'ko'):
        """æœªç¢ºå®šè¨¼æ‹ ã‚’AIåˆ†æâ†’æ—¥ä»˜æŠ½å‡ºâ†’è‡ªå‹•ã‚½ãƒ¼ãƒˆâ†’ç¢ºå®š
        
        Args:
            evidence_type: è¨¼æ‹ ç¨®åˆ¥ ('ko' ã¾ãŸã¯ 'otsu')
        """
        type_name = "ç”²å·è¨¼" if evidence_type == 'ko' else "ä¹™å·è¨¼"
        type_folder = "ç”²å·è¨¼" if evidence_type == 'ko' else "ä¹™å·è¨¼"
        
        print("\n" + "="*70)
        print(f"  æœªç¢ºå®šè¨¼æ‹ ã®åˆ†æãƒ»æ—¥ä»˜æŠ½å‡ºãƒ»è‡ªå‹•ã‚½ãƒ¼ãƒˆ [{type_name}]")
        print("="*70)
        
        # database.jsonã‹ã‚‰æœªç¢ºå®šè¨¼æ‹ ã‚’å–å¾—ï¼ˆè¨¼æ‹ ç¨®åˆ¥ã§ãƒ•ã‚£ãƒ«ã‚¿ãƒ¼ï¼‰
        database = self.load_database()
        pending_evidence = [
            e for e in database.get('evidence', []) 
            if e.get('status') == 'pending' and e.get('evidence_type', 'ko') == evidence_type
        ]
        
        if not pending_evidence:
            print(f"\n{type_name}ã®æœªç¢ºå®šè¨¼æ‹ ã¯ã‚ã‚Šã¾ã›ã‚“")
            return
        
        print(f"\n{type_name}ã®æœªç¢ºå®šè¨¼æ‹ : {len(pending_evidence)}ä»¶")
        print("\nç¾åœ¨ã®é †åº:")
        for idx, evidence in enumerate(pending_evidence, 1):
            print(f"  [{idx}] {evidence['temp_id']} - {evidence['original_filename']}")
        
        prefix = "ko" if evidence_type == 'ko' else "otsu"
        print("\nã€å‡¦ç†å†…å®¹ã€‘")
        print("  1. å„è¨¼æ‹ ã‹ã‚‰ä½œæˆå¹´æœˆæ—¥ã‚’å–å¾—ï¼ˆæ—¢ã«åˆ†ææ¸ˆã¿ãªã‚‰document_dateã‚’ä½¿ç”¨ï¼‰")
        print("  2. ä½œæˆå¹´æœˆæ—¥é †ã«è‡ªå‹•ã‚½ãƒ¼ãƒˆï¼ˆå¤ã„é †ï¼‰")
        print(f"  3. ã‚½ãƒ¼ãƒˆå¾Œã®é †åºã§ç¢ºå®šç•ªå·ï¼ˆ{prefix}001, {prefix}002, {prefix}003...ï¼‰ã‚’å‰²ã‚Šå½“ã¦")
        print(f"  4. æ•´ç†æ¸ˆã¿_æœªç¢ºå®š â†’ {type_folder} ãƒ•ã‚©ãƒ«ãƒ€ã¸ç§»å‹•")
        
        confirm = input("\nå‡¦ç†ã‚’é–‹å§‹ã—ã¾ã™ã‹ï¼Ÿ (y/n): ").strip().lower()
        if confirm != 'y':
            print("ã‚¨ãƒ©ãƒ¼: ã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
            return
        
        # ã‚¹ãƒ†ãƒƒãƒ—1: ä½œæˆå¹´æœˆæ—¥ã®å–å¾—
        print("\n" + "="*70)
        print("  [1/3] ä½œæˆå¹´æœˆæ—¥ã®å–å¾—ä¸­...")
        print("="*70)
        
        service = self.case_manager.get_google_drive_service()
        if not service:
            print("ã‚¨ãƒ©ãƒ¼: Google Driveèªè¨¼ã«å¤±æ•—ã—ã¾ã—ãŸ")
            return
        
        for idx, evidence in enumerate(pending_evidence, 1):
            print(f"\n[{idx}/{len(pending_evidence)}] {evidence['temp_id']} - {evidence['original_filename']}")
            
            # ã¾ãšã€æ—¢å­˜ã®AIåˆ†æã‹ã‚‰document_dateã‚’å–å¾—
            document_date = None
            if 'phase1_complete_analysis' in evidence:
                ai_analysis = evidence['phase1_complete_analysis'].get('ai_analysis', {})
                obj_analysis = ai_analysis.get('objective_analysis', {})
                temporal_info = obj_analysis.get('temporal_information', {})
                # v3.6.2ä»¥é™: document_dateã€v3.6.1ä»¥å‰: creation_dateï¼ˆå¾Œæ–¹äº’æ›æ€§ï¼‰
                document_date = temporal_info.get('document_date') or temporal_info.get('creation_date')
                
                if document_date:
                    print(f"  âœ… æ—¢å­˜åˆ†æã‹ã‚‰å–å¾—: {document_date}")
                    evidence['extracted_date'] = document_date
                    continue
            
            # æ—¢å­˜åˆ†æãŒãªã„å ´åˆã®ã¿ã€åˆ¥é€”æ—¥ä»˜æŠ½å‡ºã‚’å®Ÿè¡Œ
            print(f"  âš ï¸ æœªåˆ†æã®ãŸã‚æ—¥ä»˜æŠ½å‡ºã‚’å®Ÿè¡Œ...")
            
            try:
                # Google Driveã‹ã‚‰ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰
                gdrive_file_id = evidence.get('gdrive_file_id')
                if not gdrive_file_id:
                    print(f"  âš ï¸ Google Driveãƒ•ã‚¡ã‚¤ãƒ«IDãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“")
                    evidence['extracted_date'] = None
                    continue
                
                # ãƒ•ã‚¡ã‚¤ãƒ«æƒ…å ±ã‚’å–å¾—
                file_info = service.files().get(
                    fileId=gdrive_file_id,
                    supportsAllDrives=True,
                    fields='id, name, mimeType'
                ).execute()
                
                # ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰
                file_path = self._download_file_from_gdrive(file_info)
                if not file_path:
                    print(f"  âš ï¸ ãƒ•ã‚¡ã‚¤ãƒ«ã®ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ã«å¤±æ•—ã—ã¾ã—ãŸ")
                    evidence['extracted_date'] = None
                    continue
                
                # ãƒ•ã‚¡ã‚¤ãƒ«ã‚¿ã‚¤ãƒ—ã‚’æ¤œå‡º
                file_type = self._detect_file_type(file_path)
                
                # æ—¥ä»˜æŠ½å‡ºï¼ˆè»½é‡ç‰ˆï¼‰
                date_result = self.ai_analyzer.extract_date_from_evidence(
                    evidence_id=evidence['temp_id'],
                    file_path=file_path,
                    file_type=file_type,
                    original_filename=evidence['original_filename']
                )
                
                # çµæœã‚’è¨¼æ‹ æƒ…å ±ã«è¿½åŠ 
                evidence['date_extraction'] = date_result
                evidence['extracted_date'] = date_result.get('primary_date')
                
                if evidence['extracted_date']:
                    print(f"  ğŸ“… æŠ½å‡ºæ—¥ä»˜: {evidence['extracted_date']}")
                else:
                    print(f"  âš ï¸ æ—¥ä»˜ãŒæŠ½å‡ºã§ãã¾ã›ã‚“ã§ã—ãŸ")
                
            except Exception as e:
                print(f"  âŒ ã‚¨ãƒ©ãƒ¼: {e}")
                evidence['extracted_date'] = None
        
        # ã‚¹ãƒ†ãƒƒãƒ—2: ä½œæˆå¹´æœˆæ—¥é †ã«ã‚½ãƒ¼ãƒˆ
        print("\n" + "="*70)
        print("  [2/3] ä½œæˆå¹´æœˆæ—¥é †ã«ã‚½ãƒ¼ãƒˆä¸­...")
        print("="*70)
        
        # æ—¥ä»˜ãŒå–å¾—ã§ããŸã‚‚ã®ã¨å–å¾—ã§ããªã‹ã£ãŸã‚‚ã®ã«åˆ†é›¢
        with_date = [e for e in pending_evidence if e.get('extracted_date')]
        without_date = [e for e in pending_evidence if not e.get('extracted_date')]
        
        # ä½œæˆå¹´æœˆæ—¥é †ã«ã‚½ãƒ¼ãƒˆï¼ˆå¤ã„é †ï¼‰
        with_date.sort(key=lambda e: e['extracted_date'])
        
        # ã‚½ãƒ¼ãƒˆå¾Œã®é †åºï¼ˆæ—¥ä»˜ã‚ã‚Šâ†’æ—¥ä»˜ãªã—ï¼‰
        sorted_evidence = with_date + without_date
        
        print(f"\nâœ… ã‚½ãƒ¼ãƒˆå®Œäº†:")
        print(f"  - æ—¥ä»˜å–å¾—æˆåŠŸ: {len(with_date)}ä»¶")
        print(f"  - æ—¥ä»˜å–å¾—å¤±æ•—: {len(without_date)}ä»¶")
        
        print("\nã‚½ãƒ¼ãƒˆå¾Œã®é †åº:")
        for idx, evidence in enumerate(sorted_evidence, 1):
            date_str = evidence.get('extracted_date', 'æ—¥ä»˜ãªã—')
            print(f"  [{idx}] {evidence['temp_id']} - {evidence['original_filename']} ({date_str})")
        
        # ã‚¹ãƒ†ãƒƒãƒ—3: ç¢ºå®šç•ªå·å‰²ã‚Šå½“ã¦ãƒ»ç§»å‹•
        print("\n" + "="*70)
        print("  [3/3] ç¢ºå®šç•ªå·ã‚’å‰²ã‚Šå½“ã¦ä¸­...")
        print("="*70)
        
        confirm_finalize = input("\nã“ã®é †åºã§ç¢ºå®šã—ã¾ã™ã‹ï¼Ÿ (y/n): ").strip().lower()
        if confirm_finalize != 'y':
            print("ã‚¨ãƒ©ãƒ¼: ã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸï¼ˆæ—¥ä»˜æŠ½å‡ºçµæœã¯database.jsonã«ä¿å­˜ã•ã‚Œã¾ã—ãŸï¼‰")
            self.save_database(database)
            return
        
        success_count = 0
        ko_folder_id = self.current_case['ko_evidence_folder_id']
        
        # æ•´ç†æ¸ˆã¿_æœªç¢ºå®šãƒ•ã‚©ãƒ«ãƒ€IDã‚’å–å¾—
        from evidence_organizer import EvidenceOrganizer
        organizer = EvidenceOrganizer(self.case_manager, self.current_case)
        pending_folder_id = organizer.pending_folder_id
        
        for idx, evidence in enumerate(sorted_evidence, 1):
            ko_number = idx
            ko_id = f"ko{ko_number:03d}"
            ko_number_kanji = f"ç”²{ko_number:03d}"
            
            print(f"\n[{idx}/{len(sorted_evidence)}] {evidence['temp_id']} â†’ {ko_id}")
            date_str = evidence.get('extracted_date', 'æ—¥ä»˜ãªã—')
            print(f"  æ—¥ä»˜: {date_str}")
            
            try:
                # ãƒ•ã‚¡ã‚¤ãƒ«ã‚’å–å¾—
                file_id = evidence['gdrive_file_id']
                
                # æ–°ã—ã„ãƒ•ã‚¡ã‚¤ãƒ«åã‚’ç”Ÿæˆ
                old_filename = evidence['renamed_filename']
                # tmp_XXX_ ã®éƒ¨åˆ†ã‚’ koXXX_ ã«ç½®æ›
                new_filename = old_filename.replace(evidence['temp_id'], ko_id)
                
                # ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ç§»å‹•ã—ã¦ãƒªãƒãƒ¼ãƒ 
                file = service.files().update(
                    fileId=file_id,
                    addParents=ko_folder_id,
                    removeParents=pending_folder_id,
                    body={'name': new_filename},
                    supportsAllDrives=True,
                    fields='id, name'
                ).execute()
                
                print(f"  âœ… {new_filename}")
                
                # database.jsonã®è¨¼æ‹ æƒ…å ±ã‚’æ›´æ–°
                evidence['evidence_id'] = ko_id
                evidence['evidence_number'] = ko_number_kanji
                evidence['renamed_filename'] = new_filename
                evidence['status'] = 'completed'
                evidence['confirmed_at'] = datetime.now().isoformat()
                evidence['sorted_by_date'] = True  # æ—¥ä»˜ã‚½ãƒ¼ãƒˆã§ç¢ºå®šã—ãŸã“ã¨ã‚’è¨˜éŒ²
                
                success_count += 1
                
            except Exception as e:
                print(f"  âŒ ã‚¨ãƒ©ãƒ¼: {e}")
        
        # database.jsonã‚’ä¿å­˜
        self.save_database(database)
        
        print("\n" + "="*70)
        print(f"å®Œäº†: ç¢ºå®šå®Œäº†: {success_count}/{len(sorted_evidence)}ä»¶")
        print(f"æ—¥ä»˜é †ã‚½ãƒ¼ãƒˆ: {len(with_date)}ä»¶")
        print("="*70)
    
    def show_database_status(self):
        """database.jsonã®çŠ¶æ…‹è¡¨ç¤º"""
        database = self.load_database()
        
        print("\n" + "="*70)
        print("  database.json çŠ¶æ…‹ç¢ºèª")
        print("="*70)
        
        print(f"\näº‹ä»¶æƒ…å ±:")
        metadata = database.get('metadata', {})
        print(f"  äº‹ä»¶ID: {metadata.get('case_id', 'N/A')}")
        print(f"  äº‹ä»¶å: {metadata.get('case_name', 'N/A')}")
        print(f"  ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ãƒãƒ¼ã‚¸ãƒ§ãƒ³: {metadata.get('database_version', database.get('version', 'N/A'))}")
        print(f"  æœ€çµ‚æ›´æ–°: {metadata.get('last_updated', 'N/A')}")
        
        print(f"\nè¨¼æ‹ çµ±è¨ˆ:")
        print(f"  ç·è¨¼æ‹ æ•°: {len(database['evidence'])}ä»¶")
        
        completed = [e for e in database['evidence'] if e.get('status') == 'completed']
        print(f"  å®Œäº†: {len(completed)}ä»¶")
        
        pending = [e for e in database['evidence'] if e.get('status') == 'pending']
        print(f"  æœªç¢ºå®š: {len(pending)}ä»¶")
        
        in_progress = [e for e in database['evidence'] if e.get('status') == 'in_progress']
        print(f"  å‡¦ç†ä¸­: {len(in_progress)}ä»¶")
        
        if database['evidence']:
            print(f"\nè¨¼æ‹ ä¸€è¦§ (æœ€å¤§20ä»¶):")
            for evidence in database['evidence'][:20]:
                status = evidence.get('status', 'unknown')
                if status == 'completed':
                    status_text = "[å®Œäº†]"
                    evidence_id = evidence.get('evidence_number', evidence.get('evidence_id', 'N/A'))
                elif status == 'pending':
                    status_text = "[æœªç¢ºå®š]"
                    evidence_id = evidence.get('temp_id', 'N/A')
                else:
                    status_text = "[ä¸æ˜]"
                    evidence_id = 'N/A'
                
                print(f"  {status_text} {evidence_id} - {evidence.get('original_filename', 'N/A')}")
            
            if len(database['evidence']) > 20:
                print(f"  ... ä»– {len(database['evidence']) - 20}ä»¶")
        
        print("\n" + "="*70)
    
    def edit_evidence_with_ai(self, evidence_type: str = 'ko'):
        """AIå¯¾è©±å½¢å¼ã§è¨¼æ‹ å†…å®¹ã‚’ç·¨é›†
        
        Args:
            evidence_type: è¨¼æ‹ ç¨®åˆ¥ ('ko' ã¾ãŸã¯ 'otsu')
        """
        if not self.current_case:
            raise ValueError("äº‹ä»¶ãŒé¸æŠã•ã‚Œã¦ã„ã¾ã›ã‚“")
        
        type_name = "ç”²å·è¨¼" if evidence_type == 'ko' else "ä¹™å·è¨¼"
        print("\n" + "="*70)
        print(f"  AIå¯¾è©±å½¢å¼ã§è¨¼æ‹ å†…å®¹ã‚’æ”¹å–„ [{type_name}]")
        print("="*70)
        
        # è¨¼æ‹ ç•ªå·ã‚’å…¥åŠ›
        evidence_numbers = self.get_evidence_number_input(evidence_type)
        if not evidence_numbers:
            print("\nã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
            return
        
        # 1ä»¶ãšã¤å‡¦ç†
        for evidence_number in evidence_numbers:
            print(f"\nå‡¦ç†ä¸­: {evidence_number}")
            
            # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã‹ã‚‰è¨¼æ‹ ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
            database = self.db_manager.load_database()
            evidence_data = None
            
            for evidence in database.get('evidence', []):
                # evidence_id ã¾ãŸã¯ temp_id ã§æ¤œç´¢
                if (evidence.get('evidence_id') == evidence_number or 
                    evidence.get('temp_id') == evidence_number or
                    evidence.get('evidence_number') == evidence_number):
                    evidence_data = evidence
                    break
            
            if not evidence_data:
                print(f"\nã‚¨ãƒ©ãƒ¼: è¨¼æ‹  {evidence_number} ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“")
                continue
            
            # AIåˆ†æçµæœãŒå­˜åœ¨ã™ã‚‹ã‹ç¢ºèª
            if 'phase1_complete_analysis' not in evidence_data:
                print(f"\nã‚¨ãƒ©ãƒ¼: {evidence_number} ã¯ã¾ã AIåˆ†æã•ã‚Œã¦ã„ã¾ã›ã‚“")
                print("  å…ˆã«ãƒ¡ãƒ‹ãƒ¥ãƒ¼ã€Œ2ã€ã¾ãŸã¯ã€Œ3ã€ã§åˆ†æã‚’å®Ÿè¡Œã—ã¦ãã ã•ã„")
                continue
            
            # AIå¯¾è©±å½¢å¼ã§ç·¨é›†
            modified_data = self.evidence_editor.edit_evidence_interactive(
                evidence_data,
                self.db_manager
            )
            
            # ç·¨é›†ãŒã‚­ãƒ£ãƒ³ã‚»ãƒ«ã•ã‚ŒãŸå ´åˆ
            if modified_data is None:
                print(f"\n{evidence_number} ã®ç·¨é›†ã‚’ã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
                continue
            
            # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã‚’æ›´æ–°
            print(f"\n{evidence_number} ã‚’ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã«ä¿å­˜ä¸­...")
            
            for i, evidence in enumerate(database['evidence']):
                if (evidence.get('evidence_id') == evidence_number or 
                    evidence.get('temp_id') == evidence_number or
                    evidence.get('evidence_number') == evidence_number):
                    database['evidence'][i] = modified_data
                    break
            
            # ä¿å­˜
            self.db_manager.save_database(database)
            print(f"âœ… {evidence_number} ã®å¤‰æ›´ã‚’ä¿å­˜ã—ã¾ã—ãŸ")
    
    def _get_evidence_display_data(self, evidence: Dict) -> Dict:
        """è¨¼æ‹ ãƒ‡ãƒ¼ã‚¿ã‹ã‚‰è¡¨ç¤ºç”¨æƒ…å ±ã‚’æŠ½å‡ºï¼ˆãƒ‡ãƒ¼ã‚¿æ§‹é€ ã®é•ã„ã‚’å¸åï¼‰
        
        Args:
            evidence: è¨¼æ‹ ãƒ‡ãƒ¼ã‚¿
            
        Returns:
            è¡¨ç¤ºç”¨ãƒ‡ãƒ¼ã‚¿ï¼ˆfile_name, creation_date, analysis_statusï¼‰
        """
        # ãƒ•ã‚¡ã‚¤ãƒ«åã®å–å¾—ï¼ˆè¤‡æ•°ã®å ´æ‰€ã‚’è©¦è¡Œï¼‰
        file_name = (
            evidence.get('file_name') or
            evidence.get('original_filename') or
            evidence.get('complete_metadata', {}).get('basic', {}).get('file_name') or
            'ä¸æ˜'
        )
        
        # åˆ†æçµæœã®å–å¾—ï¼ˆãƒ‡ãƒ¼ã‚¿æ§‹é€ ã®é•ã„ã‚’å¸åï¼‰
        phase1_analysis = evidence.get('phase1_complete_analysis', {})
        
        # æ–°ã—ã„æ§‹é€ : phase1_complete_analysis.ai_analysis.full_content.complete_description
        ai_analysis = phase1_analysis.get('ai_analysis', {})
        
        # ä½œæˆæ—¥ã®å–å¾—ï¼ˆå„ªå…ˆé †ä½ï¼šAIåˆ†æçµæœ > PDFãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿ > ãƒ•ã‚¡ã‚¤ãƒ«ã‚·ã‚¹ãƒ†ãƒ ï¼‰
        # 1. AIåˆ†æçµæœã‹ã‚‰æ–‡æ›¸ã®ä½œæˆæ—¥ã‚’å–å¾—ï¼ˆæœ€å„ªå…ˆï¼‰
        document_date = None
        if ai_analysis:
            objective_analysis = ai_analysis.get('objective_analysis', {})
            temporal_info = objective_analysis.get('temporal_information', {})
            document_date = temporal_info.get('document_date')
        
        # 2. æ—§æ§‹é€ ã®ãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯
        if not document_date:
            document_date = (
                phase1_analysis.get('objective_analysis', {}).get('temporal_information', {}).get('document_date') or
                evidence.get('temporal_information', {}).get('document_date')
            )
        
        # 3. PDFãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿ã‹ã‚‰å–å¾—ï¼ˆãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯ï¼‰
        if not document_date:
            pdf_date = evidence.get('complete_metadata', {}).get('format_specific', {}).get('document_info', {}).get('CreationDate', '')
            if pdf_date and len(pdf_date) >= 10:
                # "D:20220103..." -> "2022-01-03" ã«å¤‰æ›
                if pdf_date.startswith('D:'):
                    pdf_date = pdf_date[2:]
                if len(pdf_date) >= 8:
                    document_date = f"{pdf_date[:4]}-{pdf_date[4:6]}-{pdf_date[6:8]}"
        
        # 4. ãƒ•ã‚¡ã‚¤ãƒ«ã‚·ã‚¹ãƒ†ãƒ ä¸Šã®ä½œæˆæ—¥ï¼ˆæœ€çµ‚ãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯ï¼‰
        if not document_date:
            fs_date = evidence.get('complete_metadata', {}).get('basic', {}).get('created_time', '')
            if fs_date:
                document_date = fs_date[:10]  # YYYY-MM-DDéƒ¨åˆ†ã®ã¿
        
        creation_date = document_date or 'ä¸æ˜'
        if ai_analysis:
            full_content = ai_analysis.get('full_content', {})
            complete_description = full_content.get('complete_description')
        else:
            # æ—§æ§‹é€ : phase1_complete_analysis.complete_description ã¾ãŸã¯ full_content.complete_description
            complete_description = (
                phase1_analysis.get('complete_description') or
                phase1_analysis.get('full_content', {}).get('complete_description') or
                evidence.get('full_content', {}).get('complete_description')
            )
        
        analysis_status = "âœ… åˆ†ææ¸ˆã¿" if complete_description else "âš ï¸  æœªåˆ†æ"
        
        return {
            'file_name': file_name,
            'creation_date': creation_date,
            'analysis_status': analysis_status,
            'complete_description': complete_description
        }
    
    def _get_evidence_export_data(self, evidence: Dict) -> Dict:
        """è¨¼æ‹ ãƒ‡ãƒ¼ã‚¿ã‹ã‚‰ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆç”¨è©³ç´°æƒ…å ±ã‚’æŠ½å‡ºï¼ˆãƒ‡ãƒ¼ã‚¿æ§‹é€ ã®é•ã„ã‚’å¸åï¼‰
        
        Args:
            evidence: è¨¼æ‹ ãƒ‡ãƒ¼ã‚¿
            
        Returns:
            ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆç”¨ãƒ‡ãƒ¼ã‚¿ï¼ˆåŸºæœ¬æƒ…å ±+åˆ†æè©³ç´°ï¼‰
        """
        # åŸºæœ¬æƒ…å ±ã‚’å–å¾—
        display_data = self._get_evidence_display_data(evidence)
        
        # åˆ†æè©³ç´°ã®å–å¾—ï¼ˆãƒ‡ãƒ¼ã‚¿æ§‹é€ ã®é•ã„ã‚’å¸åï¼‰
        phase1_analysis = evidence.get('phase1_complete_analysis', {})
        ai_analysis = phase1_analysis.get('ai_analysis', {})
        
        if ai_analysis:
            # æ–°ã—ã„æ§‹é€ : phase1_complete_analysis.ai_analysis.objective_analysis
            objective_analysis = ai_analysis.get('objective_analysis', {})
            document_type = objective_analysis.get('document_type', '')
            
            # author, recipientã®å–å¾—
            parties = objective_analysis.get('parties_mentioned', {})
            organizations = parties.get('organizations', [])
            author = organizations[0] if organizations else ''
            recipient = organizations[1] if len(organizations) > 1 else ''
        else:
            # æ—§æ§‹é€ : phase1_complete_analysisç›´ä¸‹ã¾ãŸã¯full_contentå†…
            old_analysis = phase1_analysis or evidence.get('full_content', {})
            document_type = old_analysis.get('document_type', '')
            author = old_analysis.get('author', '')
            recipient = old_analysis.get('recipient', '')
        
        return {
            **display_data,
            'document_type': document_type,
            'author': author,
            'recipient': recipient
        }
    
    def show_evidence_list(self, evidence_type: str = 'ko'):
        """è¨¼æ‹ åˆ†æä¸€è¦§ã‚’è¡¨ç¤º
        
        Args:
            evidence_type: è¨¼æ‹ ç¨®åˆ¥ ('ko' ã¾ãŸã¯ 'otsu')
        """
        type_name = "ç”²å·è¨¼" if evidence_type == 'ko' else "ä¹™å·è¨¼"
        
        print("\n" + "="*70)
        print(f"  è¨¼æ‹ åˆ†æä¸€è¦§ [{type_name}]")
        print("="*70)
        
        # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã‚’èª­ã¿è¾¼ã¿
        database = self.db_manager.load_database()
        evidence_list = database.get('evidence', [])
        
        if not evidence_list:
            print("\nâš ï¸  è¨¼æ‹ ãŒç™»éŒ²ã•ã‚Œã¦ã„ã¾ã›ã‚“")
            return
        
        # è¨¼æ‹ ç¨®åˆ¥ã§ãƒ•ã‚£ãƒ«ã‚¿ãƒ¼
        filtered_evidence = [
            e for e in evidence_list 
            if e.get('evidence_type', 'ko') == evidence_type
        ]
        
        if not filtered_evidence:
            print(f"\nâš ï¸  {type_name}ã®è¨¼æ‹ ãŒç™»éŒ²ã•ã‚Œã¦ã„ã¾ã›ã‚“")
            return
        
        # ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹åˆ¥ã«åˆ†é¡
        confirmed_evidence = []      # ç¢ºå®šæ¸ˆã¿ï¼ˆç”²å·è¨¼/ä¹™å·è¨¼ï¼‰
        pending_evidence = []        # æ•´ç†æ¸ˆã¿_æœªç¢ºå®š
        unclassified_evidence = []   # æœªåˆ†é¡
        
        for evidence in filtered_evidence:
            evidence_id = evidence.get('evidence_id', '')
            evidence_number = evidence.get('evidence_number', '')
            
            # æ•´ç†çŠ¶æ…‹ã‚’åˆ¤å®šï¼ˆå„ªå…ˆé †ä½ãŒé‡è¦ï¼‰
            # 1. æ•´ç†æ¸ˆã¿_æœªç¢ºå®š: evidence_idãŒ "tmp_" ã§å§‹ã¾ã‚‹ï¼ˆæœ€å„ªå…ˆï¼‰
            if evidence_id.startswith('tmp_'):
                pending_evidence.append(evidence)
            # 2. ç¢ºå®šæ¸ˆã¿: evidence_numberãŒ "ç”²1" ã‚„ "ä¹™2" ã®ã‚ˆã†ãªå½¢å¼
            #    ï¼ˆ"ç”²tmp" ã‚„ "ä¹™tmp" ã§å§‹ã¾ã‚‰ãªã„ã€ã‹ã¤è¨¼æ‹ ç•ªå·ãŒå­˜åœ¨ã™ã‚‹ï¼‰
            elif evidence_number and not evidence_number.startswith('ç”²tmp') and not evidence_number.startswith('ä¹™tmp'):
                confirmed_evidence.append(evidence)
            # 3. æœªåˆ†é¡: ãã®ä»–
            else:
                unclassified_evidence.append(evidence)
        
        # ç¢ºå®šæ¸ˆã¿è¨¼æ‹ ã®è¡¨ç¤º
        if confirmed_evidence:
            print(f"\nã€ç¢ºå®šæ¸ˆã¿ï¼ˆ{type_name}ï¼‰ã€‘")
            print("-"*70)
            for evidence in sorted(confirmed_evidence, key=lambda x: x.get('evidence_id', '')):
                evidence_id = evidence.get('evidence_id', 'ä¸æ˜')
                temp_id = evidence.get('temp_id', '')
                
                # ãƒ‡ãƒ¼ã‚¿æ§‹é€ ã®é•ã„ã‚’å¸åã—ã¦è¡¨ç¤ºç”¨ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
                display_data = self._get_evidence_display_data(evidence)
                
                print(f"  {evidence_id:10} | {display_data['creation_date']:12} | {display_data['analysis_status']:12} | {display_data['file_name']}")
                if temp_id:
                    print(f"             (å…ƒID: {temp_id})")
        
        # æ•´ç†æ¸ˆã¿_æœªç¢ºå®šè¨¼æ‹ ã®è¡¨ç¤º
        if pending_evidence:
            print("\nã€æ•´ç†æ¸ˆã¿_æœªç¢ºå®šã€‘")
            print("-"*70)
            for evidence in sorted(pending_evidence, key=lambda x: x.get('temp_id', '')):
                temp_id = evidence.get('temp_id', 'ä¸æ˜')
                
                # ãƒ‡ãƒ¼ã‚¿æ§‹é€ ã®é•ã„ã‚’å¸åã—ã¦è¡¨ç¤ºç”¨ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
                display_data = self._get_evidence_display_data(evidence)
                
                print(f"  {temp_id:10} | {display_data['creation_date']:12} | {display_data['analysis_status']:12} | {display_data['file_name']}")
        
        # æœªåˆ†é¡è¨¼æ‹ ã®è¡¨ç¤º
        if unclassified_evidence:
            print("\nã€æœªåˆ†é¡ã€‘")
            print("-"*70)
            for evidence in unclassified_evidence:
                file_name = evidence.get('file_name', 'ä¸æ˜')
                temp_id = evidence.get('temp_id', '')
                evidence_id = evidence.get('evidence_id', '')
                display_id = evidence_id or temp_id or 'ä¸æ˜'
                
                print(f"  {display_id:10} | {file_name}")
        
        # ã‚µãƒãƒªãƒ¼è¡¨ç¤º
        print("\n" + "="*70)
        print(f"  {type_name}ã®åˆè¨ˆ: {len(filtered_evidence)}ä»¶")
        print(f"    ç¢ºå®šæ¸ˆã¿: {len(confirmed_evidence)}ä»¶")
        print(f"    æ•´ç†æ¸ˆã¿_æœªç¢ºå®š: {len(pending_evidence)}ä»¶")
        print(f"    æœªåˆ†é¡: {len(unclassified_evidence)}ä»¶")
        print("="*70)
    
    def export_evidence_list(self, evidence_type: str = 'ko'):
        """è¨¼æ‹ ä¸€è¦§ã‚’CSV/Excelå½¢å¼ã§ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆ
        
        Args:
            evidence_type: è¨¼æ‹ ç¨®åˆ¥ ('ko' ã¾ãŸã¯ 'otsu')
        """
        type_name = "ç”²å·è¨¼" if evidence_type == 'ko' else "ä¹™å·è¨¼"
        
        print("\n" + "="*70)
        print(f"  è¨¼æ‹ ä¸€è¦§ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆ [{type_name}]")
        print("="*70)
        
        # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã‚’èª­ã¿è¾¼ã¿
        database = self.db_manager.load_database()
        evidence_list = database.get('evidence', [])
        
        if not evidence_list:
            print("\nâš ï¸  è¨¼æ‹ ãŒç™»éŒ²ã•ã‚Œã¦ã„ã¾ã›ã‚“")
            return
        
        # è¨¼æ‹ ç¨®åˆ¥ã§ãƒ•ã‚£ãƒ«ã‚¿ãƒ¼
        filtered_evidence = [
            e for e in evidence_list 
            if e.get('evidence_type', 'ko') == evidence_type
        ]
        
        if not filtered_evidence:
            print(f"\nâš ï¸  {type_name}ã®è¨¼æ‹ ãŒç™»éŒ²ã•ã‚Œã¦ã„ã¾ã›ã‚“")
            return
        
        # ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆãƒ¢ãƒ¼ãƒ‰é¸æŠ
        print("\nã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆãƒ¢ãƒ¼ãƒ‰ã‚’é¸æŠã—ã¦ãã ã•ã„:")
        print("  1. æ¨™æº–ãƒ¢ãƒ¼ãƒ‰ï¼ˆä¸»è¦ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã®ã¿ã€ç·¨é›†ã—ã‚„ã™ã„ï¼‰")
        print("  2. æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰ï¼ˆAIåˆ†æãƒ»OCRçµæœã‚’å€‹åˆ¥åˆ—ã«å±•é–‹ã€è‡ªç„¶è¨€èªç·¨é›†å¯èƒ½ï¼‰â† æ¨å¥¨")
        print("  3. å®Œå…¨ãƒ¢ãƒ¼ãƒ‰ï¼ˆdatabase.jsonå…¨ãƒ‡ãƒ¼ã‚¿ã‚’JSONåˆ—ã§å‡ºåŠ›ã€ä¸Šç´šè€…å‘ã‘ï¼‰")
        print("  4. ã‚­ãƒ£ãƒ³ã‚»ãƒ«")
        
        mode_choice = input("\n> ").strip()
        
        if mode_choice == '4' or not mode_choice:
            print("ã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
            return
        
        extended_mode = (mode_choice == '2')
        full_data = (mode_choice == '3')
        
        # å‡ºåŠ›å½¢å¼ã‚’é¸æŠ
        print("\nå‡ºåŠ›å½¢å¼ã‚’é¸æŠã—ã¦ãã ã•ã„:")
        print("  1. CSVå½¢å¼")
        print("  2. Excelå½¢å¼ (.xlsx)")
        print("  3. ã‚­ãƒ£ãƒ³ã‚»ãƒ«")
        
        format_choice = input("\n> ").strip()
        
        if format_choice == '3' or not format_choice:
            print("ã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
            return
        
        # å‡ºåŠ›ãƒ•ã‚¡ã‚¤ãƒ«åã‚’ç”Ÿæˆ
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        case_name = self.current_case.get('case_name', 'unknown').replace(' ', '_').replace('/', '_')
        type_suffix = "ko" if evidence_type == 'ko' else "otsu"
        
        if extended_mode:
            mode_suffix = "_extended"
        elif full_data:
            mode_suffix = "_full"
        else:
            mode_suffix = ""
        
        if format_choice == '1':
            # CSVå½¢å¼
            filename = f"evidence_list_{case_name}_{type_suffix}{mode_suffix}_{timestamp}.csv"
            self._export_to_csv(filtered_evidence, filename, evidence_type, full_data=full_data, extended_mode=extended_mode)
        elif format_choice == '2':
            # Excelå½¢å¼
            filename = f"evidence_list_{case_name}_{type_suffix}{mode_suffix}_{timestamp}.xlsx"
            self._export_to_excel(filtered_evidence, filename, evidence_type, full_data=full_data, extended_mode=extended_mode)
        else:
            print("ç„¡åŠ¹ãªé¸æŠã§ã™")
            return
    
    def _apply_extended_fields(self, evidence: Dict, extended_data: Dict):
        """æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰ã®è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’database.jsonæ§‹é€ ã«åæ˜ 
        
        Args:
            evidence: æ›´æ–°å¯¾è±¡ã®è¨¼æ‹ ã‚ªãƒ–ã‚¸ã‚§ã‚¯ãƒˆ
            extended_data: CSVã‹ã‚‰èª­ã¿è¾¼ã‚“ã æ‹¡å¼µãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã®è¾æ›¸
        """
        # complete_metadataæ§‹é€ ã‚’ç¢ºä¿
        if 'complete_metadata' not in evidence:
            evidence['complete_metadata'] = {}
        metadata = evidence['complete_metadata']
        
        if 'basic' not in metadata:
            metadata['basic'] = {}
        basic = metadata['basic']
        
        if 'format_specs' not in metadata:
            metadata['format_specs'] = {}
        format_spec = metadata['format_specs']
        
        # phase1_complete_analysisæ§‹é€ ã‚’ç¢ºä¿
        if 'phase1_complete_analysis' not in evidence:
            evidence['phase1_complete_analysis'] = {}
        analysis = evidence['phase1_complete_analysis']
        
        if 'ai_analysis' not in analysis:
            analysis['ai_analysis'] = {}
        ai_analysis = analysis['ai_analysis']
        
        if 'file_processing_result' not in analysis:
            analysis['file_processing_result'] = {}
        if 'content' not in analysis['file_processing_result']:
            analysis['file_processing_result']['content'] = {}
        file_content = analysis['file_processing_result']['content']
        
        # ãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰
        if 'ãƒ•ã‚¡ã‚¤ãƒ«å' in extended_data:
            basic['file_name'] = extended_data['ãƒ•ã‚¡ã‚¤ãƒ«å']
            evidence['original_filename'] = extended_data['ãƒ•ã‚¡ã‚¤ãƒ«å']
        
        if 'ãƒ•ã‚¡ã‚¤ãƒ«ã‚µã‚¤ã‚º' in extended_data:
            basic['file_size_human'] = extended_data['ãƒ•ã‚¡ã‚¤ãƒ«ã‚µã‚¤ã‚º']
        
        if 'ãƒšãƒ¼ã‚¸æ•°' in extended_data:
            try:
                format_spec['page_count'] = int(extended_data['ãƒšãƒ¼ã‚¸æ•°']) if extended_data['ãƒšãƒ¼ã‚¸æ•°'] else 0
            except (ValueError, TypeError):
                pass
        
        # è¨¼æ‹ ãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿
        if 'evidence_metadata' not in ai_analysis:
            ai_analysis['evidence_metadata'] = {}
        ev_meta = ai_analysis['evidence_metadata']
        
        if 'è¨¼æ‹ ç¨®åˆ¥_è©³ç´°' in extended_data:
            ev_meta['evidence_type'] = extended_data['è¨¼æ‹ ç¨®åˆ¥_è©³ç´°']
        
        if 'ãƒ•ã‚©ãƒ¼ãƒãƒƒãƒˆèª¬æ˜' in extended_data:
            ev_meta['format_description'] = extended_data['ãƒ•ã‚©ãƒ¼ãƒãƒƒãƒˆèª¬æ˜']
        
        # å®¢è¦³çš„åˆ†æ
        if 'objective_analysis' not in ai_analysis:
            ai_analysis['objective_analysis'] = {}
        obj = ai_analysis['objective_analysis']
        
        if 'æ–‡æ›¸ç¨®åˆ¥' in extended_data:
            obj['document_type'] = extended_data['æ–‡æ›¸ç¨®åˆ¥']
        
        # æ™‚é–“æƒ…å ±
        if 'temporal_information' not in obj:
            obj['temporal_information'] = {}
        temporal = obj['temporal_information']
        
        if 'ä½œæˆæ—¥' in extended_data:
            temporal['document_date'] = extended_data['ä½œæˆæ—¥']
        
        if 'æ™‚é–“çš„ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ' in extended_data:
            temporal['temporal_context'] = extended_data['æ™‚é–“çš„ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ']
        
        # é–¢ä¿‚è€…æƒ…å ±
        if 'parties_mentioned' not in obj:
            obj['parties_mentioned'] = {}
        parties = obj['parties_mentioned']
        
        # çµ„ç¹”æƒ…å ±ï¼ˆæœ€å¤§3ä»¶ï¼‰
        if 'organizations' not in parties:
            parties['organizations'] = []
        organizations = parties['organizations']
        
        for i in range(3):
            prefix = f'çµ„ç¹”{i+1}_'
            name = extended_data.get(f'{prefix}åç§°', '').strip()
            role = extended_data.get(f'{prefix}å½¹å‰²', '').strip()
            context = extended_data.get(f'{prefix}ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ', '').strip()
            
            if name or role or context:
                org_dict = {}
                if name:
                    org_dict['name'] = name
                if role:
                    org_dict['role'] = role
                if context:
                    org_dict['context'] = context
                
                # ã‚¤ãƒ³ãƒ‡ãƒƒã‚¯ã‚¹ã«å¿œã˜ã¦è¿½åŠ ã¾ãŸã¯æ›´æ–°
                if i < len(organizations):
                    if isinstance(organizations[i], dict):
                        organizations[i].update(org_dict)
                    else:
                        organizations[i] = org_dict
                else:
                    organizations.append(org_dict)
        
        # å€‹äººæƒ…å ±ï¼ˆæœ€å¤§2ä»¶ï¼‰
        if 'individuals' not in parties:
            parties['individuals'] = []
        individuals = parties['individuals']
        
        for i in range(2):
            prefix = f'å€‹äºº{i+1}_'
            name = extended_data.get(f'{prefix}åå‰', '').strip()
            role = extended_data.get(f'{prefix}å½¹å‰²', '').strip()
            context = extended_data.get(f'{prefix}ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ', '').strip()
            
            if name or role or context:
                ind_dict = {}
                if name:
                    ind_dict['name'] = name
                if role:
                    ind_dict['role'] = role
                if context:
                    ind_dict['context'] = context
                
                if i < len(individuals):
                    if isinstance(individuals[i], dict):
                        individuals[i].update(ind_dict)
                    else:
                        individuals[i] = ind_dict
                else:
                    individuals.append(ind_dict)
        
        # è¦–è¦šè¦ç´ 
        if 'visual_elements' not in obj:
            obj['visual_elements'] = {}
        visual = obj['visual_elements']
        
        if 'ãƒ¬ã‚¤ã‚¢ã‚¦ãƒˆèª¬æ˜' in extended_data:
            visual['layout_description'] = extended_data['ãƒ¬ã‚¤ã‚¢ã‚¦ãƒˆèª¬æ˜']
        
        if 'ãƒ†ã‚­ã‚¹ãƒˆå†…å®¹è¦ç´„' in extended_data:
            visual['text_content_summary'] = extended_data['ãƒ†ã‚­ã‚¹ãƒˆå†…å®¹è¦ç´„']
        
        if 'æ³¨ç›®ã™ã¹ãç‰¹å¾´' in extended_data:
            visual['notable_features'] = extended_data['æ³¨ç›®ã™ã¹ãç‰¹å¾´']
        
        # å®Œå…¨å†…å®¹
        if 'full_content' not in ai_analysis:
            ai_analysis['full_content'] = {}
        full_content = ai_analysis['full_content']
        
        if 'å®Œå…¨ãªèª¬æ˜' in extended_data:
            full_content['complete_description'] = extended_data['å®Œå…¨ãªèª¬æ˜']
        
        if 'è©³ç´°å†…å®¹' in extended_data:
            full_content['detailed_content'] = extended_data['è©³ç´°å†…å®¹']
        
        # æ³•çš„é‡è¦æ€§
        if 'legal_significance' not in ai_analysis:
            ai_analysis['legal_significance'] = {}
        legal = ai_analysis['legal_significance']
        
        if 'å®¢è¦³çš„äº‹å®Ÿ' in extended_data:
            legal['objective_facts'] = extended_data['å®¢è¦³çš„äº‹å®Ÿ']
        
        if 'æ–‡è„ˆä¸Šã®é‡è¦æ€§' in extended_data:
            legal['contextual_importance'] = extended_data['æ–‡è„ˆä¸Šã®é‡è¦æ€§']
        
        # è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿï¼ˆæœ€å¤§5ä»¶ï¼‰
        if 'provable_facts' not in legal:
            legal['provable_facts'] = []
        provable_facts = legal['provable_facts']
        
        for i in range(1, 6):
            fact = extended_data.get(f'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ{i}', '').strip()
            if fact:
                if i-1 < len(provable_facts):
                    provable_facts[i-1] = fact
                else:
                    provable_facts.append(fact)
        
        # ä½¿ç”¨ææ¡ˆ
        if 'usage_suggestions' not in ai_analysis:
            ai_analysis['usage_suggestions'] = {}
        usage = ai_analysis['usage_suggestions']
        
        if 'æ¨å¥¨ã•ã‚Œã‚‹ä½¿ç”¨æ–¹æ³•' in extended_data:
            usage['recommended_usage'] = extended_data['æ¨å¥¨ã•ã‚Œã‚‹ä½¿ç”¨æ–¹æ³•']
        
        # OCRçµæœ
        if 'ocr' not in file_content:
            file_content['ocr'] = {}
        
        if 'OCRãƒ†ã‚­ã‚¹ãƒˆ' in extended_data:
            file_content['ocr']['text'] = extended_data['OCRãƒ†ã‚­ã‚¹ãƒˆ']
        
        # å“è³ªã‚¹ã‚³ã‚¢
        if 'å®Œå…¨è¨€èªåŒ–ãƒ¬ãƒ™ãƒ«' in extended_data:
            try:
                ai_analysis['verbalization_level'] = int(extended_data['å®Œå…¨è¨€èªåŒ–ãƒ¬ãƒ™ãƒ«']) if extended_data['å®Œå…¨è¨€èªåŒ–ãƒ¬ãƒ™ãƒ«'] else 0
            except (ValueError, TypeError):
                pass
        
        if 'ä¿¡é ¼åº¦ã‚¹ã‚³ã‚¢' in extended_data:
            score_str = extended_data['ä¿¡é ¼åº¦ã‚¹ã‚³ã‚¢'].strip().rstrip('%')
            try:
                ai_analysis['confidence_score'] = float(score_str) / 100.0 if score_str else 0.0
            except (ValueError, TypeError):
                pass
    
    def _extract_extended_fields(self, evidence: Dict) -> Dict:
        """æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰ç”¨ã«è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’æŠ½å‡º
        
        Args:
            evidence: è¨¼æ‹ ãƒ‡ãƒ¼ã‚¿
            
        Returns:
            æ‹¡å¼µãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã®è¾æ›¸
        """
        extended = {}
        
        # ãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿
        metadata = evidence.get('complete_metadata', {})
        basic = metadata.get('basic', {})
        format_spec = metadata.get('format_specific', {})
        
        extended['ãƒ•ã‚¡ã‚¤ãƒ«å'] = basic.get('file_name', '')
        extended['ãƒ•ã‚¡ã‚¤ãƒ«ã‚µã‚¤ã‚º'] = basic.get('file_size_human', '')
        extended['ãƒšãƒ¼ã‚¸æ•°'] = format_spec.get('page_count', '') if isinstance(format_spec, dict) else ''
        
        # AIåˆ†æçµæœ
        analysis = evidence.get('phase1_complete_analysis', {})
        ai_analysis = analysis.get('ai_analysis', {})
        
        # è¨¼æ‹ ãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿
        ev_meta = ai_analysis.get('evidence_metadata', {})
        extended['è¨¼æ‹ ç¨®åˆ¥_è©³ç´°'] = ev_meta.get('evidence_type', '')
        extended['ãƒ•ã‚©ãƒ¼ãƒãƒƒãƒˆèª¬æ˜'] = ev_meta.get('format_description', '')
        
        # å®¢è¦³çš„åˆ†æ
        obj = ai_analysis.get('objective_analysis', {})
        extended['æ–‡æ›¸ç¨®åˆ¥'] = obj.get('document_type', '')
        
        # æ™‚é–“æƒ…å ±
        temporal = obj.get('temporal_information', {})
        extended['ä½œæˆæ—¥'] = temporal.get('document_date', '')
        extended['æ™‚é–“çš„ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ'] = temporal.get('temporal_context', '')
        
        # é–¢ä¿‚è€…æƒ…å ±
        parties = obj.get('parties_mentioned', {})
        organizations = parties.get('organizations', [])
        individuals = parties.get('individuals', [])
        
        # çµ„ç¹”ï¼ˆæœ€å¤§3ä»¶ï¼‰
        for i in range(3):
            prefix = f'çµ„ç¹”{i+1}_'
            if i < len(organizations):
                org = organizations[i]
                if isinstance(org, dict):
                    extended[f'{prefix}åç§°'] = org.get('name', '')
                    extended[f'{prefix}å½¹å‰²'] = org.get('role', '')
                    extended[f'{prefix}ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ'] = org.get('context', '')
                else:
                    extended[f'{prefix}åç§°'] = ''
                    extended[f'{prefix}å½¹å‰²'] = ''
                    extended[f'{prefix}ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ'] = ''
            else:
                extended[f'{prefix}åç§°'] = ''
                extended[f'{prefix}å½¹å‰²'] = ''
                extended[f'{prefix}ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ'] = ''
        
        # å€‹äººï¼ˆæœ€å¤§2ä»¶ï¼‰
        for i in range(2):
            prefix = f'å€‹äºº{i+1}_'
            if i < len(individuals):
                person = individuals[i]
                if isinstance(person, dict):
                    extended[f'{prefix}åå‰'] = person.get('name', '')
                    extended[f'{prefix}å½¹å‰²'] = person.get('role', '')
                    extended[f'{prefix}ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ'] = person.get('context', '')
                else:
                    extended[f'{prefix}åå‰'] = ''
                    extended[f'{prefix}å½¹å‰²'] = ''
                    extended[f'{prefix}ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ'] = ''
            else:
                extended[f'{prefix}åå‰'] = ''
                extended[f'{prefix}å½¹å‰²'] = ''
                extended[f'{prefix}ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ'] = ''
        
        # è¦–è¦šè¦ç´ 
        visual = obj.get('visual_elements', {})
        extended['ãƒ¬ã‚¤ã‚¢ã‚¦ãƒˆèª¬æ˜'] = visual.get('layout_description', '')
        extended['ãƒ†ã‚­ã‚¹ãƒˆå†…å®¹è¦ç´„'] = visual.get('text_content_summary', '')
        notable_features = visual.get('notable_features', [])
        extended['æ³¨ç›®ã™ã¹ãç‰¹å¾´'] = ', '.join(notable_features) if isinstance(notable_features, list) else str(notable_features)
        
        # å®Œå…¨å†…å®¹
        full_content = ai_analysis.get('full_content', {})
        extended['å®Œå…¨ãªèª¬æ˜'] = full_content.get('complete_description', '')
        extended['è©³ç´°å†…å®¹'] = full_content.get('detailed_content', '')
        
        # æ³•çš„é‡è¦æ€§
        legal = ai_analysis.get('legal_significance', {})
        extended['å®¢è¦³çš„äº‹å®Ÿ'] = legal.get('objective_facts', '')
        extended['æ–‡è„ˆä¸Šã®é‡è¦æ€§'] = legal.get('contextual_importance', '')
        
        # é–¢é€£äº‹å®Ÿï¼ˆæœ€å¤§5ä»¶ï¼‰
        # è¤‡æ•°ã®å¯èƒ½ãªãƒ‘ã‚¹ã‚’è©¦è¡Œ
        provable_facts = []
        
        # ãƒ‘ã‚¹1: related_facts.provable_facts
        related = ai_analysis.get('related_facts', {})
        provable_facts = related.get('provable_facts', [])
        
        # ãƒ‘ã‚¹2: legal_significance.provable_factsï¼ˆåˆ¥ã®å¯èƒ½ãªå ´æ‰€ï¼‰
        if not provable_facts:
            provable_facts = legal.get('provable_facts', [])
        
        for i in range(5):
            if i < len(provable_facts):
                fact = provable_facts[i]
                extended[f'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ{i+1}'] = fact if isinstance(fact, str) else ''
            else:
                extended[f'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ{i+1}'] = ''
        
        # ä½¿ç”¨ææ¡ˆ
        usage = ai_analysis.get('usage_suggestions', {})
        extended['æ¨å¥¨ã•ã‚Œã‚‹ä½¿ç”¨æ–¹æ³•'] = usage.get('recommended_usage', '')
        
        # OCRçµæœï¼ˆå®Ÿéš›ã«èª­ã¿å–ã‚‰ã‚ŒãŸæ–‡å­—åˆ—ã‚’å–å¾—ï¼‰
        ocr_text = ''
        
        # ãƒ‘ã‚¹1: ai_analysis.full_content.ocr_results.extracted_textï¼ˆVision APIã®OCRçµæœï¼‰â† æœ€å„ªå…ˆ
        ocr_results = full_content.get('ocr_results', {})
        ocr_text = ocr_results.get('extracted_text', '')
        
        # ãƒ‘ã‚¹2: ai_analysis.full_content.textual_content.extracted_textï¼ˆä»£æ›¿ãƒ‘ã‚¹ï¼‰
        if not ocr_text or ocr_text.strip() == '':
            textual_content = full_content.get('textual_content', {})
            ocr_text = textual_content.get('extracted_text', '')
        
        # ãƒ‘ã‚¹3: file_processing_result.content.total_textï¼ˆPDFç›´æ¥æŠ½å‡ºï¼‰
        if not ocr_text or ocr_text.strip() == '':
            file_content = analysis.get('file_processing_result', {}).get('content', {})
            total_text = file_content.get('total_text', '')
            # åˆ¶å¾¡æ–‡å­—ã®ã¿ã§ãªã„ã‹ç¢ºèª
            if total_text and not (len(total_text) > 0 and '\u0001' in total_text[:20]):
                ocr_text = total_text
        
        # ãƒ‘ã‚¹4: file_processing_result.content.pagesï¼ˆå„ãƒšãƒ¼ã‚¸ã®ãƒ†ã‚­ã‚¹ãƒˆã‚’çµåˆï¼‰
        if not ocr_text or ocr_text.strip() == '':
            file_content = analysis.get('file_processing_result', {}).get('content', {})
            pages = file_content.get('pages', [])
            if pages:
                page_texts = []
                for page in pages:
                    page_text = page.get('text', '')
                    # åˆ¶å¾¡æ–‡å­—ã®ã¿ã®å ´åˆã¯ã‚¹ã‚­ãƒƒãƒ—
                    if page_text and not all(c in '\u0001\n ' for c in page_text[:50]):
                        page_texts.append(f"--- ãƒšãƒ¼ã‚¸{page.get('page_number', 0)} ---\n{page_text}")
                if page_texts:
                    ocr_text = '\n\n'.join(page_texts)
        
        # ãƒ‘ã‚¹5ä»¥é™: æ—§å½¢å¼ã®æ§˜ã€…ãªãƒ‘ã‚¹
        if not ocr_text or ocr_text.strip() == '':
            file_content = analysis.get('file_processing_result', {}).get('content', {})
            ocr_text = file_content.get('text', '') or \
                      analysis.get('file_processing_result', {}).get('ocr_text', '') or \
                      evidence.get('complete_metadata', {}).get('ocr_text', '')
        
        extended['OCRãƒ†ã‚­ã‚¹ãƒˆ'] = ocr_text
        
        # å“è³ªã‚¹ã‚³ã‚¢
        extended['å®Œå…¨è¨€èªåŒ–ãƒ¬ãƒ™ãƒ«'] = ai_analysis.get('verbalization_level', 0)
        extended['ä¿¡é ¼åº¦ã‚¹ã‚³ã‚¢'] = f"{ai_analysis.get('confidence_score', 0.0):.1%}"
        
        return extended
    
    def _export_to_csv(self, evidence_list: List[Dict], filename: str, evidence_type: str = 'ko', full_data: bool = False, extended_mode: bool = False):
        """CSVå½¢å¼ã§ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆ
        
        Args:
            evidence_list: è¨¼æ‹ ãƒªã‚¹ãƒˆ
            filename: å‡ºåŠ›ãƒ•ã‚¡ã‚¤ãƒ«å
            evidence_type: è¨¼æ‹ ç¨®åˆ¥ ('ko' ã¾ãŸã¯ 'otsu')
            full_data: True ã®å ´åˆã€database.jsonã®å…¨ãƒ‡ãƒ¼ã‚¿ã‚’JSONåˆ—ã¨ã—ã¦å‡ºåŠ›
            extended_mode: True ã®å ´åˆã€è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’å€‹åˆ¥åˆ—ã«å±•é–‹ï¼ˆæ‹¡å¼µãƒ¢ãƒ¼ãƒ‰ï¼‰
        """
        import csv
        
        try:
            output_path = os.path.join(os.getcwd(), filename)
            
            with open(output_path, 'w', encoding='utf-8-sig', newline='') as csvfile:
                if extended_mode:
                    # æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰: è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’å€‹åˆ¥åˆ—ã«å±•é–‹
                    fieldnames = [
                        # åŸºæœ¬æƒ…å ±
                        'è¨¼æ‹ ç¨®åˆ¥', 'ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹', 'è¨¼æ‹ ç•ªå·', 'ä»®ç•ªå·', 'åˆ†æçŠ¶æ…‹',
                        'Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID',
                        
                        # ãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿
                        'ãƒ•ã‚¡ã‚¤ãƒ«å', 'ãƒ•ã‚¡ã‚¤ãƒ«ã‚µã‚¤ã‚º', 'ãƒšãƒ¼ã‚¸æ•°',
                        
                        # è¨¼æ‹ ãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿
                        'è¨¼æ‹ ç¨®åˆ¥_è©³ç´°', 'ãƒ•ã‚©ãƒ¼ãƒãƒƒãƒˆèª¬æ˜',
                        
                        # å®¢è¦³çš„åˆ†æ
                        'æ–‡æ›¸ç¨®åˆ¥', 'ä½œæˆæ—¥', 'æ™‚é–“çš„ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        
                        # é–¢ä¿‚è€…æƒ…å ±ï¼ˆæœ€å¤§5çµ„ç¹”ã€5å€‹äººï¼‰
                        'çµ„ç¹”1_åç§°', 'çµ„ç¹”1_å½¹å‰²', 'çµ„ç¹”1_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'çµ„ç¹”2_åç§°', 'çµ„ç¹”2_å½¹å‰²', 'çµ„ç¹”2_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'çµ„ç¹”3_åç§°', 'çµ„ç¹”3_å½¹å‰²', 'çµ„ç¹”3_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'å€‹äºº1_åå‰', 'å€‹äºº1_å½¹å‰²', 'å€‹äºº1_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'å€‹äºº2_åå‰', 'å€‹äºº2_å½¹å‰²', 'å€‹äºº2_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        
                        # è¦–è¦šè¦ç´ 
                        'ãƒ¬ã‚¤ã‚¢ã‚¦ãƒˆèª¬æ˜', 'ãƒ†ã‚­ã‚¹ãƒˆå†…å®¹è¦ç´„', 'æ³¨ç›®ã™ã¹ãç‰¹å¾´',
                        
                        # å®Œå…¨å†…å®¹
                        'å®Œå…¨ãªèª¬æ˜', 'è©³ç´°å†…å®¹',
                        
                        # æ³•çš„é‡è¦æ€§
                        'å®¢è¦³çš„äº‹å®Ÿ', 'æ–‡è„ˆä¸Šã®é‡è¦æ€§',
                        
                        # é–¢é€£äº‹å®Ÿï¼ˆæœ€å¤§10ä»¶ï¼‰
                        'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ1', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ2', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ3',
                        'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ4', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ5',
                        
                        # ä½¿ç”¨ææ¡ˆ
                        'æ¨å¥¨ã•ã‚Œã‚‹ä½¿ç”¨æ–¹æ³•',
                        
                        # OCRçµæœ
                        'OCRãƒ†ã‚­ã‚¹ãƒˆ',
                        
                        # å“è³ªã‚¹ã‚³ã‚¢
                        'å®Œå…¨è¨€èªåŒ–ãƒ¬ãƒ™ãƒ«', 'ä¿¡é ¼åº¦ã‚¹ã‚³ã‚¢'
                    ]
                elif full_data:
                    # å…¨ãƒ‡ãƒ¼ã‚¿ãƒ¢ãƒ¼ãƒ‰: ä¸»è¦ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ + JSONåˆ—
                    fieldnames = [
                        'è¨¼æ‹ ç¨®åˆ¥', 'ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹', 'è¨¼æ‹ ç•ªå·', 'ä»®ç•ªå·', 'ä½œæˆæ—¥', 
                        'åˆ†æçŠ¶æ…‹', 'ãƒ•ã‚¡ã‚¤ãƒ«å', 'æ–‡æ›¸ç¨®åˆ¥', 'ä½œæˆè€…',
                        'å®›å…ˆ', 'è¦ç´„', 'Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID',
                        'complete_metadata_json', 'phase1_complete_analysis_json'
                    ]
                else:
                    # ç°¡æ˜“ãƒ¢ãƒ¼ãƒ‰: ä¸»è¦ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã®ã¿
                    fieldnames = [
                        'è¨¼æ‹ ç¨®åˆ¥', 'ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹', 'è¨¼æ‹ ç•ªå·', 'ä»®ç•ªå·', 'ä½œæˆæ—¥', 
                        'åˆ†æçŠ¶æ…‹', 'ãƒ•ã‚¡ã‚¤ãƒ«å', 'æ–‡æ›¸ç¨®åˆ¥', 'ä½œæˆè€…',
                        'å®›å…ˆ', 'è¦ç´„', 'Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID'
                    ]
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                writer.writeheader()
                
                # ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹ã§ã‚½ãƒ¼ãƒˆï¼ˆç¢ºå®šæ¸ˆã¿â†’æ•´ç†æ¸ˆã¿_æœªç¢ºå®šâ†’æœªåˆ†é¡ï¼‰
                def get_organization_status(evidence):
                    """è¨¼æ‹ ã®æ•´ç†çŠ¶æ…‹ã‚’åˆ¤å®š"""
                    evidence_id = evidence.get('evidence_id', '')
                    evidence_number = evidence.get('evidence_number', '')
                    
                    if evidence_number and not evidence_number.startswith('ç”²tmp') and not evidence_number.startswith('ä¹™tmp'):
                        return 'ç¢ºå®šæ¸ˆã¿'
                    elif evidence_id.startswith('tmp_'):
                        return 'æ•´ç†æ¸ˆã¿_æœªç¢ºå®š'
                    else:
                        return 'æœªåˆ†é¡'
                
                status_order = {'ç¢ºå®šæ¸ˆã¿': 1, 'æ•´ç†æ¸ˆã¿_æœªç¢ºå®š': 2, 'æœªåˆ†é¡': 3}
                sorted_evidence = sorted(
                    evidence_list, 
                    key=lambda x: (
                        status_order.get(get_organization_status(x), 99),
                        x.get('evidence_id', ''),
                        x.get('temp_id', '')
                    )
                )
                
                type_name = "ç”²å·è¨¼" if evidence_type == 'ko' else "ä¹™å·è¨¼"
                
                for evidence in sorted_evidence:
                    status = get_organization_status(evidence)
                    evidence_id = evidence.get('evidence_id', '')
                    temp_id = evidence.get('temp_id', '')
                    
                    # ãƒ‡ãƒ¼ã‚¿æ§‹é€ ã®é•ã„ã‚’å¸åã—ã¦ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆç”¨ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
                    export_data = self._get_evidence_export_data(evidence)
                    
                    gdrive_file_id = evidence.get('gdrive_file_id', '') or evidence.get('complete_metadata', {}).get('gdrive', {}).get('file_id', '')
                    
                    # åˆ†æçŠ¶æ…‹
                    analysis_status = "åˆ†ææ¸ˆã¿" if export_data['complete_description'] else "æœªåˆ†æ"
                    
                    summary = export_data['complete_description'] or ''
                    
                    # æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰ã®å ´åˆã€è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’å€‹åˆ¥åˆ—ã«å±•é–‹
                    if extended_mode:
                        # æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰å°‚ç”¨ã®åŸºæœ¬æƒ…å ±
                        row_data = {
                            'è¨¼æ‹ ç¨®åˆ¥': type_name,
                            'ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹': status,
                            'è¨¼æ‹ ç•ªå·': evidence_id,
                            'ä»®ç•ªå·': temp_id,
                            'åˆ†æçŠ¶æ…‹': analysis_status,
                            'Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID': gdrive_file_id
                        }
                        # æ‹¡å¼µãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’è¿½åŠ 
                        extended_data = self._extract_extended_fields(evidence)
                        row_data.update(extended_data)
                    
                    # å…¨ãƒ‡ãƒ¼ã‚¿ãƒ¢ãƒ¼ãƒ‰ã®å ´åˆã€æ¨™æº–ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ + JSONåˆ—
                    elif full_data:
                        row_data = {
                            'è¨¼æ‹ ç¨®åˆ¥': type_name,
                            'ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹': status,
                            'è¨¼æ‹ ç•ªå·': evidence_id,
                            'ä»®ç•ªå·': temp_id,
                            'ä½œæˆæ—¥': export_data['creation_date'],
                            'åˆ†æçŠ¶æ…‹': analysis_status,
                            'ãƒ•ã‚¡ã‚¤ãƒ«å': export_data['file_name'],
                            'æ–‡æ›¸ç¨®åˆ¥': export_data['document_type'],
                            'ä½œæˆè€…': export_data['author'],
                            'å®›å…ˆ': export_data['recipient'],
                            'è¦ç´„': summary[:100] + '...' if len(summary) > 100 else summary,
                            'Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID': gdrive_file_id
                        }
                        # complete_metadataã‚’JSONæ–‡å­—åˆ—ã¨ã—ã¦å‡ºåŠ›
                        complete_metadata = evidence.get('complete_metadata', {})
                        row_data['complete_metadata_json'] = json.dumps(complete_metadata, ensure_ascii=False) if complete_metadata else ''
                        
                        # phase1_complete_analysisã‚’JSONæ–‡å­—åˆ—ã¨ã—ã¦å‡ºåŠ›
                        phase1_analysis = evidence.get('phase1_complete_analysis', {})
                        row_data['phase1_complete_analysis_json'] = json.dumps(phase1_analysis, ensure_ascii=False) if phase1_analysis else ''
                    
                    # æ¨™æº–ãƒ¢ãƒ¼ãƒ‰
                    else:
                        row_data = {
                            'è¨¼æ‹ ç¨®åˆ¥': type_name,
                            'ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹': status,
                            'è¨¼æ‹ ç•ªå·': evidence_id,
                            'ä»®ç•ªå·': temp_id,
                            'ä½œæˆæ—¥': export_data['creation_date'],
                            'åˆ†æçŠ¶æ…‹': analysis_status,
                            'ãƒ•ã‚¡ã‚¤ãƒ«å': export_data['file_name'],
                            'æ–‡æ›¸ç¨®åˆ¥': export_data['document_type'],
                            'ä½œæˆè€…': export_data['author'],
                            'å®›å…ˆ': export_data['recipient'],
                            'è¦ç´„': summary[:100] + '...' if len(summary) > 100 else summary,
                            'Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID': gdrive_file_id
                        }
                    
                    writer.writerow(row_data)
            
            print(f"\nâœ… CSVå½¢å¼ã§ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆã—ã¾ã—ãŸ")
            print(f"   ãƒ•ã‚¡ã‚¤ãƒ«: {output_path}")
            print(f"   ä»¶æ•°: {len(evidence_list)}ä»¶")
            
            # Google Driveã¸ã®ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰ã‚’ç¢ºèª
            upload_choice = input("\nGoogle Driveã«ã‚‚ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰ã—ã¾ã™ã‹ï¼Ÿ (y/n): ").strip().lower()
            if upload_choice == 'y':
                gdrive_url = self._upload_export_file_to_gdrive(output_path, filename)
                if gdrive_url:
                    print(f"\nâœ… Google Driveã«ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰ã—ã¾ã—ãŸ")
                    print(f"   URL: {gdrive_url}")
            
        except Exception as e:
            print(f"\nâŒ ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆã«å¤±æ•—ã—ã¾ã—ãŸ: {e}")
            import traceback
            traceback.print_exc()
    
    def _upload_export_file_to_gdrive(self, local_path: str, filename: str) -> Optional[str]:
        """ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆãƒ•ã‚¡ã‚¤ãƒ«ã‚’Google Driveã«ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰
        
        Args:
            local_path: ãƒ­ãƒ¼ã‚«ãƒ«ãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹
            filename: ãƒ•ã‚¡ã‚¤ãƒ«å
        
        Returns:
            Google Driveã®URLï¼ˆå¤±æ•—æ™‚ã¯Noneï¼‰
        """
        try:
            service = self.case_manager.get_google_drive_service()
            if not service:
                print("âŒ Google Driveã‚µãƒ¼ãƒ“ã‚¹ã«æ¥ç¶šã§ãã¾ã›ã‚“")
                return None
            
            from googleapiclient.http import MediaFileUpload
            
            # ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆãƒ•ã‚©ãƒ«ãƒ€IDã‚’å–å¾—
            database = self.db_manager.load_database()
            export_folder_id = database.get('metadata', {}).get('export_folder_id')
            
            # ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆãƒ•ã‚©ãƒ«ãƒ€IDãŒãªã„å ´åˆã¯äº‹ä»¶ãƒ•ã‚©ãƒ«ãƒ€IDã«ãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯
            if not export_folder_id:
                case_folder_id = database.get('metadata', {}).get('case_folder_id')
                if not case_folder_id:
                    print("âŒ äº‹ä»¶ãƒ•ã‚©ãƒ«ãƒ€IDãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“")
                    return None
                print("âš ï¸  ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆãƒ•ã‚©ãƒ«ãƒ€IDãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã€‚äº‹ä»¶ãƒ•ã‚©ãƒ«ãƒ€ã«ä¿å­˜ã—ã¾ã™ã€‚")
                target_folder_id = case_folder_id
            else:
                target_folder_id = export_folder_id
            
            # ãƒ•ã‚¡ã‚¤ãƒ«ã‚¿ã‚¤ãƒ—ã‚’åˆ¤å®š
            file_ext = os.path.splitext(filename)[1].lower()
            if file_ext == '.csv':
                mime_type = 'text/csv'
            elif file_ext == '.xlsx':
                mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            else:
                mime_type = 'application/octet-stream'
            
            file_metadata = {
                'name': filename,
                'parents': [target_folder_id],
                'mimeType': mime_type
            }
            
            media = MediaFileUpload(local_path, mimetype=mime_type, resumable=True)
            
            file = service.files().create(
                body=file_metadata,
                media_body=media,
                supportsAllDrives=True,
                fields='id, name, webViewLink'
            ).execute()
            
            return file.get('webViewLink', f"https://drive.google.com/file/d/{file['id']}/view")
            
        except Exception as e:
            print(f"âŒ Google Driveã¸ã®ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰å¤±æ•—: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _download_file_from_gdrive_url(self, gdrive_url: str, output_path: str) -> bool:
        """Google Drive URLã‹ã‚‰ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰
        
        Args:
            gdrive_url: Google Driveã®URL
                - https://drive.google.com/file/d/FILE_ID/view
                - https://docs.google.com/spreadsheets/d/FILE_ID/edit
            output_path: ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰å…ˆãƒ‘ã‚¹
        
        Returns:
            æˆåŠŸæ™‚Trueã€å¤±æ•—æ™‚False
        """
        try:
            # URLã‹ã‚‰ãƒ•ã‚¡ã‚¤ãƒ«IDã‚’æŠ½å‡º
            import re
            
            # ãƒ‘ã‚¿ãƒ¼ãƒ³1: https://drive.google.com/file/d/FILE_ID/view
            match = re.search(r'/file/d/([a-zA-Z0-9_-]+)', gdrive_url)
            if not match:
                # ãƒ‘ã‚¿ãƒ¼ãƒ³2: https://docs.google.com/spreadsheets/d/FILE_ID/edit
                match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', gdrive_url)
            if not match:
                # ãƒ‘ã‚¿ãƒ¼ãƒ³3: https://drive.google.com/open?id=FILE_ID
                match = re.search(r'[?&]id=([a-zA-Z0-9_-]+)', gdrive_url)
            if not match:
                # ãƒ‘ã‚¿ãƒ¼ãƒ³4: æ±ç”¨çš„ãª /d/ ãƒ‘ã‚¿ãƒ¼ãƒ³
                match = re.search(r'/d/([a-zA-Z0-9_-]+)', gdrive_url)
            
            if not match:
                print(f"âŒ ç„¡åŠ¹ãªGoogle Drive URL: {gdrive_url}")
                return False
            
            file_id = match.group(1)
            print(f"ğŸ“¥ ãƒ•ã‚¡ã‚¤ãƒ«ID: {file_id}")
            
            service = self.case_manager.get_google_drive_service()
            if not service:
                print("âŒ Google Driveã‚µãƒ¼ãƒ“ã‚¹ã«æ¥ç¶šã§ãã¾ã›ã‚“")
                return False
            
            # ãƒ•ã‚¡ã‚¤ãƒ«æƒ…å ±ã‚’å–å¾—ã—ã¦MIMEã‚¿ã‚¤ãƒ—ã‚’ç¢ºèª
            import io
            from googleapiclient.http import MediaIoBaseDownload
            
            file_metadata = service.files().get(
                fileId=file_id,
                fields='mimeType, name',
                supportsAllDrives=True
            ).execute()
            
            mime_type = file_metadata.get('mimeType', '')
            file_name = file_metadata.get('name', 'unknown')
            
            print(f"   ãƒ•ã‚¡ã‚¤ãƒ«å: {file_name}")
            print(f"   ãƒ•ã‚¡ã‚¤ãƒ«ã‚¿ã‚¤ãƒ—: {mime_type}")
            
            # Google Spreadsheetsã®å ´åˆã¯Excelå½¢å¼ã§ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆ
            if mime_type == 'application/vnd.google-apps.spreadsheet':
                print("   ğŸ“Š Google Spreadsheetsã‚’æ¤œå‡º â†’ Excelå½¢å¼ã§ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆä¸­...")
                
                # Excelå½¢å¼ã§ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆ
                request = service.files().export_media(
                    fileId=file_id,
                    mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                
                done = False
                while not done:
                    status, done = downloader.next_chunk()
                    if status:
                        print(f"   ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰é€²è¡Œä¸­: {int(status.progress() * 100)}%")
                
                # ãƒ•ã‚¡ã‚¤ãƒ«ã«æ›¸ãè¾¼ã¿
                with open(output_path, 'wb') as f:
                    f.write(fh.getvalue())
                
                print(f"âœ… ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰å®Œäº†: {output_path}")
            else:
                # é€šå¸¸ã®ãƒ•ã‚¡ã‚¤ãƒ«ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰
                request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                
                done = False
                while not done:
                    status, done = downloader.next_chunk()
                    if status:
                        print(f"   ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰é€²è¡Œä¸­: {int(status.progress() * 100)}%")
                
                # ãƒ•ã‚¡ã‚¤ãƒ«ã«æ›¸ãè¾¼ã¿
                with open(output_path, 'wb') as f:
                    f.write(fh.getvalue())
                
                print(f"âœ… ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰å®Œäº†: {output_path}")
            return True
            
        except Exception as e:
            print(f"âŒ Google Driveã‹ã‚‰ã®ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰å¤±æ•—: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def import_csv_updates(self):
        """CSVç·¨é›†å†…å®¹ã‚’database.jsonã«åæ˜ """
        print("\n" + "="*70)
        print("  CSVç·¨é›†å†…å®¹ã‚’database.jsonã«åæ˜ ")
        print("="*70)
        print("\nç·¨é›†ã—ãŸCSVãƒ•ã‚¡ã‚¤ãƒ«ã‹ã‚‰è¨¼æ‹ æƒ…å ±ã‚’æ›´æ–°ã—ã¾ã™")
        print("ç·¨é›†å¯èƒ½ãªé …ç›®:")
        print("  æ¨™æº–ãƒ¢ãƒ¼ãƒ‰: ä½œæˆæ—¥ã€ãƒ•ã‚¡ã‚¤ãƒ«åã€æ–‡æ›¸ç¨®åˆ¥ã€ä½œæˆè€…ã€å®›å…ˆ")
        print("  æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰: è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ï¼ˆOCRã€AIåˆ†æã€è¦–è¦šæƒ…å ±ãªã©ï¼‰")
        print("  å®Œå…¨ãƒ¢ãƒ¼ãƒ‰: database.jsonã®å…¨ãƒ‡ãƒ¼ã‚¿ï¼ˆJSONåˆ—ï¼‰")
        print("â€»ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹ã€è¨¼æ‹ ç•ªå·ãªã©ã®è­˜åˆ¥å­ã¯ç·¨é›†ã§ãã¾ã›ã‚“")
        
        # ãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹ã¾ãŸã¯Google Drive URLã®å…¥åŠ›
        print("\nç·¨é›†ã—ãŸCSV/Excelãƒ•ã‚¡ã‚¤ãƒ«ã‚’æŒ‡å®šã—ã¦ãã ã•ã„:")
        print("  1. ãƒ­ãƒ¼ã‚«ãƒ«ãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹ï¼ˆçµ¶å¯¾ãƒ‘ã‚¹ã¾ãŸã¯ç›¸å¯¾ãƒ‘ã‚¹ã€æ‹¡å¼µå­: .csv, .xlsxï¼‰")
        print("  2. Google Drive ãƒ•ã‚¡ã‚¤ãƒ«URLï¼ˆhttps://drive.google.com/file/d/.../viewï¼‰")
        print("  3. Google Spreadsheets URLï¼ˆhttps://docs.google.com/spreadsheets/d/.../editï¼‰")
        file_input = input("\nãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹ã¾ãŸã¯URL: ").strip()
        
        if not file_input:
            print("ã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
            return
        
        # Google Drive URLã‹ã©ã†ã‹ã‚’åˆ¤å®š
        is_gdrive_url = file_input.startswith('http') and ('drive.google.com' in file_input or 'docs.google.com' in file_input)
        
        if is_gdrive_url:
            # Google Drive URLã‹ã‚‰ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰
            print("\nğŸ“¥ Google Driveã‹ã‚‰ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ä¸­...")
            
            # ä¸€æ™‚ãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹ã‚’ç”Ÿæˆ
            import tempfile
            temp_dir = tempfile.gettempdir()
            
            # Google Spreadsheetsã®URLã‹ã©ã†ã‹ã‚’åˆ¤å®š
            import re
            is_spreadsheet = bool(re.search(r'/spreadsheets/d/', file_input))
            
            if is_spreadsheet:
                # Google Spreadsheetsã®å ´åˆã¯è‡ªå‹•çš„ã«Excelå½¢å¼
                print("   ğŸ“Š Google Spreadsheetsã‚’æ¤œå‡º â†’ Excelå½¢å¼ã¨ã—ã¦å‡¦ç†ã—ã¾ã™")
                file_ext = '.xlsx'
            else:
                # é€šå¸¸ã®Google Driveãƒ•ã‚¡ã‚¤ãƒ«ã®å ´åˆã¯ãƒ¦ãƒ¼ã‚¶ãƒ¼ã«ç¢ºèª
                print("\nãƒ•ã‚¡ã‚¤ãƒ«å½¢å¼ã‚’é¸æŠã—ã¦ãã ã•ã„:")
                print("  1. CSVå½¢å¼ (.csv)")
                print("  2. Excelå½¢å¼ (.xlsx)")
                format_choice = input("\n> ").strip()
                
                if format_choice == '1':
                    file_ext = '.csv'
                elif format_choice == '2':
                    file_ext = '.xlsx'
                else:
                    print("âŒ ç„¡åŠ¹ãªé¸æŠã§ã™")
                    return
            
            temp_filename = f"imported_evidence{file_ext}"
            file_path = os.path.join(temp_dir, temp_filename)
            
            # Google Driveã‹ã‚‰ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰
            if not self._download_file_from_gdrive_url(file_input, file_path):
                print("âŒ ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ã«å¤±æ•—ã—ã¾ã—ãŸ")
                return
        else:
            # ãƒ­ãƒ¼ã‚«ãƒ«ãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹
            file_path = file_input
            
            # ãƒ•ã‚¡ã‚¤ãƒ«ã®å­˜åœ¨ç¢ºèª
            if not os.path.exists(file_path):
                print(f"\nâŒ ãƒ•ã‚¡ã‚¤ãƒ«ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“: {file_path}")
                return
        
        try:
            import csv
            
            # ãƒ•ã‚¡ã‚¤ãƒ«æ‹¡å¼µå­ã§åˆ¤å®š
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.xlsx':
                # Excelãƒ•ã‚¡ã‚¤ãƒ«ã‚’èª­ã¿è¾¼ã¿
                print("\nExcelãƒ•ã‚¡ã‚¤ãƒ«ã‚’èª­ã¿è¾¼ã‚“ã§ã„ã¾ã™...")
                csv_data = []
                
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # ãƒ˜ãƒƒãƒ€ãƒ¼è¡Œã‚’å–å¾—
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value if cell.value else '')
                
                # ãƒ‡ãƒ¼ã‚¿è¡Œã‚’èª­ã¿è¾¼ã¿
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = str(value) if value is not None else ''
                    csv_data.append(row_dict)
            
            elif file_ext == '.csv':
                # CSVãƒ•ã‚¡ã‚¤ãƒ«ã‚’èª­ã¿è¾¼ã¿
                print("\nCSVãƒ•ã‚¡ã‚¤ãƒ«ã‚’èª­ã¿è¾¼ã‚“ã§ã„ã¾ã™...")
                csv_data = []
                
                with open(file_path, 'r', encoding='utf-8-sig') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        csv_data.append(row)
            
            else:
                print(f"\nâŒ ã‚µãƒãƒ¼ãƒˆã•ã‚Œã¦ã„ãªã„ãƒ•ã‚¡ã‚¤ãƒ«å½¢å¼ã§ã™: {file_ext}")
                print("   å¯¾å¿œå½¢å¼: .csv, .xlsx")
                return
            
            if not csv_data:
                print("\nâŒ CSVãƒ•ã‚¡ã‚¤ãƒ«ã«ãƒ‡ãƒ¼ã‚¿ãŒã‚ã‚Šã¾ã›ã‚“")
                return
            
            print(f"âœ… {len(csv_data)}ä»¶ã®ãƒ¬ã‚³ãƒ¼ãƒ‰ã‚’èª­ã¿è¾¼ã¿ã¾ã—ãŸ")
            
            # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã‚’èª­ã¿è¾¼ã¿
            database = self.db_manager.load_database()
            evidence_list = database.get('evidence', [])
            
            if not evidence_list:
                print("\nâŒ ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã«è¨¼æ‹ ãŒç™»éŒ²ã•ã‚Œã¦ã„ã¾ã›ã‚“")
                return
            
            # è¨¼æ‹ ç•ªå·ã§ã‚¤ãƒ³ãƒ‡ãƒƒã‚¯ã‚¹ã‚’ä½œæˆï¼ˆé«˜é€Ÿæ¤œç´¢ç”¨ï¼‰
            evidence_index = {}
            for evidence in evidence_list:
                evidence_id = evidence.get('evidence_id', '')
                if evidence_id:
                    evidence_index[evidence_id] = evidence
            
            # CSVãƒ¢ãƒ¼ãƒ‰ã‚’åˆ¤å®šï¼ˆå®Œå…¨/æ‹¡å¼µ/æ¨™æº–ï¼‰
            first_row = csv_data[0] if csv_data else {}
            has_json_columns = 'complete_metadata_json' in first_row or 'phase1_complete_analysis_json' in first_row
            has_extended_columns = 'OCRãƒ†ã‚­ã‚¹ãƒˆ' in first_row or 'å®Œå…¨ãªèª¬æ˜' in first_row or 'ãƒ¬ã‚¤ã‚¢ã‚¦ãƒˆèª¬æ˜' in first_row
            
            if has_json_columns:
                csv_mode = 'full'
                print("\nâœ… å®Œå…¨ãƒ¢ãƒ¼ãƒ‰CSVæ¤œå‡º: JSONåˆ—ã‹ã‚‰database.jsonã®å…¨ãƒ‡ãƒ¼ã‚¿ã‚’èª­ã¿å–ã‚Šã¾ã™")
            elif has_extended_columns:
                csv_mode = 'extended'
                print("\nâœ… æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰CSVæ¤œå‡º: è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’å€‹åˆ¥ã«æ›´æ–°ã—ã¾ã™")
            else:
                csv_mode = 'standard'
                print("\nâœ… æ¨™æº–ãƒ¢ãƒ¼ãƒ‰CSVæ¤œå‡º: ä¸»è¦ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã®ã¿ã‚’æ›´æ–°ã—ã¾ã™")
            
            # æ›´æ–°å†…å®¹ã®ãƒ—ãƒ¬ãƒ“ãƒ¥ãƒ¼
            updates = []
            
            for row in csv_data:
                evidence_id = row.get('è¨¼æ‹ ç•ªå·', '').strip()
                
                if not evidence_id:
                    continue
                
                if evidence_id not in evidence_index:
                    print(f"\nâš ï¸  è¨¼æ‹ ç•ªå· {evidence_id} ãŒãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã«è¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ï¼ˆã‚¹ã‚­ãƒƒãƒ—ï¼‰")
                    continue
                
                evidence = evidence_index[evidence_id]
                
                # å¤‰æ›´ç®‡æ‰€ã‚’æ¤œå‡º
                changes = {}
                csv_update_data = {}
                
                if csv_mode == 'full':
                    # å®Œå…¨ãƒ¢ãƒ¼ãƒ‰: JSONåˆ—ã‹ã‚‰å…¨ãƒ‡ãƒ¼ã‚¿ã‚’èª­ã¿å–ã‚‹
                    csv_metadata_json = row.get('complete_metadata_json', '').strip()
                    csv_analysis_json = row.get('phase1_complete_analysis_json', '').strip()
                    
                    # JSONæ–‡å­—åˆ—ã‚’ãƒ‘ãƒ¼ã‚¹
                    if csv_metadata_json:
                        try:
                            csv_metadata = json.loads(csv_metadata_json)
                            current_metadata = evidence.get('complete_metadata', {})
                            
                            # å¤‰æ›´æ¤œå‡ºï¼ˆç°¡æ˜“æ¯”è¼ƒï¼‰
                            if csv_metadata != current_metadata:
                                changes['complete_metadata'] = ('å¤‰æ›´ã‚ã‚Š', 'å¤‰æ›´ã‚ã‚Š')
                                csv_update_data['complete_metadata'] = csv_metadata
                        except json.JSONDecodeError as e:
                            print(f"\nâš ï¸  {evidence_id}: complete_metadata_jsonã®è§£æã‚¨ãƒ©ãƒ¼: {e}")
                    
                    if csv_analysis_json:
                        try:
                            csv_analysis = json.loads(csv_analysis_json)
                            current_analysis = evidence.get('phase1_complete_analysis', {})
                            
                            # å¤‰æ›´æ¤œå‡ºï¼ˆç°¡æ˜“æ¯”è¼ƒï¼‰
                            if csv_analysis != current_analysis:
                                changes['phase1_complete_analysis'] = ('å¤‰æ›´ã‚ã‚Š', 'å¤‰æ›´ã‚ã‚Š')
                                csv_update_data['phase1_complete_analysis'] = csv_analysis
                        except json.JSONDecodeError as e:
                            print(f"\nâš ï¸  {evidence_id}: phase1_complete_analysis_jsonã®è§£æã‚¨ãƒ©ãƒ¼: {e}")
                
                elif csv_mode == 'extended':
                    # æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰: è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’å€‹åˆ¥ã«æ›´æ–°
                    current_extended = self._extract_extended_fields(evidence)
                    
                    # æ‹¡å¼µãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã®ãƒªã‚¹ãƒˆ
                    extended_fields = [
                        'ãƒ•ã‚¡ã‚¤ãƒ«å', 'ãƒ•ã‚¡ã‚¤ãƒ«ã‚µã‚¤ã‚º', 'ãƒšãƒ¼ã‚¸æ•°',
                        'è¨¼æ‹ ç¨®åˆ¥_è©³ç´°', 'ãƒ•ã‚©ãƒ¼ãƒãƒƒãƒˆèª¬æ˜',
                        'æ–‡æ›¸ç¨®åˆ¥', 'ä½œæˆæ—¥', 'æ™‚é–“çš„ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'çµ„ç¹”1_åç§°', 'çµ„ç¹”1_å½¹å‰²', 'çµ„ç¹”1_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'çµ„ç¹”2_åç§°', 'çµ„ç¹”2_å½¹å‰²', 'çµ„ç¹”2_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'çµ„ç¹”3_åç§°', 'çµ„ç¹”3_å½¹å‰²', 'çµ„ç¹”3_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'å€‹äºº1_åå‰', 'å€‹äºº1_å½¹å‰²', 'å€‹äºº1_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'å€‹äºº2_åå‰', 'å€‹äºº2_å½¹å‰²', 'å€‹äºº2_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'ãƒ¬ã‚¤ã‚¢ã‚¦ãƒˆèª¬æ˜', 'ãƒ†ã‚­ã‚¹ãƒˆå†…å®¹è¦ç´„', 'æ³¨ç›®ã™ã¹ãç‰¹å¾´',
                        'å®Œå…¨ãªèª¬æ˜', 'è©³ç´°å†…å®¹',
                        'å®¢è¦³çš„äº‹å®Ÿ', 'æ–‡è„ˆä¸Šã®é‡è¦æ€§',
                        'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ1', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ2', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ3',
                        'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ4', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ5',
                        'æ¨å¥¨ã•ã‚Œã‚‹ä½¿ç”¨æ–¹æ³•',
                        'OCRãƒ†ã‚­ã‚¹ãƒˆ',
                        'å®Œå…¨è¨€èªåŒ–ãƒ¬ãƒ™ãƒ«', 'ä¿¡é ¼åº¦ã‚¹ã‚³ã‚¢'
                    ]
                    
                    # CSVã‹ã‚‰æ‹¡å¼µãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’èª­ã¿å–ã‚Šã€å¤‰æ›´ã‚’æ¤œå‡º
                    csv_extended_data = {}
                    for field in extended_fields:
                        csv_value = row.get(field, '').strip()
                        current_value = str(current_extended.get(field, '')).strip()
                        
                        # å€¤ãŒå¤‰æ›´ã•ã‚Œã¦ã„ã‚‹å ´åˆï¼ˆå‹å¤‰æ›ã®ã¿ã®å¤‰æ›´ã¯ç„¡è¦–ï¼‰
                        if csv_value != current_value:
                            # æ•°å€¤å‹å¤‰æ›ã®ã¿ã®å¤‰æ›´ã‚’ç„¡è¦–ï¼ˆä¾‹: 2 â†’ 2.0ï¼‰
                            is_numeric_conversion_only = False
                            try:
                                # ä¸¡æ–¹ãŒæ•°å€¤ã«å¤‰æ›å¯èƒ½ã§ã€å€¤ãŒç­‰ã—ã„å ´åˆã¯å‹å¤‰æ›ã®ã¿ã¨ã¿ãªã™
                                if csv_value and current_value:
                                    csv_num = float(csv_value)
                                    current_num = float(current_value)
                                    if csv_num == current_num:
                                        is_numeric_conversion_only = True
                            except (ValueError, TypeError):
                                pass
                            
                            # å‹å¤‰æ›ã®ã¿ã®å¤‰æ›´ã§ãªã‘ã‚Œã°ã€å¤‰æ›´ã¨ã—ã¦è¨˜éŒ²
                            if not is_numeric_conversion_only:
                                changes[field] = (current_value[:50] + '...' if len(current_value) > 50 else current_value,
                                                csv_value[:50] + '...' if len(csv_value) > 50 else csv_value)
                                csv_extended_data[field] = csv_value
                    
                    if csv_extended_data:
                        csv_update_data['extended_fields'] = csv_extended_data
                
                else:
                    # æ¨™æº–ãƒ¢ãƒ¼ãƒ‰: ä¸»è¦ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã®ã¿æ›´æ–°
                    current_data = self._get_evidence_export_data(evidence)
                    
                    csv_file_name = row.get('ãƒ•ã‚¡ã‚¤ãƒ«å', '').strip()
                    csv_creation_date = row.get('ä½œæˆæ—¥', '').strip()
                    csv_document_type = row.get('æ–‡æ›¸ç¨®åˆ¥', '').strip()
                    csv_author = row.get('ä½œæˆè€…', '').strip()
                    csv_recipient = row.get('å®›å…ˆ', '').strip()
                    
                    if csv_file_name and csv_file_name != current_data['file_name']:
                        changes['file_name'] = (current_data['file_name'], csv_file_name)
                        csv_update_data['file_name'] = csv_file_name
                    
                    if csv_creation_date and csv_creation_date != current_data['creation_date']:
                        changes['creation_date'] = (current_data['creation_date'], csv_creation_date)
                        csv_update_data['creation_date'] = csv_creation_date
                    
                    if csv_document_type and csv_document_type != current_data['document_type']:
                        changes['document_type'] = (current_data['document_type'], csv_document_type)
                        csv_update_data['document_type'] = csv_document_type
                    
                    if csv_author and csv_author != current_data['author']:
                        changes['author'] = (current_data['author'], csv_author)
                        csv_update_data['author'] = csv_author
                    
                    if csv_recipient and csv_recipient != current_data['recipient']:
                        changes['recipient'] = (current_data['recipient'], csv_recipient)
                        csv_update_data['recipient'] = csv_recipient
                
                if changes:
                    updates.append({
                        'evidence_id': evidence_id,
                        'evidence': evidence,
                        'changes': changes,
                        'csv_data': csv_update_data,
                        'csv_mode': csv_mode
                    })
            
            if not updates:
                print("\nâœ… å¤‰æ›´ç®‡æ‰€ãŒã‚ã‚Šã¾ã›ã‚“ã€‚database.jsonã¯æ›´æ–°ã•ã‚Œã¾ã›ã‚“ã§ã—ãŸã€‚")
                return
            
            # å¤‰æ›´å†…å®¹ã®ãƒ—ãƒ¬ãƒ“ãƒ¥ãƒ¼è¡¨ç¤º
            print(f"\nã€æ›´æ–°ãƒ—ãƒ¬ãƒ“ãƒ¥ãƒ¼ã€‘ {len(updates)}ä»¶ã®è¨¼æ‹ ã«å¤‰æ›´ãŒã‚ã‚Šã¾ã™")
            print("-"*70)
            
            for i, update in enumerate(updates, 1):
                evidence_id = update['evidence_id']
                changes = update['changes']
                mode = update.get('csv_mode', 'standard')
                
                print(f"\n{i}. è¨¼æ‹ ç•ªå·: {evidence_id}")
                
                if mode == 'full':
                    # å®Œå…¨ãƒ¢ãƒ¼ãƒ‰: JSONå…¨ä½“ã®å¤‰æ›´
                    if 'complete_metadata' in changes:
                        print(f"   complete_metadata: å¤‰æ›´ã‚ã‚Š")
                    if 'phase1_complete_analysis' in changes:
                        print(f"   phase1_complete_analysis: å¤‰æ›´ã‚ã‚Š")
                
                elif mode == 'extended':
                    # æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰: ä¸»è¦ãªå¤‰æ›´ã‚’è¡¨ç¤ºï¼ˆæœ€å¤§5ä»¶ï¼‰
                    displayed_count = 0
                    for field_name, (old_val, new_val) in changes.items():
                        if displayed_count < 5:
                            print(f"   {field_name}: {old_val} â†’ {new_val}")
                            displayed_count += 1
                    
                    if len(changes) > 5:
                        print(f"   ... ä»– {len(changes) - 5} ä»¶ã®ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ãŒå¤‰æ›´ã•ã‚Œã¾ã—ãŸ")
                
                else:
                    # æ¨™æº–ãƒ¢ãƒ¼ãƒ‰: å€‹åˆ¥ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰
                    if 'file_name' in changes:
                        old, new = changes['file_name']
                        print(f"   ãƒ•ã‚¡ã‚¤ãƒ«å: {old} â†’ {new}")
                    
                    if 'creation_date' in changes:
                        old, new = changes['creation_date']
                        print(f"   ä½œæˆæ—¥: {old} â†’ {new}")
                    
                    if 'document_type' in changes:
                        old, new = changes['document_type']
                        print(f"   æ–‡æ›¸ç¨®åˆ¥: {old} â†’ {new}")
                    
                    if 'author' in changes:
                        old, new = changes['author']
                        print(f"   ä½œæˆè€…: {old} â†’ {new}")
                    
                    if 'recipient' in changes:
                        old, new = changes['recipient']
                        print(f"   å®›å…ˆ: {old} â†’ {new}")
            
            # ç¢ºèª
            print("\n" + "-"*70)
            print("ã“ã®å†…å®¹ã§database.jsonã‚’æ›´æ–°ã—ã¾ã™ã‹ï¼Ÿ")
            confirm = input("ç¶šè¡Œã™ã‚‹å ´åˆã¯ 'yes' ã¨å…¥åŠ›: ").strip().lower()
            
            if confirm != 'yes':
                print("\nã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
                return
            
            # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã‚’æ›´æ–°
            print("\nãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã‚’æ›´æ–°ä¸­...")
            
            for update in updates:
                evidence = update['evidence']
                csv_data = update['csv_data']
                mode = update.get('csv_mode', 'standard')
                
                if mode == 'full':
                    # å®Œå…¨ãƒ¢ãƒ¼ãƒ‰: JSONåˆ—ã‹ã‚‰å…¨ãƒ‡ãƒ¼ã‚¿ã‚’ç›´æ¥åæ˜ 
                    if 'complete_metadata' in csv_data:
                        evidence['complete_metadata'] = csv_data['complete_metadata']
                    
                    if 'phase1_complete_analysis' in csv_data:
                        evidence['phase1_complete_analysis'] = csv_data['phase1_complete_analysis']
                
                elif mode == 'extended':
                    # æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰: è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’å€‹åˆ¥ã«åæ˜ 
                    extended_fields = csv_data.get('extended_fields', {})
                    if extended_fields:
                        self._apply_extended_fields(evidence, extended_fields)
                
                else:
                    # æ¨™æº–ãƒ¢ãƒ¼ãƒ‰: ä¸»è¦ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã®ã¿æ›´æ–°
                    # ãƒ•ã‚¡ã‚¤ãƒ«åã®æ›´æ–°
                    if 'file_name' in csv_data:
                        evidence['original_filename'] = csv_data['file_name']
                        if 'complete_metadata' not in evidence:
                            evidence['complete_metadata'] = {}
                        if 'basic' not in evidence['complete_metadata']:
                            evidence['complete_metadata']['basic'] = {}
                        evidence['complete_metadata']['basic']['file_name'] = csv_data['file_name']
                    
                    # ä½œæˆæ—¥ã®æ›´æ–°ï¼ˆAIåˆ†æçµæœå†…ã«æ ¼ç´ï¼‰
                    if 'creation_date' in csv_data:
                        phase1_analysis = evidence.get('phase1_complete_analysis', {})
                        ai_analysis = phase1_analysis.get('ai_analysis', {})
                        
                        if ai_analysis:
                            # æ–°ã—ã„æ§‹é€ : phase1_complete_analysis.ai_analysis.objective_analysis.temporal_information.document_date
                            if 'objective_analysis' not in ai_analysis:
                                ai_analysis['objective_analysis'] = {}
                            if 'temporal_information' not in ai_analysis['objective_analysis']:
                                ai_analysis['objective_analysis']['temporal_information'] = {}
                            ai_analysis['objective_analysis']['temporal_information']['document_date'] = csv_data['creation_date']
                        else:
                            # æ—§æ§‹é€ ç”¨ã®ãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯
                            if 'temporal_information' not in evidence:
                                evidence['temporal_information'] = {}
                            evidence['temporal_information']['document_date'] = csv_data['creation_date']
                    
                    # æ–‡æ›¸ç¨®åˆ¥ã®æ›´æ–°
                    if 'document_type' in csv_data:
                        phase1_analysis = evidence.get('phase1_complete_analysis', {})
                        ai_analysis = phase1_analysis.get('ai_analysis', {})
                        
                        if ai_analysis:
                            if 'objective_analysis' not in ai_analysis:
                                ai_analysis['objective_analysis'] = {}
                            ai_analysis['objective_analysis']['document_type'] = csv_data['document_type']
                        else:
                            # æ—§æ§‹é€ ç”¨ã®ãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯
                            phase1_analysis['document_type'] = csv_data['document_type']
                    
                    # ä½œæˆè€…ãƒ»å®›å…ˆã®æ›´æ–°
                    if 'author' in csv_data or 'recipient' in csv_data:
                        phase1_analysis = evidence.get('phase1_complete_analysis', {})
                        ai_analysis = phase1_analysis.get('ai_analysis', {})
                        
                        if ai_analysis:
                            if 'objective_analysis' not in ai_analysis:
                                ai_analysis['objective_analysis'] = {}
                            if 'parties_mentioned' not in ai_analysis['objective_analysis']:
                                ai_analysis['objective_analysis']['parties_mentioned'] = {}
                            if 'organizations' not in ai_analysis['objective_analysis']['parties_mentioned']:
                                ai_analysis['objective_analysis']['parties_mentioned']['organizations'] = []
                            
                            organizations = ai_analysis['objective_analysis']['parties_mentioned']['organizations']
                            
                            # ä½œæˆè€…ï¼ˆorganizations[0]ï¼‰
                            if 'author' in csv_data:
                                if len(organizations) == 0:
                                    organizations.append(csv_data['author'])
                                else:
                                    organizations[0] = csv_data['author']
                            
                            # å®›å…ˆï¼ˆorganizations[1]ï¼‰
                            if 'recipient' in csv_data:
                                if len(organizations) <= 1:
                                    organizations.append(csv_data['recipient'])
                                else:
                                    organizations[1] = csv_data['recipient']
                        else:
                            # æ—§æ§‹é€ ç”¨ã®ãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯
                            if 'author' in csv_data:
                                phase1_analysis['author'] = csv_data['author']
                            if 'recipient' in csv_data:
                                phase1_analysis['recipient'] = csv_data['recipient']
            
            # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã‚’ä¿å­˜ï¼ˆGoogle Driveã«åŒæœŸï¼‰
            print("Google Driveã«åŒæœŸä¸­...")
            self.save_database(database)
            
            print(f"\nâœ… {len(updates)}ä»¶ã®è¨¼æ‹ ã‚’æ›´æ–°ã—ã€Google Driveã«åŒæœŸã—ã¾ã—ãŸï¼")
            
        except Exception as e:
            print(f"\nâŒ ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {e}")
            import traceback
            traceback.print_exc()
    
    def _export_to_excel(self, evidence_list: List[Dict], filename: str, evidence_type: str = 'ko', full_data: bool = False, extended_mode: bool = False):
        """Excelå½¢å¼ã§ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆ
        
        Args:
            evidence_list: è¨¼æ‹ ãƒªã‚¹ãƒˆ
            filename: å‡ºåŠ›ãƒ•ã‚¡ã‚¤ãƒ«å
            evidence_type: è¨¼æ‹ ç¨®åˆ¥ ('ko' ã¾ãŸã¯ 'otsu')
            full_data: True ã®å ´åˆã€database.jsonã®å…¨ãƒ‡ãƒ¼ã‚¿ã‚’JSONåˆ—ã¨ã—ã¦å‡ºåŠ›
            extended_mode: True ã®å ´åˆã€è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’å€‹åˆ¥åˆ—ã«å±•é–‹ï¼ˆæ‹¡å¼µãƒ¢ãƒ¼ãƒ‰ï¼‰
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("\nâš ï¸  openpyxlãŒã‚¤ãƒ³ã‚¹ãƒˆãƒ¼ãƒ«ã•ã‚Œã¦ã„ã¾ã›ã‚“")
            print("ä»¥ä¸‹ã®ã‚³ãƒãƒ³ãƒ‰ã§ã‚¤ãƒ³ã‚¹ãƒˆãƒ¼ãƒ«ã—ã¦ãã ã•ã„:")
            print("  pip3 install openpyxl")
            print("\nã¾ãŸã¯:")
            print("  pip3 install --break-system-packages openpyxl")
            return
        
        try:
            type_name = "ç”²å·è¨¼" if evidence_type == 'ko' else "ä¹™å·è¨¼"
            output_path = os.path.join(os.getcwd(), filename)
            
            # ãƒ¯ãƒ¼ã‚¯ãƒ–ãƒƒã‚¯ã¨ã‚·ãƒ¼ãƒˆã‚’ä½œæˆ
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{type_name}ä¸€è¦§"
            
            # ãƒ˜ãƒƒãƒ€ãƒ¼è¡Œã®ã‚¹ã‚¿ã‚¤ãƒ«
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ãƒ˜ãƒƒãƒ€ãƒ¼è¡Œ
            if extended_mode:
                # æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰: è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’å€‹åˆ¥åˆ—ã«å±•é–‹ï¼ˆCSV ã¨åŒã˜ï¼‰
                headers = [
                    'è¨¼æ‹ ç¨®åˆ¥', 'ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹', 'è¨¼æ‹ ç•ªå·', 'ä»®ç•ªå·', 'åˆ†æçŠ¶æ…‹',
                    'Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID',
                    'ãƒ•ã‚¡ã‚¤ãƒ«å', 'ãƒ•ã‚¡ã‚¤ãƒ«ã‚µã‚¤ã‚º', 'ãƒšãƒ¼ã‚¸æ•°',
                    'è¨¼æ‹ ç¨®åˆ¥_è©³ç´°', 'ãƒ•ã‚©ãƒ¼ãƒãƒƒãƒˆèª¬æ˜',
                    'æ–‡æ›¸ç¨®åˆ¥', 'ä½œæˆæ—¥', 'æ™‚é–“çš„ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                    'çµ„ç¹”1_åç§°', 'çµ„ç¹”1_å½¹å‰²', 'çµ„ç¹”1_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                    'çµ„ç¹”2_åç§°', 'çµ„ç¹”2_å½¹å‰²', 'çµ„ç¹”2_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                    'çµ„ç¹”3_åç§°', 'çµ„ç¹”3_å½¹å‰²', 'çµ„ç¹”3_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                    'å€‹äºº1_åå‰', 'å€‹äºº1_å½¹å‰²', 'å€‹äºº1_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                    'å€‹äºº2_åå‰', 'å€‹äºº2_å½¹å‰²', 'å€‹äºº2_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                    'ãƒ¬ã‚¤ã‚¢ã‚¦ãƒˆèª¬æ˜', 'ãƒ†ã‚­ã‚¹ãƒˆå†…å®¹è¦ç´„', 'æ³¨ç›®ã™ã¹ãç‰¹å¾´',
                    'å®Œå…¨ãªèª¬æ˜', 'è©³ç´°å†…å®¹',
                    'å®¢è¦³çš„äº‹å®Ÿ', 'æ–‡è„ˆä¸Šã®é‡è¦æ€§',
                    'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ1', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ2', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ3',
                    'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ4', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ5',
                    'æ¨å¥¨ã•ã‚Œã‚‹ä½¿ç”¨æ–¹æ³•',
                    'OCRãƒ†ã‚­ã‚¹ãƒˆ',
                    'å®Œå…¨è¨€èªåŒ–ãƒ¬ãƒ™ãƒ«', 'ä¿¡é ¼åº¦ã‚¹ã‚³ã‚¢'
                ]
            elif full_data:
                headers = [
                    'è¨¼æ‹ ç¨®åˆ¥', 'ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹', 'è¨¼æ‹ ç•ªå·', 'ä»®ç•ªå·', 'ä½œæˆæ—¥', 
                    'åˆ†æçŠ¶æ…‹', 'ãƒ•ã‚¡ã‚¤ãƒ«å', 'æ–‡æ›¸ç¨®åˆ¥', 'ä½œæˆè€…',
                    'å®›å…ˆ', 'è¦ç´„', 'Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID',
                    'complete_metadata_json', 'phase1_complete_analysis_json'
                ]
            else:
                headers = [
                    'è¨¼æ‹ ç¨®åˆ¥', 'ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹', 'è¨¼æ‹ ç•ªå·', 'ä»®ç•ªå·', 'ä½œæˆæ—¥', 
                    'åˆ†æçŠ¶æ…‹', 'ãƒ•ã‚¡ã‚¤ãƒ«å', 'æ–‡æ›¸ç¨®åˆ¥', 'ä½œæˆè€…',
                    'å®›å…ˆ', 'è¦ç´„', 'Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID'
                ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # åˆ—å¹…ã‚’è¨­å®š
            if extended_mode:
                # æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰: å…¨åˆ—ã«ãƒ‡ãƒ•ã‚©ãƒ«ãƒˆåˆ—å¹…ã‚’è¨­å®š
                for col_idx in range(1, len(headers) + 1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    # åŸºæœ¬åˆ—ã¯ç‹­ã‚ã€èª¬æ˜åˆ—ã¯åºƒã‚ã«è¨­å®š
                    if col_idx <= 6:  # åŸºæœ¬æƒ…å ±
                        ws.column_dimensions[col_letter].width = 15
                    elif 'èª¬æ˜' in headers[col_idx-1] or 'å†…å®¹' in headers[col_idx-1] or 'ãƒ†ã‚­ã‚¹ãƒˆ' in headers[col_idx-1]:
                        ws.column_dimensions[col_letter].width = 60
                    elif 'äº‹å®Ÿ' in headers[col_idx-1]:
                        ws.column_dimensions[col_letter].width = 50
                    else:
                        ws.column_dimensions[col_letter].width = 25
            elif full_data:
                column_widths = {
                    'A': 12,  # è¨¼æ‹ ç¨®åˆ¥
                    'B': 15,  # ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹
                    'C': 12,  # è¨¼æ‹ ç•ªå·
                    'D': 12,  # ä»®ç•ªå·
                    'E': 12,  # ä½œæˆæ—¥
                    'F': 12,  # åˆ†æçŠ¶æ…‹
                    'G': 40,  # ãƒ•ã‚¡ã‚¤ãƒ«å
                    'H': 15,  # æ–‡æ›¸ç¨®åˆ¥
                    'I': 20,  # ä½œæˆè€…
                    'J': 20,  # å®›å…ˆ
                    'K': 60,  # è¦ç´„
                    'L': 35,  # Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID
                    'M': 80,  # complete_metadata_json
                    'N': 80   # phase1_complete_analysis_json
                }
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
            else:
                column_widths = {
                    'A': 12,  # è¨¼æ‹ ç¨®åˆ¥
                    'B': 15,  # ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹
                    'C': 12,  # è¨¼æ‹ ç•ªå·
                    'D': 12,  # ä»®ç•ªå·
                    'E': 12,  # ä½œæˆæ—¥
                    'F': 12,  # åˆ†æçŠ¶æ…‹
                    'G': 40,  # ãƒ•ã‚¡ã‚¤ãƒ«å
                    'H': 15,  # æ–‡æ›¸ç¨®åˆ¥
                    'I': 20,  # ä½œæˆè€…
                    'J': 20,  # å®›å…ˆ
                    'K': 60,  # è¦ç´„
                    'L': 35   # Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID
                }
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
            
            # ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹ã§ã‚½ãƒ¼ãƒˆ
            def get_organization_status(evidence):
                """è¨¼æ‹ ã®æ•´ç†çŠ¶æ…‹ã‚’åˆ¤å®š"""
                evidence_id = evidence.get('evidence_id', '')
                evidence_number = evidence.get('evidence_number', '')
                
                if evidence_number and not evidence_number.startswith('ç”²tmp') and not evidence_number.startswith('ä¹™tmp'):
                    return 'ç¢ºå®šæ¸ˆã¿'
                elif evidence_id.startswith('tmp_'):
                    return 'æ•´ç†æ¸ˆã¿_æœªç¢ºå®š'
                else:
                    return 'æœªåˆ†é¡'
            
            status_order = {'ç¢ºå®šæ¸ˆã¿': 1, 'æ•´ç†æ¸ˆã¿_æœªç¢ºå®š': 2, 'æœªåˆ†é¡': 3}
            sorted_evidence = sorted(
                evidence_list, 
                key=lambda x: (
                    status_order.get(get_organization_status(x), 99),
                    x.get('evidence_id', ''),
                    x.get('temp_id', '')
                )
            )
            
            # ãƒ‡ãƒ¼ã‚¿è¡Œ
            for row_idx, evidence in enumerate(sorted_evidence, 2):
                status = get_organization_status(evidence)
                evidence_id = evidence.get('evidence_id', '')
                temp_id = evidence.get('temp_id', '')
                
                # ãƒ‡ãƒ¼ã‚¿æ§‹é€ ã®é•ã„ã‚’å¸åã—ã¦ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆç”¨ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
                export_data = self._get_evidence_export_data(evidence)
                
                # Google Drive ãƒ•ã‚¡ã‚¤ãƒ«ID
                gdrive_file_id = evidence.get('gdrive_file_id', '') or evidence.get('complete_metadata', {}).get('gdrive', {}).get('file_id', '')
                
                summary = export_data['complete_description'] or ''
                
                # åˆ†æçŠ¶æ…‹
                analysis_status = "åˆ†ææ¸ˆã¿" if export_data['complete_description'] else "æœªåˆ†æ"
                
                # ã‚»ãƒ«ã«å€¤ã‚’è¨­å®š
                if extended_mode:
                    # æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰: è‡ªç„¶è¨€èªãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’å€‹åˆ¥åˆ—ã«å±•é–‹
                    extended_data = self._extract_extended_fields(evidence)
                    
                    # åŸºæœ¬æƒ…å ±ï¼ˆæœ€åˆã®6åˆ—ï¼‰
                    row_data = [
                        type_name,          # è¨¼æ‹ ç¨®åˆ¥
                        status,             # ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹
                        evidence_id,        # è¨¼æ‹ ç•ªå·
                        temp_id,            # ä»®ç•ªå·
                        analysis_status,    # åˆ†æçŠ¶æ…‹
                        gdrive_file_id      # Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID
                    ]
                    
                    # æ‹¡å¼µãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’é †ç•ªã«è¿½åŠ ï¼ˆãƒ˜ãƒƒãƒ€ãƒ¼ã¨åŒã˜é †åºï¼‰
                    extended_field_order = [
                        'ãƒ•ã‚¡ã‚¤ãƒ«å', 'ãƒ•ã‚¡ã‚¤ãƒ«ã‚µã‚¤ã‚º', 'ãƒšãƒ¼ã‚¸æ•°',
                        'è¨¼æ‹ ç¨®åˆ¥_è©³ç´°', 'ãƒ•ã‚©ãƒ¼ãƒãƒƒãƒˆèª¬æ˜',
                        'æ–‡æ›¸ç¨®åˆ¥', 'ä½œæˆæ—¥', 'æ™‚é–“çš„ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'çµ„ç¹”1_åç§°', 'çµ„ç¹”1_å½¹å‰²', 'çµ„ç¹”1_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'çµ„ç¹”2_åç§°', 'çµ„ç¹”2_å½¹å‰²', 'çµ„ç¹”2_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'çµ„ç¹”3_åç§°', 'çµ„ç¹”3_å½¹å‰²', 'çµ„ç¹”3_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'å€‹äºº1_åå‰', 'å€‹äºº1_å½¹å‰²', 'å€‹äºº1_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'å€‹äºº2_åå‰', 'å€‹äºº2_å½¹å‰²', 'å€‹äºº2_ã‚³ãƒ³ãƒ†ã‚­ã‚¹ãƒˆ',
                        'ãƒ¬ã‚¤ã‚¢ã‚¦ãƒˆèª¬æ˜', 'ãƒ†ã‚­ã‚¹ãƒˆå†…å®¹è¦ç´„', 'æ³¨ç›®ã™ã¹ãç‰¹å¾´',
                        'å®Œå…¨ãªèª¬æ˜', 'è©³ç´°å†…å®¹',
                        'å®¢è¦³çš„äº‹å®Ÿ', 'æ–‡è„ˆä¸Šã®é‡è¦æ€§',
                        'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ1', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ2', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ3',
                        'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ4', 'è¨¼æ˜å¯èƒ½ãªäº‹å®Ÿ5',
                        'æ¨å¥¨ã•ã‚Œã‚‹ä½¿ç”¨æ–¹æ³•',
                        'OCRãƒ†ã‚­ã‚¹ãƒˆ',
                        'å®Œå…¨è¨€èªåŒ–ãƒ¬ãƒ™ãƒ«', 'ä¿¡é ¼åº¦ã‚¹ã‚³ã‚¢'
                    ]
                    
                    for field_name in extended_field_order:
                        row_data.append(extended_data.get(field_name, ''))
                
                elif full_data:
                    # å…¨ãƒ‡ãƒ¼ã‚¿ãƒ¢ãƒ¼ãƒ‰: æ¨™æº–ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ + JSONåˆ—
                    row_data = [
                        type_name,                      # è¨¼æ‹ ç¨®åˆ¥
                        status,                         # ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹
                        evidence_id,                    # è¨¼æ‹ ç•ªå·
                        temp_id,                        # ä»®ç•ªå·
                        export_data['creation_date'],   # ä½œæˆæ—¥
                        analysis_status,                # åˆ†æçŠ¶æ…‹
                        export_data['file_name'],       # ãƒ•ã‚¡ã‚¤ãƒ«å
                        export_data['document_type'],   # æ–‡æ›¸ç¨®åˆ¥
                        export_data['author'],          # ä½œæˆè€…
                        export_data['recipient'],       # å®›å…ˆ
                        summary,                        # è¦ç´„
                        gdrive_file_id                  # Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID
                    ]
                    
                    # JSONåˆ—ã‚’è¿½åŠ 
                    complete_metadata = evidence.get('complete_metadata', {})
                    row_data.append(json.dumps(complete_metadata, ensure_ascii=False) if complete_metadata else '')
                    
                    phase1_analysis = evidence.get('phase1_complete_analysis', {})
                    row_data.append(json.dumps(phase1_analysis, ensure_ascii=False) if phase1_analysis else '')
                
                else:
                    # æ¨™æº–ãƒ¢ãƒ¼ãƒ‰: ä¸»è¦ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã®ã¿
                    row_data = [
                        type_name,                      # è¨¼æ‹ ç¨®åˆ¥
                        status,                         # ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹
                        evidence_id,                    # è¨¼æ‹ ç•ªå·
                        temp_id,                        # ä»®ç•ªå·
                        export_data['creation_date'],   # ä½œæˆæ—¥
                        analysis_status,                # åˆ†æçŠ¶æ…‹
                        export_data['file_name'],       # ãƒ•ã‚¡ã‚¤ãƒ«å
                        export_data['document_type'],   # æ–‡æ›¸ç¨®åˆ¥
                        export_data['author'],          # ä½œæˆè€…
                        export_data['recipient'],       # å®›å…ˆ
                        summary,                        # è¦ç´„
                        gdrive_file_id                  # Google Driveãƒ•ã‚¡ã‚¤ãƒ«ID
                    ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    
                    # ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹ã«å¿œã˜ã¦èƒŒæ™¯è‰²ã‚’è¨­å®š
                    if col_idx == 2:  # ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹åˆ—ï¼ˆè¨¼æ‹ ç¨®åˆ¥ã®æ¬¡ï¼‰
                        if status == 'ç¢ºå®šæ¸ˆã¿':
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif status == 'æ•´ç†æ¸ˆã¿_æœªç¢ºå®š':
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    # åˆ†æçŠ¶æ…‹ã«å¿œã˜ã¦èƒŒæ™¯è‰²ã‚’è¨­å®š
                    # æ‹¡å¼µãƒ¢ãƒ¼ãƒ‰ã¯5åˆ—ç›®ã€æ¨™æº–ãƒ»å…¨ãƒ‡ãƒ¼ã‚¿ãƒ¢ãƒ¼ãƒ‰ã¯6åˆ—ç›®
                    analysis_status_col = 5 if extended_mode else 6
                    if col_idx == analysis_status_col:
                        if analysis_status == "åˆ†ææ¸ˆã¿":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            # ãƒ•ãƒªãƒ¼ã‚ºãƒšã‚¤ãƒ³ï¼ˆãƒ˜ãƒƒãƒ€ãƒ¼è¡Œã‚’å›ºå®šï¼‰
            ws.freeze_panes = "A2"
            
            # ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ä¿å­˜
            wb.save(output_path)
            
            print(f"\nâœ… Excelå½¢å¼ã§ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆã—ã¾ã—ãŸ")
            print(f"   ãƒ•ã‚¡ã‚¤ãƒ«: {output_path}")
            print(f"   ä»¶æ•°: {len(evidence_list)}ä»¶")
            
            # Google Driveã¸ã®ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰ã‚’ç¢ºèª
            upload_choice = input("\nGoogle Driveã«ã‚‚ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰ã—ã¾ã™ã‹ï¼Ÿ (y/n): ").strip().lower()
            if upload_choice == 'y':
                gdrive_url = self._upload_export_file_to_gdrive(output_path, filename)
                if gdrive_url:
                    print(f"\nâœ… Google Driveã«ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰ã—ã¾ã—ãŸ")
                    print(f"   URL: {gdrive_url}")
            
        except Exception as e:
            print(f"\nâŒ ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆã«å¤±æ•—ã—ã¾ã—ãŸ: {e}")
            import traceback
            traceback.print_exc()
    
    def generate_timeline_story(self):
        """æ™‚ç³»åˆ—ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã®ç”Ÿæˆï¼ˆæ‹¡å¼µç‰ˆï¼‰
        
        è¨¼æ‹ ã‚’æ™‚ç³»åˆ—é †ã«ä¸¦ã¹ã¦ã€å®¢è¦³çš„ãªäº‹å®Ÿã®æµã‚Œã‚’ç”Ÿæˆã—ã¾ã™ã€‚
        æ³•çš„åˆ¤æ–­ã¯å«ã¾ãšã€è¨¼æ‹ ã‹ã‚‰æŠ½å‡ºã•ã‚ŒãŸäº‹å®Ÿã®ã¿ã‚’æ™‚ç³»åˆ—ã§æ•´ç†ã—ã¾ã™ã€‚
        AIæ©Ÿèƒ½ã‚’ä½¿ç”¨ã™ã‚‹ã¨ã€ã‚ˆã‚Šèª­ã¿ã‚„ã™ã„å®¢è¦³çš„ãªã‚¹ãƒˆãƒ¼ãƒªãƒ¼ãŒç”Ÿæˆã•ã‚Œã¾ã™ã€‚
        """
        print("\n" + "="*70)
        print("  æ™‚ç³»åˆ—ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã®ç”Ÿæˆï¼ˆæ‹¡å¼µç‰ˆï¼‰")
        print("="*70)
        print("\nè¨¼æ‹ ã‚’æ™‚ç³»åˆ—é †ã«æ•´ç†ã—ã¦ã€å®¢è¦³çš„ãªäº‹å®Ÿã®æµã‚Œã‚’ç”Ÿæˆã—ã¾ã™ã€‚")
        print("æ³•çš„åˆ¤æ–­ã¯å«ã¾ãšã€è¨¼æ‹ ã‹ã‚‰æŠ½å‡ºã•ã‚ŒãŸäº‹å®Ÿã®ã¿ã‚’è¨˜è¼‰ã—ã¾ã™ã€‚")
        
        # AIæ©Ÿèƒ½ã®ä½¿ç”¨ç¢ºèª
        print("\nã€AIæ©Ÿèƒ½ã®é¸æŠã€‘")
        print("  1. AIç”Ÿæˆã®å®¢è¦³çš„ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ï¼ˆæ¨å¥¨ï¼šèª­ã¿ã‚„ã™ã„ï¼‰")
        print("  2. åŸºæœ¬çš„ãªã‚¹ãƒˆãƒ¼ãƒªãƒ¼ï¼ˆAIä¸ä½¿ç”¨ï¼‰")
        print("  0. ã‚­ãƒ£ãƒ³ã‚»ãƒ«")
        
        ai_choice = input("\né¸æŠ (0-2): ").strip()
        
        if ai_choice == '0':
            print("\nã‚­ãƒ£ãƒ³ã‚»ãƒ«ã—ã¾ã—ãŸ")
            return
        
        use_ai = (ai_choice == '1')
        
        try:
            # TimelineBuilderã‚’åˆæœŸåŒ–
            print("\nè¨¼æ‹ ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã‚’èª­ã¿è¾¼ã¿ä¸­...")
            builder = TimelineBuilder(self.case_manager, self.current_case, use_ai=use_ai)
            
            # ã‚¿ã‚¤ãƒ ãƒ©ã‚¤ãƒ³ã‚’æ§‹ç¯‰
            print("è¨¼æ‹ ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã‚’åˆ†æä¸­...")
            timeline_events = builder.build_timeline()
            
            if not timeline_events:
                print("\nâš ï¸ ã‚¿ã‚¤ãƒ ãƒ©ã‚¤ãƒ³ã‚’æ§‹ç¯‰ã§ãã¾ã›ã‚“ã§ã—ãŸã€‚")
                print("è¨¼æ‹ ãŒç™»éŒ²ã•ã‚Œã¦ã„ãªã„ã‹ã€è¨¼æ‹ ã«æ—¥ä»˜æƒ…å ±ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")
                return
            
            # ãƒŠãƒ©ãƒ†ã‚£ãƒ–ï¼ˆç‰©èªï¼‰ã‚’ç”Ÿæˆ
            narrative_result = None
            if use_ai:
                print("\nğŸ¤– Claude AIã«ã‚ˆã‚‹å®¢è¦³çš„ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã‚’ç”Ÿæˆä¸­...")
                print("ï¼ˆé«˜å“è³ªãªåˆ†æã®ãŸã‚ã€æ•°åç§’ã‹ã‹ã‚‹å ´åˆãŒã‚ã‚Šã¾ã™ï¼‰")
                narrative_result = builder.generate_objective_narrative(timeline_events)
                narrative_text = narrative_result.get("narrative", "") if isinstance(narrative_result, dict) else narrative_result
            else:
                print("\næ™‚ç³»åˆ—ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã‚’ç”Ÿæˆä¸­...")
                narrative_text = builder.generate_narrative(timeline_events)
            
            # ç”»é¢ã«è¡¨ç¤º
            print("\n" + "="*70)
            print("ã€ç”Ÿæˆã•ã‚ŒãŸæ™‚ç³»åˆ—ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã€‘")
            print("="*70)
            print("\n" + narrative_text)
            
            # äº‹å®Ÿã¨è¨¼æ‹ ã®ç´ä»˜ã‘ã‚’è¡¨ç¤ºï¼ˆAIä½¿ç”¨æ™‚ã®ã¿ï¼‰
            if use_ai and isinstance(narrative_result, dict):
                fact_mapping = narrative_result.get("fact_evidence_mapping", [])
                if fact_mapping:
                    print("\n" + "="*70)
                    print("ã€äº‹å®Ÿã¨è¨¼æ‹ ã®ç´ä»˜ã‘ã€‘")
                    print("="*70)
                    for i, fact in enumerate(fact_mapping[:5], 1):  # æœ€åˆã®5ä»¶ã®ã¿è¡¨ç¤º
                        print(f"\n{i}. {fact.get('fact_description', 'ä¸æ˜ãªäº‹å®Ÿ')}")
                        print(f"   æ—¥ä»˜: {fact.get('date', 'ä¸æ˜')}")
                        print(f"   è£ä»˜ã‘è¨¼æ‹ : {', '.join(fact.get('evidence_numbers', []))}")
                        print(f"   ç¢ºå®Ÿæ€§: {fact.get('confidence', 'unknown')}")
                    
                    if len(fact_mapping) > 5:
                        print(f"\n   ... ä»– {len(fact_mapping) - 5}ä»¶ã®äº‹å®Ÿï¼ˆè©³ç´°ã¯ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆãƒ•ã‚¡ã‚¤ãƒ«ã‚’å‚ç…§ï¼‰")
            
            # è‡ªç„¶è¨€èªã«ã‚ˆã‚‹æ”¹å–„ã‚ªãƒ—ã‚·ãƒ§ãƒ³ï¼ˆAIä½¿ç”¨æ™‚ã®ã¿ï¼‰
            if use_ai and isinstance(narrative_result, dict):
                print("\n" + "="*70)
                print("ã€ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã®æ”¹å–„ã€‘")
                print("è‡ªç„¶è¨€èªã§æ”¹å–„æŒ‡ç¤ºã‚’å…¥åŠ›ã§ãã¾ã™ï¼ˆä¾‹ï¼šã€Œã‚‚ã£ã¨è©³ã—ãã€ã€Œç°¡æ½”ã«ã€ï¼‰")
                print("æ”¹å–„ã—ãªã„å ´åˆã¯ç©ºEnterã‚’æŠ¼ã—ã¦ãã ã•ã„ã€‚")
                print("="*70)
                
                user_instruction = input("\næ”¹å–„æŒ‡ç¤º: ").strip()
                
                if user_instruction:
                    print("\nğŸ”„ ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã‚’æ”¹å–„ä¸­...")
                    improved_result = builder.refine_narrative_with_instruction(narrative_result, user_instruction)
                    
                    # æ”¹å–„å¾Œã®ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã‚’è¡¨ç¤º
                    improved_narrative = improved_result.get("narrative", "")
                    print("\n" + "="*70)
                    print("ã€æ”¹å–„å¾Œã®æ™‚ç³»åˆ—ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã€‘")
                    print("="*70)
                    print("\n" + improved_narrative)
                    
                    # æ”¹å–„çµæœã‚’ä½¿ç”¨ã™ã‚‹ã‹ç¢ºèª
                    use_improved = input("\næ”¹å–„å¾Œã®ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã‚’ä½¿ç”¨ã—ã¾ã™ã‹ï¼Ÿ (y/n): ").strip().lower()
                    if use_improved == 'y':
                        narrative_result = improved_result
                        narrative_text = improved_narrative
                        print("âœ… æ”¹å–„å¾Œã®ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã‚’æ¡ç”¨ã—ã¾ã—ãŸ")
                    else:
                        print("å…ƒã®ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã‚’ä½¿ç”¨ã—ã¾ã™")
            
            # ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆå½¢å¼ã‚’é¸æŠ
            print("\n" + "="*70)
            print("ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆå½¢å¼ã‚’é¸æŠã—ã¦ãã ã•ã„:")
            print("  1. JSONå½¢å¼ï¼ˆãƒ—ãƒ­ã‚°ãƒ©ãƒ ã§å‡¦ç†å¯èƒ½ï¼‰")
            print("  2. Markdownå½¢å¼ï¼ˆèª­ã¿ã‚„ã™ã„ï¼‰")
            print("  3. ãƒ†ã‚­ã‚¹ãƒˆå½¢å¼ï¼ˆã‚·ãƒ³ãƒ—ãƒ«ï¼‰")
            print("  4. HTMLå½¢å¼ï¼ˆãƒ–ãƒ©ã‚¦ã‚¶ã§é–²è¦§å¯èƒ½ï¼‰")
            print("  5. ã™ã¹ã¦ã®å½¢å¼")
            print("  0. ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆã—ãªã„")
            
            choice = input("\né¸æŠ (0-5): ").strip()
            
            export_formats = []
            if choice == '1':
                export_formats = ['json']
            elif choice == '2':
                export_formats = ['markdown']
            elif choice == '3':
                export_formats = ['text']
            elif choice == '4':
                export_formats = ['html']
            elif choice == '5':
                export_formats = ['json', 'markdown', 'text', 'html']
            elif choice == '0':
                print("\nã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆã‚’ã‚¹ã‚­ãƒƒãƒ—ã—ã¾ã—ãŸ")
                return
            else:
                print("\nç„¡åŠ¹ãªé¸æŠã§ã™")
                return
            
            # ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆ
            print("\nãƒ•ã‚¡ã‚¤ãƒ«ã‚’ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆä¸­...")
            for format_type in export_formats:
                output_path = builder.export_timeline(
                    timeline_events, 
                    output_format=format_type,
                    include_ai_narrative=use_ai
                )
                if output_path:
                    print(f"  âœ… {format_type.upper()}: {output_path}")
            
            print("\nâœ… æ™‚ç³»åˆ—ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã®ç”ŸæˆãŒå®Œäº†ã—ã¾ã—ãŸï¼")
            
            # è¨¼æ‹ é–“ã®é–¢é€£æ€§åˆ†æã‚’è¡¨ç¤º
            if use_ai:
                print("\nğŸ“Š è¨¼æ‹ é–“ã®é–¢é€£æ€§åˆ†æ:")
                relationships = builder.analyze_evidence_relationships(timeline_events)
                
                if relationships["temporal_clusters"]:
                    print("\nã€æ™‚é–“çš„ã«è¿‘æ¥ã—ãŸè¨¼æ‹ ã‚°ãƒ«ãƒ¼ãƒ—ã€‘")
                    for i, cluster in enumerate(relationships["temporal_clusters"], 1):
                        print(f"  {i}. {cluster['period']}: {cluster['evidence_count']}ä»¶")
                        print(f"     è¨¼æ‹ : {', '.join(cluster['evidence_numbers'])}")
                
                if relationships["chronological_gaps"]:
                    print("\nã€æ™‚ç³»åˆ—ä¸Šã®ã‚®ãƒ£ãƒƒãƒ—ã€‘")
                    for gap in relationships["chronological_gaps"]:
                        print(f"  - {gap['description']}")
            
        except Exception as e:
            print(f"\nâŒ ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {e}")
            import traceback
            traceback.print_exc()
    
    def manage_client_statements(self):
        """ä¾é ¼è€…ç™ºè¨€ãƒ»ãƒ¡ãƒ¢ã®ç®¡ç†"""
        from timeline_builder import TimelineBuilder
        
        print("\n" + "="*70)
        print("  ä¾é ¼è€…ç™ºè¨€ãƒ»ãƒ¡ãƒ¢ã®ç®¡ç†")
        print("="*70)
        print("\nä¾é ¼è€…ã‹ã‚‰ã®ãƒ’ã‚¢ãƒªãƒ³ã‚°å†…å®¹ã€ç™ºè¨€ã€ãƒ¡ãƒ¢ãªã©ã‚’è¨˜éŒ²ã—ã¾ã™ã€‚")
        print("ã“ã‚Œã‚‰ã¯æ™‚ç³»åˆ—ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ç”Ÿæˆæ™‚ã«è¨¼æ‹ ã¨çµ±åˆã•ã‚Œã¾ã™ã€‚")
        
        # TimelineBuilderã‚’åˆæœŸåŒ–
        builder = TimelineBuilder(self.case_manager, self.current_case, use_ai=False)
        
        while True:
            print("\n" + "-"*70)
            print("ã€ä¾é ¼è€…æƒ…å ±ç®¡ç†ãƒ¡ãƒ‹ãƒ¥ãƒ¼ã€‘")
            print("\nã€æ—¥ä»˜æŒ‡å®šã‚ã‚Šï¼ˆç‰¹å®šã®å‡ºæ¥äº‹ï¼‰ã€‘")
            print("  1. ä¾é ¼è€…ç™ºè¨€ã‚’è¿½åŠ ")
            print("  2. ç™»éŒ²æ¸ˆã¿ç™ºè¨€ã‚’è¡¨ç¤º")
            print("  3. ãƒ•ã‚¡ã‚¤ãƒ«ã‹ã‚‰ä¸€æ‹¬ã‚¤ãƒ³ãƒãƒ¼ãƒˆï¼ˆãƒ†ã‚­ã‚¹ãƒˆå½¢å¼ï¼‰")
            print("\nã€åŒ…æ‹¬çš„æƒ…å ±ï¼ˆæ—¥ä»˜ãªã—ã€è¤‡æ•°äº‹å®Ÿã«ã‚ãŸã‚‹ï¼‰ã€‘ğŸ†•")
            print("  4. åŒ…æ‹¬çš„ãªç™ºè¨€ãƒ»ãƒ¡ãƒ¢ã‚’è¿½åŠ ")
            print("  5. ç™»éŒ²æ¸ˆã¿åŒ…æ‹¬çš„ç™ºè¨€ã‚’è¡¨ç¤º")
            print("\n  0. ãƒ¡ã‚¤ãƒ³ãƒ¡ãƒ‹ãƒ¥ãƒ¼ã«æˆ»ã‚‹")
            print("-"*70)
            
            choice = input("\né¸æŠ (0-5): ").strip()
            
            if choice == '1':
                # ä¾é ¼è€…ç™ºè¨€ã‚’è¿½åŠ 
                print("\nã€ä¾é ¼è€…ç™ºè¨€ã®è¿½åŠ ã€‘")
                print("ç™ºè¨€ã¾ãŸã¯å‡ºæ¥äº‹ã«é–¢ã™ã‚‹æ—¥ä»˜ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„ï¼ˆYYYY-MM-DDå½¢å¼ï¼‰")
                print("ä¾‹: 2023-05-15")
                date = input("æ—¥ä»˜: ").strip()
                
                if not date:
                    print("âŒ æ—¥ä»˜ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„")
                    continue
                
                print("\nç™ºè¨€å†…å®¹ã¾ãŸã¯ãƒ¡ãƒ¢ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„ï¼ˆè¤‡æ•°è¡Œå¯ã€çµ‚äº†ã¯ç©ºè¡Œï¼‰")
                statement_lines = []
                while True:
                    line = input()
                    if not line:
                        break
                    statement_lines.append(line)
                
                statement = "\n".join(statement_lines)
                
                if not statement:
                    print("âŒ ç™ºè¨€å†…å®¹ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„")
                    continue
                
                print("\nçŠ¶æ³èª¬æ˜ã‚„èƒŒæ™¯æƒ…å ±ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„ï¼ˆçœç•¥å¯ã€çµ‚äº†ã¯ç©ºè¡Œï¼‰")
                context_lines = []
                while True:
                    line = input()
                    if not line:
                        break
                    context_lines.append(line)
                
                context = "\n".join(context_lines) if context_lines else None
                
                # ç™ºè¨€ã‚’è¿½åŠ 
                if builder.add_client_statement(date, statement, context):
                    print("\nâœ… ä¾é ¼è€…ç™ºè¨€ã‚’è¿½åŠ ã—ã¾ã—ãŸ")
                else:
                    print("\nâŒ ç™ºè¨€ã®è¿½åŠ ã«å¤±æ•—ã—ã¾ã—ãŸ")
            
            elif choice == '2':
                # ç™»éŒ²æ¸ˆã¿ç™ºè¨€ã‚’è¡¨ç¤º
                statements = builder._load_client_statements()
                
                if not statements:
                    print("\nç™»éŒ²ã•ã‚Œã¦ã„ã‚‹ç™ºè¨€ã¯ã‚ã‚Šã¾ã›ã‚“")
                    continue
                
                print(f"\nã€ç™»éŒ²æ¸ˆã¿ç™ºè¨€ä¸€è¦§ã€‘ï¼ˆ{len(statements)}ä»¶ï¼‰")
                print("-"*70)
                
                for i, stmt in enumerate(statements, 1):
                    print(f"\n{i}. ID: {stmt.get('statement_id', 'ä¸æ˜')}")
                    print(f"   æ—¥ä»˜: {stmt.get('date', 'ä¸æ˜')}")
                    print(f"   ç™ºè¨€: {stmt.get('statement', '')[:100]}...")
                    if stmt.get('context'):
                        print(f"   çŠ¶æ³: {stmt.get('context', '')[:100]}...")
                    print(f"   ç™»éŒ²æ—¥æ™‚: {stmt.get('added_at', 'ä¸æ˜')}")
            
            elif choice == '3':
                # ãƒ•ã‚¡ã‚¤ãƒ«ã‹ã‚‰ä¸€æ‹¬ã‚¤ãƒ³ãƒãƒ¼ãƒˆ
                print("\nã€ãƒ•ã‚¡ã‚¤ãƒ«ã‹ã‚‰ä¸€æ‹¬ã‚¤ãƒ³ãƒãƒ¼ãƒˆã€‘")
                print("ãƒ†ã‚­ã‚¹ãƒˆãƒ•ã‚¡ã‚¤ãƒ«ã®ãƒ‘ã‚¹ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„")
                print("ãƒ•ã‚©ãƒ¼ãƒãƒƒãƒˆä¾‹:")
                print("  [2023-05-15]")
                print("  ä¾é ¼è€…ã®ç™ºè¨€å†…å®¹")
                print("  è¤‡æ•°è¡Œå¯èƒ½")
                print("  ")
                print("  [2023-06-20]")
                print("  åˆ¥ã®ç™ºè¨€...")
                
                file_path = input("\nãƒ•ã‚¡ã‚¤ãƒ«ãƒ‘ã‚¹: ").strip()
                
                if not file_path or not os.path.exists(file_path):
                    print("âŒ ãƒ•ã‚¡ã‚¤ãƒ«ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“")
                    continue
                
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    # ç°¡å˜ãªãƒ‘ãƒ¼ã‚¹: [YYYY-MM-DD] ã§åŒºåˆ‡ã‚‹
                    import re
                    entries = re.split(r'\[(\d{4}-\d{2}-\d{2})\]', content)
                    
                    added_count = 0
                    for i in range(1, len(entries), 2):
                        if i + 1 < len(entries):
                            date = entries[i].strip()
                            statement = entries[i + 1].strip()
                            
                            if date and statement:
                                if builder.add_client_statement(date, statement, None):
                                    added_count += 1
                    
                    print(f"\nâœ… {added_count}ä»¶ã®ç™ºè¨€ã‚’è¿½åŠ ã—ã¾ã—ãŸ")
                    
                except Exception as e:
                    print(f"\nâŒ ã‚¤ãƒ³ãƒãƒ¼ãƒˆã«å¤±æ•—ã—ã¾ã—ãŸ: {e}")
            
            elif choice == '4':
                # åŒ…æ‹¬çš„ãªç™ºè¨€ãƒ»ãƒ¡ãƒ¢ã‚’è¿½åŠ 
                print("\nã€åŒ…æ‹¬çš„ãªç™ºè¨€ãƒ»ãƒ¡ãƒ¢ã®è¿½åŠ ã€‘")
                print("â€» æ—¥ä»˜æŒ‡å®šãªã—ã€è¤‡æ•°ã®äº‹å®Ÿã«ã‚ãŸã‚‹å…¨ä½“çš„ãªæƒ…å ±ã‚’è¨˜éŒ²ã—ã¾ã™")
                print("â€» AIãŒæ™‚ç³»åˆ—ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ç”Ÿæˆæ™‚ã«å…¨ä½“ã®æ–‡è„ˆã¨ã—ã¦æ´»ç”¨ã—ã¾ã™\n")
                
                print("ã‚¿ã‚¤ãƒˆãƒ«ã¾ãŸã¯æ¦‚è¦ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„")
                print("ä¾‹: äº‹ä»¶ã®å…¨ä½“çš„ãªçµŒç·¯ã€å½“äº‹è€…é–“ã®é–¢ä¿‚æ€§ã€èƒŒæ™¯æƒ…å ±")
                title = input("ã‚¿ã‚¤ãƒˆãƒ«: ").strip()
                
                if not title:
                    print("âŒ ã‚¿ã‚¤ãƒˆãƒ«ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„")
                    continue
                
                print("\nã‚«ãƒ†ã‚´ãƒªã‚’é¸æŠã—ã¦ãã ã•ã„")
                print("  1. äº‹ä»¶ã®èƒŒæ™¯")
                print("  2. äººç‰©é–¢ä¿‚")
                print("  3. å…¨ä½“çš„ãªçµŒç·¯")
                print("  4. ãã®ä»–")
                cat_choice = input("é¸æŠ (1-4): ").strip()
                
                category_map = {
                    '1': 'äº‹ä»¶ã®èƒŒæ™¯',
                    '2': 'äººç‰©é–¢ä¿‚',
                    '3': 'å…¨ä½“çš„ãªçµŒç·¯',
                    '4': 'ãã®ä»–'
                }
                category = category_map.get(cat_choice, 'ãã®ä»–')
                
                print(f"\n{title} ã®è©³ç´°å†…å®¹ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„ï¼ˆè¤‡æ•°è¡Œå¯ã€çµ‚äº†ã¯ç©ºè¡Œï¼‰")
                content_lines = []
                while True:
                    line = input()
                    if not line:
                        break
                    content_lines.append(line)
                
                content = "\n".join(content_lines)
                
                if not content:
                    print("âŒ å†…å®¹ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„")
                    continue
                
                # åŒ…æ‹¬çš„ç™ºè¨€ã‚’è¿½åŠ 
                if builder.add_general_context(title, content, category):
                    print("\nâœ… åŒ…æ‹¬çš„ãªç™ºè¨€ãƒ»ãƒ¡ãƒ¢ã‚’è¿½åŠ ã—ã¾ã—ãŸ")
                    print("   æ™‚ç³»åˆ—ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ç”Ÿæˆæ™‚ã«AIãŒè‡ªå‹•çš„ã«è€ƒæ…®ã—ã¾ã™")
                else:
                    print("\nâŒ è¿½åŠ ã«å¤±æ•—ã—ã¾ã—ãŸ")
            
            elif choice == '5':
                # ç™»éŒ²æ¸ˆã¿åŒ…æ‹¬çš„ç™ºè¨€ã‚’è¡¨ç¤º
                contexts = builder._load_general_context()
                
                if not contexts:
                    print("\nç™»éŒ²ã•ã‚Œã¦ã„ã‚‹åŒ…æ‹¬çš„ç™ºè¨€ã¯ã‚ã‚Šã¾ã›ã‚“")
                    continue
                
                print(f"\nã€ç™»éŒ²æ¸ˆã¿åŒ…æ‹¬çš„ç™ºè¨€ä¸€è¦§ã€‘ï¼ˆ{len(contexts)}ä»¶ï¼‰")
                print("-"*70)
                
                for i, ctx in enumerate(contexts, 1):
                    print(f"\n{i}. ID: {ctx.get('context_id', 'ä¸æ˜')}")
                    print(f"   ã‚¿ã‚¤ãƒˆãƒ«: {ctx.get('title', 'ä¸æ˜')}")
                    print(f"   ã‚«ãƒ†ã‚´ãƒª: {ctx.get('category', 'ä¸æ˜')}")
                    print(f"   å†…å®¹: {ctx.get('content', '')[:150]}...")
                    print(f"   ç™»éŒ²æ—¥æ™‚: {ctx.get('added_at', 'ä¸æ˜')}")
            
            elif choice == '0':
                # ãƒ¡ã‚¤ãƒ³ãƒ¡ãƒ‹ãƒ¥ãƒ¼ã«æˆ»ã‚‹
                break
            
            else:
                print("\nâŒ ç„¡åŠ¹ãªé¸æŠã§ã™")
    
    def run(self):
        """ãƒ¡ã‚¤ãƒ³å®Ÿè¡Œãƒ«ãƒ¼ãƒ—"""
        # æœ€åˆã«äº‹ä»¶ã‚’é¸æŠ
        if not self.select_case():
            print("\nâŒ äº‹ä»¶ãŒé¸æŠã•ã‚Œã¦ã„ãªã„ãŸã‚çµ‚äº†ã—ã¾ã™")
            return
        
        # ãƒ¡ã‚¤ãƒ³ãƒ«ãƒ¼ãƒ—
        while True:
            self.display_main_menu()
            choice = input("\né¸æŠ (0-9): ").strip()
            
            if choice == '1':
                # è¨¼æ‹ æ•´ç†ï¼ˆæœªåˆ†é¡ãƒ•ã‚©ãƒ«ãƒ€ã‹ã‚‰æ•´ç†æ¸ˆã¿_æœªç¢ºå®šã¸ï¼‰
                try:
                    # è¨¼æ‹ ç¨®åˆ¥ã‚’é¸æŠ
                    evidence_type = self.select_evidence_type()
                    if evidence_type:
                        organizer = EvidenceOrganizer(self.case_manager, self.current_case)
                        organizer.interactive_organize(evidence_type=evidence_type)
                except Exception as e:
                    print(f"\nâŒ ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {str(e)}")
                    import traceback
                    traceback.print_exc()
                        
            elif choice == '2':
                # è¨¼æ‹ åˆ†æï¼ˆç•ªå·æŒ‡å®šãƒ»ç¯„å›²æŒ‡å®šã«å¯¾å¿œï¼‰
                # è¨¼æ‹ ç¨®åˆ¥ã‚’é¸æŠ
                evidence_type = self.select_evidence_type()
                if evidence_type:
                    evidence_numbers = self.get_evidence_number_input(evidence_type)
                    if evidence_numbers:
                        # è¤‡æ•°ä»¶ã®å ´åˆã¯ç¢ºèª
                        if len(evidence_numbers) > 1:
                            type_name = "ç”²å·è¨¼" if evidence_type == 'ko' else "ä¹™å·è¨¼"
                            print(f"\nå‡¦ç†å¯¾è±¡ [{type_name}]: {', '.join(evidence_numbers)}")
                            confirm = input("å‡¦ç†ã‚’é–‹å§‹ã—ã¾ã™ã‹ï¼Ÿ (y/n): ").strip().lower()
                            if confirm != 'y':
                                continue
                        
                        # åˆ†æå®Ÿè¡Œ
                        for evidence_number in evidence_numbers:
                            gdrive_file_info = self._get_gdrive_info_from_database(evidence_number, evidence_type)
                            self.process_evidence(evidence_number, gdrive_file_info, evidence_type)
                        
            elif choice == '3':
                # AIå¯¾è©±å½¢å¼ã§åˆ†æå†…å®¹ã‚’æ”¹å–„
                try:
                    # è¨¼æ‹ ç¨®åˆ¥ã‚’é¸æŠ
                    evidence_type = self.select_evidence_type()
                    if evidence_type:
                        self.edit_evidence_with_ai(evidence_type)
                except Exception as e:
                    print(f"\nã‚¨ãƒ©ãƒ¼: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    
            elif choice == '4':
                # æ—¥ä»˜é †ã«ä¸¦ã³æ›¿ãˆã¦ç¢ºå®šï¼ˆæ•´ç†æ¸ˆã¿_æœªç¢ºå®š â†’ ç”²å·è¨¼/ä¹™å·è¨¼ï¼‰
                try:
                    # è¨¼æ‹ ç¨®åˆ¥ã‚’é¸æŠ
                    evidence_type = self.select_evidence_type()
                    if evidence_type:
                        self.analyze_and_sort_pending_evidence(evidence_type)
                except Exception as e:
                    print(f"\nã‚¨ãƒ©ãƒ¼: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    
            elif choice == '5':
                # è¨¼æ‹ åˆ†æä¸€è¦§ã‚’è¡¨ç¤º
                try:
                    # è¨¼æ‹ ç¨®åˆ¥ã‚’é¸æŠ
                    evidence_type = self.select_evidence_type()
                    if evidence_type:
                        self.show_evidence_list(evidence_type)
                except Exception as e:
                    print(f"\nã‚¨ãƒ©ãƒ¼: {str(e)}")
                    import traceback
                    traceback.print_exc()
            
            elif choice == '6':
                # è¨¼æ‹ ä¸€è¦§ã‚’ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆ
                try:
                    # è¨¼æ‹ ç¨®åˆ¥ã‚’é¸æŠ
                    evidence_type = self.select_evidence_type()
                    if evidence_type:
                        self.export_evidence_list(evidence_type)
                except Exception as e:
                    print(f"\nã‚¨ãƒ©ãƒ¼: {str(e)}")
                    import traceback
                    traceback.print_exc()
            
            elif choice == '7':
                # CSVç·¨é›†å†…å®¹ã‚’database.jsonã«åæ˜ 
                try:
                    self.import_csv_updates()
                except Exception as e:
                    print(f"\nã‚¨ãƒ©ãƒ¼: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    
            elif choice == '8':
                # æ™‚ç³»åˆ—ã‚¹ãƒˆãƒ¼ãƒªãƒ¼ã®ç”Ÿæˆ
                try:
                    self.generate_timeline_story()
                except Exception as e:
                    print(f"\nã‚¨ãƒ©ãƒ¼: {str(e)}")
                    import traceback
                    traceback.print_exc()
                
            elif choice == '9':
                # ä¾é ¼è€…ç™ºè¨€ãƒ»ãƒ¡ãƒ¢ã®ç®¡ç†
                try:
                    self.manage_client_statements()
                except Exception as e:
                    print(f"\nã‚¨ãƒ©ãƒ¼: {str(e)}")
                    import traceback
                    traceback.print_exc()
                
            elif choice == '10':
                # database.jsonã®çŠ¶æ…‹ç¢ºèª
                self.show_database_status()
                
            elif choice == '11':
                # äº‹ä»¶ã‚’åˆ‡ã‚Šæ›¿ãˆ
                if self.select_case():
                    print("\nâœ… äº‹ä»¶ã‚’åˆ‡ã‚Šæ›¿ãˆã¾ã—ãŸ")
                    
            elif choice == '0':
                # çµ‚äº†
                print("\nPhase1_Evidence Analysis Systemã‚’çµ‚äº†ã—ã¾ã™")
                break
                
            else:
                print("\nã‚¨ãƒ©ãƒ¼: ç„¡åŠ¹ãªé¸æŠã§ã™ã€‚0-11ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„ã€‚")
            
            input("\nEnterã‚­ãƒ¼ã‚’æŠ¼ã—ã¦ç¶šè¡Œ...")


def main():
    """ãƒ¡ã‚¤ãƒ³é–¢æ•°"""
    print("\n" + "="*70)
    print("  Phase1_Evidence Analysis Systemï¼ˆãƒãƒ«ãƒäº‹ä»¶å¯¾å¿œç‰ˆï¼‰èµ·å‹•ä¸­...")
    print("="*70)
    
    # ç’°å¢ƒãƒã‚§ãƒƒã‚¯
    if not os.getenv('OPENAI_API_KEY'):
        print("\nâŒ ã‚¨ãƒ©ãƒ¼: OPENAI_API_KEYãŒè¨­å®šã•ã‚Œã¦ã„ã¾ã›ã‚“")
        print("\nè¨­å®šæ–¹æ³•:")
        print("  export OPENAI_API_KEY='sk-your-api-key'")
        print("\nã¾ãŸã¯ .env ãƒ•ã‚¡ã‚¤ãƒ«ã«è¨˜è¼‰:")
        print("  OPENAI_API_KEY=sk-your-api-key")
        return
    
    # Googleèªè¨¼ãƒã‚§ãƒƒã‚¯
    if not os.path.exists('credentials.json'):
        print("\nâš ï¸ è­¦å‘Š: credentials.jsonãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“")
        print("Google Drive APIæ©Ÿèƒ½ãŒä½¿ç”¨ã§ãã¾ã›ã‚“")
        print("\nGoogle Cloud Consoleã‹ã‚‰ credentials.json ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ã—ã¦ãã ã•ã„")
        response = input("\nç¶šè¡Œã—ã¾ã™ã‹ï¼Ÿ (y/n): ").strip().lower()
        if response != 'y':
            return
    
    # global_config.py ãƒã‚§ãƒƒã‚¯
    if not hasattr(gconfig, 'SHARED_DRIVE_ROOT_ID') or not gconfig.SHARED_DRIVE_ROOT_ID:
        print("\nâŒ ã‚¨ãƒ©ãƒ¼: global_config.py ã§ SHARED_DRIVE_ROOT_ID ãŒè¨­å®šã•ã‚Œã¦ã„ã¾ã›ã‚“")
        print("\nglobal_config.py ã‚’é–‹ã„ã¦ã€å¤§å…ƒã®å…±æœ‰ãƒ‰ãƒ©ã‚¤ãƒ–IDã‚’è¨­å®šã—ã¦ãã ã•ã„:")
        print("  SHARED_DRIVE_ROOT_ID = 'your-shared-drive-id'")
        return
    
    print(f"\nâœ… å…±æœ‰ãƒ‰ãƒ©ã‚¤ãƒ–ID: {gconfig.SHARED_DRIVE_ROOT_ID}")
    
    # å®Ÿè¡Œ
    runner = Phase1MultiRunner()
    runner.run()


if __name__ == "__main__":
    main()
